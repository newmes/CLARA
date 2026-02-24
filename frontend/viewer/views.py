"""
Views for Clinical Trial Simulation Viewer.

Generative Agents 스타일의 day-by-day viewer +
Concordia 스타일의 SSE 실시간 업데이트 +
Interactive Game Mode (Care Agent 대신 사람이 참여).
"""
import json
import logging
import os
import time
from pathlib import Path
from urllib.error import URLError
from urllib.request import Request, urlopen

from django.conf import settings
from django.http import HttpResponse, JsonResponse, StreamingHttpResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_GET

# ─── Data helpers ─────────────────────────────────────────────

DATA_DIR = settings.DATA_DIR
MAP_ASSETS_DIR = Path(settings.BASE_DIR) / "static_dirs" / "assets" / "map"


def _get_runs():
    """Available simulation runs, newest first."""
    runs_dir = DATA_DIR / "runs"
    if not runs_dir.exists():
        return []
    runs = []
    for d in sorted(runs_dir.iterdir(), reverse=True):
        if d.is_dir() and (d / "simulations").exists():
            modes = []
            natural_files = [f for f in (d / "simulations").glob("*_natural.jsonl")
                           if "_hospital" not in f.stem]
            care_ai_files = [f for f in (d / "simulations").glob("*_care_ai.jsonl")
                           if "_hospital" not in f.stem]
            if natural_files:
                modes.append("natural")
            if care_ai_files:
                modes.append("care_ai")

            run_info = {
                "id": d.name,
                "path": str(d),
                "modes": modes,
                "status": "completed",
            }

            # Check for run_meta.json
            meta_path = d / "run_meta.json"
            if meta_path.exists():
                try:
                    meta = json.loads(meta_path.read_text(encoding="utf-8"))
                    run_info["drug_name"] = meta.get("drug_name", "")
                    run_info["indication"] = meta.get("indication", "")
                    run_info["n_patients"] = meta.get("n_patients")
                    run_info["total_days"] = meta.get("total_days")
                    run_info["status"] = meta.get("status", "completed")
                    run_info["started_at"] = meta.get("started_at")
                except Exception:
                    pass

            # Count patients
            patients_dir = d / "patients"
            if patients_dir.exists() and "n_patients" not in run_info:
                run_info["n_patients"] = len(list(patients_dir.glob("*.json")))

            runs.append(run_info)

    # Also include runs that are still generating (simulations/ might not exist yet)
    for d in sorted(runs_dir.iterdir(), reverse=True):
        if d.is_dir() and not (d / "simulations").exists():
            meta_path = d / "run_meta.json"
            if meta_path.exists():
                try:
                    meta = json.loads(meta_path.read_text(encoding="utf-8"))
                    if meta.get("status") in ("running", "generating_patients",
                                               "starting"):
                        runs.insert(0, {
                            "id": d.name,
                            "path": str(d),
                            "modes": [],
                            "status": meta.get("status", "running"),
                            "drug_name": meta.get("drug_name", ""),
                            "n_patients": meta.get("n_patients"),
                            "total_days": meta.get("total_days"),
                        })
                except Exception:
                    pass

    return runs


def _get_run_path(run_id: str) -> Path:
    return DATA_DIR / "runs" / run_id


def _load_patient_profile(run_path: Path, patient_id: str) -> dict:
    """Load patient JSON profile (demographics, persona, etc.)."""
    f = run_path / "patients" / f"{patient_id}.json"
    if f.exists():
        with open(f) as fh:
            return json.load(fh)
    return {}


def _load_run_meta(run_path: Path) -> dict:
    """Load run_meta.json for this run."""
    f = run_path / "run_meta.json"
    if f.exists():
        try:
            with open(f) as fh:
                return json.load(fh)
        except Exception:
            pass
    return {}


def _load_rule_set(run_path: Path) -> dict:
    """Load rule_set.json for this run."""
    f = run_path / "rule_set.json"
    if f.exists():
        with open(f) as fh:
            return json.load(fh)
    return {}


def _extract_lab_ranges(run_path: Path, mode: str = "natural") -> dict:
    """Extract lab reference ranges from LB data (LBORNRLO/LBORNRHI).

    Used when rule_set.json is missing or has no lab_reference_ranges.
    Reads the first patient's first day with LB data and builds a ranges dict.
    """
    from frontend.viewer.crf_aggregator import LAB_ABBREVIATIONS, _lab_display_name
    sim_dir = run_path / "simulations"
    if not sim_dir.exists():
        return {}
    # Find any JSONL file to extract ranges from
    for fpath in sorted(sim_dir.glob(f"*_{mode}.jsonl")):
        if "_hospital" in fpath.stem:
            continue
        with open(fpath, encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                record = json.loads(line)
                lb = record.get("LB")
                if not lb or not lb.get("LBPERF"):
                    continue
                results = lb.get("results", {})
                ranges = {}
                for test_name, vals in results.items():
                    lo = vals.get("LBORNRLO")
                    hi = vals.get("LBORNRHI")
                    if lo is not None or hi is not None:
                        display = _lab_display_name(test_name)
                        ranges[display] = {
                            "unit": vals.get("LBORRESU", ""),
                            "normal_range": {"min": lo, "max": hi},
                            "LLN": lo,
                            "ULN": hi,
                        }
                if ranges:
                    return ranges
    return {}


def _list_patients(run_path: Path) -> list[str]:
    """List patient IDs from simulation files (exclude _hospital variants).
    Falls back to patients/ directory if no simulation files exist yet."""
    ids = set()
    sim_dir = run_path / "simulations"
    if sim_dir.exists():
        for f in sim_dir.glob("*_natural.jsonl"):
            if "_hospital" in f.stem:
                continue
            pid = f.stem.replace("_natural", "")
            ids.add(pid)
        for f in sim_dir.glob("*_care_ai.jsonl"):
            if "_hospital" in f.stem:
                continue
            pid = f.stem.replace("_care_ai", "")
            ids.add(pid)
    # Fallback: count from patients/ directory
    if not ids:
        patients_dir = run_path / "patients"
        if patients_dir.exists():
            for f in patients_dir.glob("*.json"):
                ids.add(f.stem)
    return sorted(ids)


def _load_day_for_patient(run_path: Path, patient_id: str, day: int,
                          mode: str = "natural") -> dict | None:
    """Load a single day's data for a patient from JSONL.

    If the requested day is beyond the last entry AND the patient is
    deceased on the last day, return the death-day record so that
    downstream code still sees location='DECEASED'.
    """
    f = run_path / "simulations" / f"{patient_id}_{mode}.jsonl"
    if not f.exists():
        return None
    last_record = None
    with open(f) as fh:
        for line in fh:
            record = json.loads(line)
            if record.get("day") == day:
                return record
            last_record = record
    # Day not found — check if patient died before this day
    if last_record and last_record.get("day", 0) < day:
        loc = (last_record.get("objective") or {}).get("location", "")
        if loc == "DECEASED":
            return last_record
    return None


def _load_all_days_for_patient(run_path: Path, patient_id: str,
                               mode: str = "natural") -> list[dict]:
    """Load all days for a patient."""
    f = run_path / "simulations" / f"{patient_id}_{mode}.jsonl"
    if not f.exists():
        return []
    days = []
    with open(f) as fh:
        for line in fh:
            days.append(json.loads(line))
    return days


def _find_last_hr_observation(run_path: Path, patient_id: str, day: int,
                              mode: str = "natural") -> tuple[dict, int]:
    """Find the last day with non-empty hospital_record labs/vitals.

    Returns (hr_objective, stale_days). If none found, returns ({}, 0).
    Reads the JSONL backwards from the target day.
    """
    f = run_path / "simulations" / f"{patient_id}_{mode}.jsonl"
    if not f.exists():
        return {}, 0
    candidates = []
    with open(f) as fh:
        for line in fh:
            record = json.loads(line)
            d_num = record.get("day", 0)
            if d_num >= day:
                break
            hr = record.get("hospital_record", {})
            hr_obj = hr.get("objective", {})
            if hr_obj.get("labs") or hr_obj.get("vitals"):
                candidates.append((d_num, hr_obj))
    if candidates:
        last_day_num, last_hr_obj = candidates[-1]
        return last_hr_obj, day - last_day_num
    return {}, 0


def _count_days(run_path: Path, mode: str | None = None) -> int:
    """Find max day across all patients.
    
    If mode is specified, only count that mode's files.
    Otherwise, count across all modes.
    """
    sim_dir = run_path / "simulations"
    max_day = 0
    patterns = []
    if mode:
        patterns.append(f"*_{mode}.jsonl")
    else:
        patterns.extend(["*_natural.jsonl", "*_care_ai.jsonl"])
    for pattern in patterns:
        for f in sim_dir.glob(pattern):
            with open(f) as fh:
                for line in fh:
                    d = json.loads(line).get("day", 0)
                    if d > max_day:
                        max_day = d
    return max_day


def _is_hr_tumor_stuck(all_days: list[dict], current_day: int) -> bool:
    """Detect if HR tumor data is stuck (never updated from baseline).

    Legacy runs generated before the RECIST fix have HR tumor frozen at
    the Day 1 value. We detect this by checking if HR tumor is identical
    across multiple visit days.
    """
    tumor_vals = set()
    visit_count = 0
    for d in all_days:
        d_num = d.get("day", 0)
        if d_num > current_day:
            break
        hr = d.get("hospital_record", {})
        obs_types = hr.get("observation_types", [])
        hr_obj = hr.get("objective", {})
        hr_t = hr_obj.get("tumor")
        if obs_types and hr_t:
            pct = hr_t.get("estimated_change_pct")
            if pct is not None:
                tumor_vals.add(round(pct, 2))
                visit_count += 1
    # If we've had 3+ visit days with tumor data and they all have the same value,
    # it's almost certainly stuck
    return visit_count >= 3 and len(tumor_vals) <= 1


def _extract_day_events(day_data: dict) -> list[dict]:
    """Extract notable events from a day's data for the summary panel."""
    events = []
    pid = day_data.get("patient_id", "?")

    # AE events
    for ae in day_data.get("AE", []):
        ae_term = ae.get("AETERM", "unknown")
        grade = ae.get("_grade", "?")
        status = ae.get("_status", "")
        days_active = ae.get("_days_active", 0)

        if days_active <= 1:
            events.append({
                "type": "ae_onset",
                "severity": "high" if grade >= 3 else "medium",
                "icon": "🔴" if grade >= 3 else "🟡",
                "text": f"{pid}: {ae_term} Grade {grade} onset",
            })
        elif "worsened" in status:
            events.append({
                "type": "ae_worsened",
                "severity": "high" if grade >= 3 else "medium",
                "icon": "🔴" if grade >= 3 else "🟡",
                "text": f"{pid}: {ae_term} worsened to Grade {grade}",
            })

    # Resolved AEs (checking _status)
    for ae in day_data.get("AE", []):
        if ae.get("_status") == "resolved":
            events.append({
                "type": "ae_resolved",
                "severity": "low",
                "icon": "🟢",
                "text": f"{pid}: {ae.get('AETERM', '?')} resolved",
            })

    # Dose modifications
    for ec in day_data.get("EC", []):
        if ec.get("ECDOSADJ"):
            drug = ec.get("ECREFID", "?")
            adj = ec.get("ECADJ", "modified")
            events.append({
                "type": "dose_mod",
                "severity": "medium",
                "icon": "💊",
                "text": f"{pid}: {drug} {adj}",
            })

    # Treatment administration
    for ec in day_data.get("EC", []):
        if ec.get("ECTRTCMP") and not ec.get("ECDOSADJ"):
            drug = ec.get("ECREFID", "?")
            events.append({
                "type": "treatment",
                "severity": "info",
                "icon": "💉",
                "text": f"{pid}: {drug} administered",
            })

    # RECIST scan
    rs = day_data.get("RS")
    if rs:
        if isinstance(rs, list):
            for r in rs:
                events.append({
                    "type": "recist",
                    "severity": "info",
                    "icon": "📋",
                    "text": f"{pid}: RECIST scan — {r.get('RSORRESU', '?')}",
                })
        elif isinstance(rs, dict):
            events.append({
                "type": "recist",
                "severity": "info",
                "icon": "📋",
                "text": f"{pid}: RECIST scan — {rs.get('RSORRESU', '?')}",
            })

    # Discontinuation
    ds = day_data.get("DS")
    if ds:
        events.append({
            "type": "discontinuation",
            "severity": "high",
            "icon": "⛔",
            "text": f"{pid}: Discontinued — {ds.get('DSDECOD', '?')}",
        })

    # Video call / Care record
    for cr in day_data.get("care_record", []):
        assessment = cr.get("nurse_assessment", {})
        severity_level = assessment.get("severity_level", "green")
        summary_text = assessment.get("summary", "Care AI interaction")
        sev_map = {"green": "info", "yellow": "info", "orange": "medium", "red": "high"}
        icon_map = {"green": "📹", "yellow": "📹", "orange": "🟠", "red": "🔴"}
        events.append({
            "type": "video_call",
            "severity": sev_map.get(severity_level, "info"),
            "icon": icon_map.get(severity_level, "📹"),
            "text": f"{pid}: Video call [{severity_level.upper()}] — {summary_text}",
        })
        for action in cr.get("actions", []):
            act = action.get("action", "")
            if act not in ("no_action", "monitor_closely"):
                events.append({
                    "type": "care_action",
                    "severity": "medium",
                    "icon": "🩺",
                    "text": f"{pid}: Care AI → {act}: {action.get('reason', '')}",
                })

    # Observation events
    for obs in day_data.get("observation_events", []):
        obs_type = obs.get("type", "")
        if obs_type == "self_report":
            events.append({
                "type": "self_report", "severity": "info", "icon": "📞",
                "text": f"{pid}: Self-reported symptoms to clinic",
            })
        elif obs_type == "er_visit":
            events.append({
                "type": "er_visit", "severity": "high", "icon": "🚑",
                "text": f"{pid}: Emergency room visit",
            })

    return events


def _patient_summary(profile: dict, day_data: dict | None,
                     view_mode: str = "gt",
                     run_path: Path = None, mode: str = "natural") -> dict:
    """Build a summary dict for a patient card.

    Args:
        view_mode: "gt" for Ground Truth, "hr" for Hospital Record
        run_path: needed for HR carry-forward when hospital_record is empty
        mode: simulation mode for HR carry-forward
    """
    dm = profile.get("DM", {})
    persona = profile.get("persona", {})

    summary = {
        "patient_id": profile.get("patient_id", "?"),
        "age": dm.get("AGE", "?"),
        "sex": dm.get("SEX", "?"),
        "race": dm.get("RACE", ""),
        "persona_type": persona.get("type", "unknown"),
        "persona_desc": persona.get("description", ""),
        "view_mode": view_mode,
    }

    if day_data:
        hr_data = day_data.get("hospital_record", {})
        hr_obj = hr_data.get("objective", {})
        gt_obj = day_data.get("objective", {})
        obs_types = hr_data.get("observation_types", [])
        is_visit = ("scheduled_visit" in obs_types or "er_visit" in obs_types)

        if view_mode == "hr":
            # ── Hospital Record mode: ONLY what the hospital knows ──
            # Carry-forward: if current HR has no labs/vitals, find last known
            if not hr_obj.get("labs") and not hr_obj.get("vitals") and run_path:
                pid = profile.get("patient_id", "?")
                cur_day = day_data.get("day", 0)
                last_hr_obj, stale = _find_last_hr_observation(
                    run_path, pid, cur_day, mode)
                if last_hr_obj:
                    hr_obj = dict(hr_obj) if hr_obj else {}
                    if not hr_obj.get("labs"):
                        hr_obj["labs"] = last_hr_obj.get("labs", {})
                        hr_obj["labs_stale_days"] = stale
                    if not hr_obj.get("vitals"):
                        hr_obj["vitals"] = last_hr_obj.get("vitals", {})
                        hr_obj["vitals_stale_days"] = stale
                    for k in ("active_aes", "treatment_status", "ecog",
                              "location", "tumor"):
                        if k not in hr_obj and k in last_hr_obj:
                            hr_obj[k] = last_hr_obj[k]

            summary["location"] = hr_obj.get("location", gt_obj.get("location", "?"))
            summary["treatment_status"] = hr_obj.get(
                "treatment_status", gt_obj.get("treatment_status", "?"))
            summary["ecog"] = hr_obj.get("ecog", "?")
            tumor = hr_obj.get("tumor") or {}
            summary["tumor_change_pct"] = tumor.get("estimated_change_pct")
            summary["is_visit_day"] = is_visit
            summary["observation_types"] = obs_types

            # AEs: only detected
            active_aes = []
            for ae in hr_obj.get("active_aes", []):
                active_aes.append({
                    "term": ae.get("ae", "?"),
                    "grade": ae.get("grade", "?"),
                    "days_active": ae.get("days_active", 0),
                    "detected_day": ae.get("detected_day"),
                    "detection_delay": ae.get("detection_delay"),
                    "channel": ae.get("channel", ""),
                })
            summary["active_aes"] = active_aes

            # Labs: from HR (stale on non-visit days)
            hr_labs = hr_obj.get("labs", {})
            summary["labs"] = {}
            for name, info in hr_labs.items():
                if isinstance(info, dict):
                    summary["labs"][name] = info.get("value")
                else:
                    summary["labs"][name] = info
            summary["labs_stale_days"] = hr_obj.get("labs_stale_days", 0)

            # Vitals: from HR (stale on non-visit days)
            hr_vitals = hr_obj.get("vitals", {})
            summary["vitals"] = {
                "temp": hr_vitals.get("BT"),
                "bp": f"{hr_vitals.get('SBP', '?')}/{hr_vitals.get('DBP', '?')}",
                "hr": hr_vitals.get("HR"),
                "spo2": hr_vitals.get("SpO2"),
                "rr": hr_vitals.get("RR"),
                "weight": hr_vitals.get("weight_kg"),
            }
            summary["vitals_stale_days"] = hr_obj.get("vitals_stale_days", 0)

            # Subjective: only from HR (empty on non-visit days)
            hr_subj = hr_data.get("subjective", {})
            if hr_subj:
                summary["awareness"] = hr_subj.get("overall_awareness", "?")
                summary["symptoms_perceived"] = hr_subj.get(
                    "symptoms_patient_perceives", [])
            else:
                summary["awareness"] = "UNKNOWN"
                summary["symptoms_perceived"] = []

        else:
            # ── Ground Truth mode: full picture ──
            summary["location"] = gt_obj.get("location", "?")
            summary["treatment_status"] = gt_obj.get("treatment_status", "?")
            summary["ecog"] = gt_obj.get("ecog", "?")
            gt_tumor = gt_obj.get("tumor") or {}
            summary["tumor_change_pct"] = gt_tumor.get(
                "estimated_change_pct", 0)

            active_aes = []
            for ae in day_data.get("AE", []):
                if ae.get("AEONGO") or ae.get("_status", "").startswith("active"):
                    active_aes.append({
                        "term": ae.get("AETERM", "?"),
                        "grade": ae.get("_grade", "?"),
                        "days_active": ae.get("_days_active", 0),
                    })
            summary["active_aes"] = active_aes

            # All labs from GT
            lb = day_data.get("LB", {}).get("results", {})
            summary["labs"] = {}
            for name, info in lb.items():
                if isinstance(info, dict):
                    summary["labs"][name] = info.get("LBORRES")
                else:
                    summary["labs"][name] = info

            # Vitals from GT
            vs = day_data.get("VS", {})
            summary["vitals"] = {
                "temp": vs.get("TEMP_VSORRES"),
                "bp": f"{vs.get('SYSBP_VSORRES', '?')}/{vs.get('DIABP_VSORRES', '?')}",
                "hr": vs.get("PULSE_VSORRES"),
                "spo2": vs.get("_SpO2") or vs.get("OXYSAT_VSORRES"),
                "rr": vs.get("RESP_VSORRES"),
                "weight": vs.get("WEIGHT_VSORRES"),
            }

            # Subjective from GT
            subj = day_data.get("subjective", {})
            summary["awareness"] = subj.get("overall_awareness", "?")
            summary["symptoms_perceived"] = subj.get(
                "symptoms_patient_perceives", [])

        # Sim metadata
        sim = day_data.get("_sim", {})
        summary["generation_mode"] = sim.get("generation_mode", "?")
        summary["mortality_risk"] = sim.get("mortality_risk", 0)

        # ── New: Mood state ──
        summary["mood"] = day_data.get("mood_state", {})

        # ── New: Hospital record (what the hospital knows) ──
        raw_hr = day_data.get("hospital_record", {})
        raw_hr_obj = raw_hr.get("objective", {})
        # Carry-forward: if HR has no labs/vitals, look back
        if not raw_hr_obj.get("labs") and not raw_hr_obj.get("vitals") and run_path:
            pid_cf = profile.get("patient_id", "?")
            cur_day_cf = day_data.get("day", 0)
            cf_hr_obj, cf_stale = _find_last_hr_observation(
                run_path, pid_cf, cur_day_cf, mode)
            if cf_hr_obj:
                merged_hr = dict(raw_hr)
                merged_obj = dict(raw_hr_obj) if raw_hr_obj else {}
                if not merged_obj.get("labs"):
                    merged_obj["labs"] = cf_hr_obj.get("labs", {})
                    merged_obj["labs_stale_days"] = cf_stale
                if not merged_obj.get("vitals"):
                    merged_obj["vitals"] = cf_hr_obj.get("vitals", {})
                    merged_obj["vitals_stale_days"] = cf_stale
                for k in ("active_aes", "treatment_status", "ecog",
                          "location", "tumor"):
                    if k not in merged_obj and k in cf_hr_obj:
                        merged_obj[k] = cf_hr_obj[k]
                merged_hr["objective"] = merged_obj
                raw_hr = merged_hr
        summary["hospital_record"] = raw_hr

        # ── New: Observation events ──
        summary["observation_events"] = day_data.get("observation_events", [])

        # ── New: Care AI record ──
        care_records = day_data.get("care_record", [])
        if care_records:
            cr = care_records[0] if isinstance(care_records, list) else care_records
            summary["care_record"] = {
                "severity_level": cr.get("nurse_assessment", {}).get("severity_level", ""),
                "summary": cr.get("nurse_assessment", {}).get("summary", ""),
                "actions": cr.get("actions", []),
                "detection": cr.get("detection", {}),
                "turns": cr.get("turns", []),
                "terminated_early": cr.get("terminated_early", False),
                "mood_snapshot": cr.get("mood_snapshot", {}),
                "interaction_quality": cr.get("interaction_quality", {}),
                "grade_distortion": cr.get("grade_distortion", 0),
            }
        else:
            summary["care_record"] = None

    return summary


# ─── Page views ───────────────────────────────────────────────

def landing(request):
    """Landing page: pure technology showcase + demo map preview."""
    # Fixed to the 100-patient Etoposide + Cisplatin run for a visually rich map
    demo_run_id = "20260224_061414_Etoposide___Cisplatin_100pt_126d"
    # Verify the run exists, fall back to dynamic selection if not
    run_path = _get_run_path(demo_run_id)
    if not run_path.exists() or not (run_path / "simulations").exists():
        runs = _get_runs()
        demo_run_id = ""
        for r in runs:
            if r.get("status") == "completed" and (r.get("n_patients") or 0) >= 5:
                demo_run_id = r["id"]
                break
        if not demo_run_id and runs:
            demo_run_id = runs[0]["id"]
    context = {"demo_run_id": demo_run_id}
    return render(request, "landing/landing.html", context)


def simulation_list(request):
    """Simulation list page: all runs, stats, and new-sim modal."""
    runs = _get_runs()
    context = {"runs": runs}
    return render(request, "simulation/simulation_list.html", context)


# ─── Demo Pages ─────────────────────────────────────────────

def demo_anti_hallucination(request):
    """Anti-Hallucination technology demo page."""
    return render(request, "demo/anti_hallucination.html")


def demo_medgemma(request):
    """MedGemma Vision technology demo page."""
    return render(request, "demo/medgemma.html")


_MEDGEMMA_PROMPT = (
    "You are a clinical dermatology expert. You are given two images of the same patient:\n"
    "- Image 1: Baseline photograph (before treatment)\n"
    "- Image 2: Current photograph\n\n"
    "Compare the two images and identify any NEW adverse events (AEs) visible in Image 2 "
    "that were NOT present in Image 1. Use the CTCAE categories and grading criteria below "
    "to classify and grade each finding. If the patient's appearance is unchanged between "
    "the two images, return an empty list.\n\n"
    "## AE Categories & Grading\n\n"
    "### rash_maculopapular\n"
    "- Grade 1: Faint pink macules/papules scattered on cheeks (<10% BSA); mild erythema, no scaling, subtle and localized\n"
    "- Grade 2: Visible red macules/papules spreading across cheeks and forehead (10-30% BSA); moderate erythema with fine scaling at lesion edges\n"
    "- Grade 3: Severe confluent rash covering entire face including cheeks, forehead, chin, and nose (>30% BSA); intense erythema, coarse scaling, and facial edema\n\n"
    "### rash_acneiform\n"
    "- Grade 1: Few small papules on forehead (<10% BSA); non-inflamed or mildly inflamed, skin-colored to pink\n"
    "- Grade 2: Multiple erythematous papules and pustules on cheeks and forehead (10-30% BSA); visible pus-filled lesions, surrounding redness\n"
    "- Grade 3: Dense pustules covering entire face - forehead, cheeks, nose, chin (>30% BSA); confluent inflammation, crusting, signs of secondary infection\n\n"
    "### periorbital_edema\n"
    "- Grade 1: Slight puffiness of upper and lower eyelids, barely noticeable; periorbital skin appears mildly swollen\n"
    "- Grade 2: Obvious bilateral periorbital swelling; puffy, baggy eyelids with visible tissue distension; eyes appear partially narrowed\n"
    "- Grade 3: Severe periorbital edema causing near-closure of eyes; tense, shiny skin around orbital rims, eye-opening significantly impaired\n\n"
    "### sjs_prodrome\n"
    "- Grade 1: Lip redness and dryness; vermilion border appears erythematous, slight chapping without blistering\n"
    "- Grade 2: Lip and oral mucosal blistering; fluid-filled vesicles on lip surface and inner mouth; erosions with crusting at lip margins\n"
    "- Grade 3: Beginning of epidermal detachment on lips and perioral skin; large erosions, bleeding mucosa, severe crusting extending beyond lip borders\n\n"
    "### stomatitis\n"
    "- Grade 1: Mild redness or minor aphthous-like ulcer on lip mucosa; slight discomfort, no visible swelling from outside\n"
    "- Grade 2: Visible cracking and erythema at lip corners with shallow erosions; perioral redness, mild swelling of the lips\n"
    "- Grade 3: Severe lip swelling and deep erosions visible on external lip surface; crusting, bleeding, perioral inflammation\n\n"
    "### pruritus\n"
    "- Grade 1: Mild localized skin excoriation marks on forehead or cheeks; faint scratch marks, minimal erythema\n"
    "- Grade 2: Moderate visible scratch marks and erythema across face; dry, irritated skin with diffuse redness\n"
    "- Grade 3: Severe widespread excoriations with lichenification; intense erythema, bleeding scratch marks, facial edema from chronic scratching\n\n"
    "### alopecia\n"
    "- Grade 1: Mild hair thinning visible at temples and frontal hairline; slightly widened part line, subtle compared to baseline\n"
    "- Grade 2: Obvious diffuse hair thinning with clearly visible scalp through hair; temporal recession, noticeably sparse hair\n\n"
    "## Output Format\n"
    "Return a JSON array of detected AEs. Each element:\n"
    '{"ae_term": "<category>", "grade": <1|2|3>, "confidence": <0.0-1.0>, '
    '"reasoning": "<clinical description of what changed from Image 1 to Image 2>"}\n\n'
    "If no change from baseline is detected, return: []"
)


def _medgemma_infer(baseline_name, current_name, vllm_url, model_id):
    """Shared inference logic for MedGemma vLLM calls."""
    import base64
    import re
    import urllib.request
    import urllib.error

    img_dir = Path(settings.BASE_DIR) / "static_dirs" / "assets" / "medgemma"
    baseline_path = img_dir / baseline_name
    current_path = img_dir / current_name

    if not baseline_path.exists() or not current_path.exists():
        return None, "Image not found"

    b64_baseline = base64.b64encode(baseline_path.read_bytes()).decode("utf-8")
    b64_current = base64.b64encode(current_path.read_bytes()).decode("utf-8")

    payload = json.dumps({
        "model": model_id,
        "messages": [{
            "role": "user",
            "content": [
                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64_baseline}"}},
                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64_current}"}},
                {"type": "text", "text": _MEDGEMMA_PROMPT},
            ],
        }],
        "max_completion_tokens": 1024,
        "temperature": 0,
    }).encode("utf-8")

    req = urllib.request.Request(
        f"{vllm_url}/chat/completions",
        data=payload,
        headers={"Content-Type": "application/json"},
    )

    t0 = time.time()
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            result = json.loads(resp.read().decode("utf-8"))
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as e:
        return None, f"vLLM request failed: {e}"
    latency_ms = (time.time() - t0) * 1000

    raw_text = result.get("choices", [{}])[0].get("message", {}).get("content", "")

    detected = []
    try:
        data = json.loads(raw_text)
        if isinstance(data, list):
            detected = data
    except (json.JSONDecodeError, TypeError):
        match = re.search(r'\[.*\]', raw_text, re.DOTALL)
        if match:
            try:
                data = json.loads(match.group())
                if isinstance(data, list):
                    detected = data
            except (json.JSONDecodeError, TypeError):
                pass

    return {
        "detected_aes": detected,
        "raw_output": raw_text,
        "latency_ms": round(latency_ms, 1),
    }, None


@csrf_exempt
@require_POST
def api_medgemma_analyze(request):
    """Run live MedGemma 4B finetuned inference via vLLM."""
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    baseline_name = body.get("baseline")
    current_name = body.get("current")
    if not baseline_name or not current_name:
        return JsonResponse({"error": "baseline and current are required"}, status=400)

    vllm_url = os.environ.get("MEDGEMMA4B_VLLM_BASE_URL", "http://clara-medgemma4b-ft:8000/v1")
    model_id = os.environ.get("MEDGEMMA4B_MODEL_ID", "medgemma-4b-ctcae")

    result, err = _medgemma_infer(baseline_name, current_name, vllm_url, model_id)
    if err:
        status = 404 if "not found" in err else 502
        return JsonResponse({"error": err}, status=status)
    return JsonResponse(result)


@csrf_exempt
@require_POST
def api_medgemma_analyze_base(request):
    """Run live MedGemma 4B base (zero-shot) inference via vLLM."""
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    baseline_name = body.get("baseline")
    current_name = body.get("current")
    if not baseline_name or not current_name:
        return JsonResponse({"error": "baseline and current are required"}, status=400)

    vllm_url = os.environ.get("MEDGEMMA4B_BASE_VLLM_BASE_URL", "http://clara-medgemma4b:8000/v1")
    model_id = os.environ.get("MEDGEMMA4B_BASE_MODEL_ID", "google/medgemma-4b-it")

    result, err = _medgemma_infer(baseline_name, current_name, vllm_url, model_id)
    if err:
        status = 404 if "not found" in err else 502
        return JsonResponse({"error": err}, status=status)
    return JsonResponse(result)


def demo_hazard(request):
    """Hazard Engine technology demo page."""
    return render(request, "demo/hazard.html")


def demo_patient_init(request):
    """Patient Initialization demo — single patient generation with avatar."""
    return render(request, "demo/patient_init.html")


def demo_daily_sim(request):
    """Daily Simulation demo — step-by-step daily simulation with hazard engine."""
    return render(request, "demo/daily_sim.html")


def demo_validate_sim(request):
    """Validate Simulation — rule-set vs simulation statistical comparison."""
    import json as _json
    run_id = "20260224_061414_Etoposide___Cisplatin_100pt_126d"
    run_dir = DATA_DIR / "runs" / run_id
    ctx = {"run_id": run_id}
    val_path = run_dir / "validation" / "ruleset_validation_natural_v4.json"
    if val_path.exists():
        ctx["validation_json"] = val_path.read_text(encoding="utf-8")
    rs_path = run_dir / "rule_set.json"
    if rs_path.exists():
        rs = _json.loads(rs_path.read_text(encoding="utf-8"))
        ctx["drug_name"] = rs.get("drug_name", "Unknown")
    return render(request, "demo/validate_sim.html", ctx)


# ─── AntiHallu API ───────────────────────────────────────────

ANTIHALLU_ASSETS = Path(settings.BASE_DIR) / "static_dirs" / "assets" / "antihallu"
_antihallu_log = logging.getLogger("antihallu")

# FastAPI endpoint for antihallu demo
_AH_FASTAPI_URL = os.environ.get("ANTIHALLU_FASTAPI_URL", "http://clara-antihallu:8000").rstrip("/")


@require_GET
def api_antihallu_examples(request):
    """Return AntiHallu example questions."""
    try:
        data = json.loads((ANTIHALLU_ASSETS / "examples.json").read_text(encoding="utf-8"))
    except FileNotFoundError:
        return JsonResponse({"error": "examples.json not found"}, status=404)
    return JsonResponse(data)


def _fastapi_antihallu_generate(question: str):
    """Call the AntiHallu FastAPI server (/api/generate). Returns dict or None."""
    if not _AH_FASTAPI_URL:
        return None
    payload = {"question": question}
    body_bytes = json.dumps(payload).encode("utf-8")
    req = Request(
        f"{_AH_FASTAPI_URL}/api/generate",
        data=body_bytes,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urlopen(req, timeout=120) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        return data
    except (URLError, OSError, json.JSONDecodeError, TimeoutError, KeyError) as exc:
        _antihallu_log.warning("AntiHallu FastAPI call failed (%s): %s", _AH_FASTAPI_URL, exc)
        return None


def _cache_lookup(question: str):
    """Look up a question in the local AntiHallu cache. Returns dict or None."""
    try:
        cache = json.loads(
            (ANTIHALLU_ASSETS / "cache.json").read_text(encoding="utf-8")
        )
    except FileNotFoundError:
        return None
    return cache.get(question.lower())


@csrf_exempt
@require_POST
def api_antihallu_generate(request):
    """Generate AntiHallu comparison: FastAPI live inference with cache fallback."""
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON body"}, status=400)

    question = body.get("question", "").strip()
    if not question:
        return JsonResponse({"error": "question is required"}, status=400)

    # 1) Try live inference via FastAPI server
    live_result = _fastapi_antihallu_generate(question)
    if live_result is not None:
        live_result["live"] = True
        return JsonResponse(live_result)

    # 2) Fallback to local cache
    entry = _cache_lookup(question)
    if entry is not None:
        return JsonResponse({
            "question": entry["question"],
            "original": entry["original"],
            "defended": entry["defended"],
            "cached": True,
            "live": False,
        })

    return JsonResponse(
        {"error": "AntiHallu server unavailable and question not in cache"},
        status=503,
    )


def trial_viewer(request, run_id: str, day: int = 1):
    """Main trial viewer page — Generative Agents demo style."""
    run_path = _get_run_path(run_id)
    if not run_path.exists():
        return HttpResponse("Run not found", status=404)

    # Check which modes are available
    sim_dir = run_path / "simulations"
    if not sim_dir.exists():
        sim_dir.mkdir(parents=True, exist_ok=True)
    available_modes = []
    natural_files = [f for f in sim_dir.glob("*_natural.jsonl")
                   if "_hospital" not in f.stem]
    care_files = [f for f in sim_dir.glob("*_care_ai.jsonl")
                if "_hospital" not in f.stem]
    if natural_files:
        available_modes.append("natural")
    if care_files:
        available_modes.append("care_ai")

    # Default to first available mode if requested mode doesn't exist
    mode = request.GET.get("mode", "")
    if mode not in available_modes:
        mode = available_modes[0] if available_modes else "natural"

    view_mode = request.GET.get("view", "hr")  # default Hospital Record

    patient_ids = _list_patients(run_path)
    total_days = _count_days(run_path, mode)
    rule_set = _load_rule_set(run_path)

    # Ensure map is generated for this run
    try:
        _ensure_map_for_run(run_path, len(patient_ids))
    except Exception:
        pass  # non-critical; map will fall back to default

    # Load profiles for patient cards
    patients = []
    for pid in patient_ids:
        profile = _load_patient_profile(run_path, pid)
        day_data = _load_day_for_patient(run_path, pid, day, mode)
        patients.append(_patient_summary(
            profile, day_data, view_mode, run_path=run_path, mode=mode))

    # Collect day events across all patients
    all_events = []
    for pid in patient_ids:
        day_data = _load_day_for_patient(run_path, pid, day, mode)
        if day_data:
            all_events.extend(_extract_day_events(day_data))

    # Sort events: high severity first
    severity_order = {"high": 0, "medium": 1, "info": 2, "low": 3}
    all_events.sort(key=lambda e: severity_order.get(e["severity"], 9))

    # Cycle info
    cycle_length = 21
    if rule_set:
        td = rule_set.get("trial_design", {})
        cycle_length = td.get("cycle_length_days", 21)
    cycle = (day - 1) // cycle_length + 1
    cycle_day = (day - 1) % cycle_length + 1

    drug_name = rule_set.get("drug_name", "Unknown")
    indication = rule_set.get("indication", "")

    # Live simulation detection
    is_live = request.GET.get("live") == "1" or run_id in _live_sims

    context = {
        "run_id": run_id,
        "day": day,
        "total_days": total_days,
        "cycle": cycle,
        "cycle_day": cycle_day,
        "cycle_length": cycle_length,
        "drug_name": drug_name,
        "indication": indication,
        "mode": mode,
        "view_mode": view_mode,
        "available_modes": available_modes,
        "available_modes_json": json.dumps(available_modes),
        "patients": patients,
        "patients_json": json.dumps(patients),
        "events": all_events,
        "events_json": json.dumps(all_events),
        "patient_ids": patient_ids,
        "patient_ids_json": json.dumps(patient_ids),
        "is_live": is_live,
        "model_name": _load_run_meta(run_path).get("model", ""),
        "lab_ranges_json": json.dumps(rule_set.get("lab_reference_ranges", {}) or _extract_lab_ranges(run_path)),
    }
    return render(request, "trial/trial.html", context)


def patient_state(request, run_id: str, patient_id: str, day: int = None):
    """Patient detail page — like Generative Agents persona_state."""
    run_path = _get_run_path(run_id)
    if not run_path.exists():
        return HttpResponse("Run not found", status=404)

    # Auto-detect mode if not specified
    sim_dir = run_path / "simulations"
    avail = []
    if [f for f in sim_dir.glob("*_natural.jsonl") if "_hospital" not in f.stem]:
        avail.append("natural")
    if [f for f in sim_dir.glob("*_care_ai.jsonl") if "_hospital" not in f.stem]:
        avail.append("care_ai")
    mode = request.GET.get("mode", "")
    if mode not in avail:
        mode = avail[0] if avail else "natural"

    view_mode = request.GET.get("view", "hr")

    profile = _load_patient_profile(run_path, patient_id)
    all_days = _load_all_days_for_patient(run_path, patient_id, mode)

    if not day and all_days:
        day = all_days[-1].get("day", 1)
    current_day_data = None
    for d in all_days:
        if d.get("day") == day:
            current_day_data = d
            break

    # Django templates disallow underscore-prefixed attributes.
    # Remap _grade, _status, _days_active etc. into template-safe keys.
    if current_day_data:
        current_day_data = dict(current_day_data)

        # ── Compute RECIST scan schedule (for HR tumor reconstruction) ──
        rule_path = run_path / "rule_set.json"
        cycle_len = 21
        if rule_path.exists():
            try:
                rs = json.loads(rule_path.read_text(encoding="utf-8"))
                cycle_len = rs.get("trial_design", {}).get("cycle_length_days", 21)
            except Exception:
                pass
        total_d = len(all_days)
        first_scan = cycle_len * 2 + 7  # e.g. 49
        scan_interval = cycle_len * 2   # e.g. 42
        recist_scan_days = set()
        sd = first_scan
        while sd <= total_d:
            recist_scan_days.add(sd)
            sd += scan_interval

        if view_mode == "hr":
            # ── Hospital Record mode: replace GT fields with HR data ──
            hr = current_day_data.get("hospital_record", {})
            hr_obj = hr.get("objective", {})
            obs_types = hr.get("observation_types", [])

            # If current day has empty hospital_record, search backwards
            # for the last day with actual HR data (carry-forward logic).
            if not hr_obj.get("labs") and not hr_obj.get("vitals"):
                last_hr_day_num = None
                last_hr = {}
                last_hr_obj = {}
                for prev_d in reversed(all_days):
                    pd_num = prev_d.get("day", 0)
                    if pd_num >= day:
                        continue
                    prev_hr = prev_d.get("hospital_record", {})
                    prev_hr_obj = prev_hr.get("objective", {})
                    if prev_hr_obj.get("labs") or prev_hr_obj.get("vitals"):
                        last_hr_day_num = pd_num
                        last_hr = prev_hr
                        last_hr_obj = prev_hr_obj
                        break
                if last_hr_day_num is not None:
                    stale_days = day - last_hr_day_num
                    # Merge: use last known data but keep current day's
                    # obs_types and any AEs if present
                    if not hr_obj.get("labs"):
                        hr_obj = dict(hr_obj)
                        hr_obj["labs"] = last_hr_obj.get("labs", {})
                        hr_obj["labs_stale_days"] = stale_days
                    if not hr_obj.get("vitals"):
                        hr_obj = dict(hr_obj)
                        hr_obj["vitals"] = last_hr_obj.get("vitals", {})
                        hr_obj["vitals_stale_days"] = stale_days
                    if not hr_obj.get("active_aes"):
                        hr_obj.setdefault("active_aes",
                                          last_hr_obj.get("active_aes", []))
                    for k in ("treatment_status", "ecog", "location", "tumor"):
                        if k not in hr_obj and k in last_hr_obj:
                            hr_obj[k] = last_hr_obj[k]

            # AEs: only hospital-detected
            hr_aes = []
            for ae in hr_obj.get("active_aes", []):
                hr_aes.append({
                    "AETERM": ae.get("ae", ""),
                    "grade": ae.get("grade", 0),
                    "status": "active",
                    "days_active": ae.get("days_active", 0),
                    "channel": ae.get("channel", ""),
                    "detection_delay": ae.get("detection_delay"),
                    "detected_day": ae.get("detected_day"),
                    "AEONGO": True,
                })
            current_day_data["AE"] = hr_aes
            current_day_data["safe_AE"] = hr_aes

            # Labs: from HR (stale values)
            hr_labs = hr_obj.get("labs", {})
            hr_results = {}
            for name, info in hr_labs.items():
                val = info.get("value") if isinstance(info, dict) else info
                trend = info.get("trend", "") if isinstance(info, dict) else ""
                hr_results[name] = {
                    "LBORRES": val,
                    "LBORRESU": info.get("unit", "") if isinstance(info, dict) else "",
                    "_trend": trend,
                }
            current_day_data["LB"] = {
                "results": hr_results,
                "labs_stale_days": hr_obj.get("labs_stale_days", 0),
            }

            # Vitals: from HR
            hr_vitals = hr_obj.get("vitals", {})
            current_day_data["VS"] = {
                "TEMP_VSORRES": hr_vitals.get("BT"),
                "SYSBP_VSORRES": hr_vitals.get("SBP"),
                "DIABP_VSORRES": hr_vitals.get("DBP"),
                "PULSE_VSORRES": hr_vitals.get("HR"),
                "RESP_VSORRES": hr_vitals.get("RR"),
                "OXYSAT_VSORRES": hr_vitals.get("SpO2"),
                "WEIGHT_VSORRES": hr_vitals.get("weight_kg"),
                "vitals_stale_days": hr_obj.get("vitals_stale_days", 0),
            }

            # Subjective: from HR (empty on non-visit days)
            hr_subj = hr.get("subjective", {})
            if hr_subj:
                current_day_data["subjective"] = hr_subj
            else:
                current_day_data["subjective"] = {
                    "overall_awareness": "UNKNOWN",
                    "symptoms_patient_perceives": [],
                }

            # Objective: replace tumor and ecog with HR values
            gt_obj = current_day_data.get("objective", {})
            hr_objective = dict(gt_obj)
            hr_objective["treatment_status"] = hr_obj.get(
                "treatment_status", gt_obj.get("treatment_status"))
            hr_objective["ecog"] = hr_obj.get(
                "ecog", gt_obj.get("ecog"))
            hr_objective["_ecog_note"] = "Last clinical assessment"
            hr_objective["_labs_stale_days"] = hr_obj.get("labs_stale_days", 0)
            hr_objective["_vitals_stale_days"] = hr_obj.get("vitals_stale_days", 0)

            # ── HR Tumor: reconstruct from scan schedule if stuck ──
            hr_tumor = hr_obj.get("tumor")
            hr_tumor_stuck = _is_hr_tumor_stuck(all_days, day)
            if hr_tumor_stuck:
                # Legacy run: HR tumor never updated properly.
                # Reconstruct: find the last RECIST scan day <= current day
                # and use GT tumor from that day.
                last_scan_tumor = None
                last_scan_day_num = None
                for d in all_days:
                    d_num = d.get("day", 0)
                    if d_num > day:
                        break
                    if d_num in recist_scan_days:
                        gt_t = d.get("objective", {}).get("tumor", {})
                        if gt_t and gt_t.get("estimated_change_pct") is not None:
                            last_scan_tumor = gt_t
                            last_scan_day_num = d_num
                if last_scan_tumor:
                    hr_objective["tumor"] = last_scan_tumor
                    hr_objective["_tumor_note"] = f"RECIST scan (Day {last_scan_day_num})"
                else:
                    hr_objective["tumor"] = hr_tumor
                    hr_objective["_tumor_note"] = "Baseline only — no RECIST scan yet"
            else:
                hr_objective["tumor"] = hr_tumor
                if hr_tumor:
                    hr_objective["_tumor_note"] = "Last RECIST scan"
                else:
                    hr_objective["_tumor_note"] = "No scan performed yet"

            current_day_data["objective"] = hr_objective

            # Location for display
            current_day_data["_display_location"] = hr_obj.get("location", gt_obj.get("location", "HOME"))
            current_day_data["_hr_obs_types"] = obs_types
            current_day_data["_hr_is_visit"] = ("scheduled_visit" in obs_types or "er_visit" in obs_types)
        else:
            # GT mode: standard safe_AE mapping
            safe_aes = []
            for ae in current_day_data.get("AE", []):
                safe_ae = {k.lstrip("_") if k.startswith("_") else k: v
                           for k, v in ae.items()}
                safe_aes.append(safe_ae)
            current_day_data["safe_AE"] = safe_aes

            # GT mode: normalise VS keys for template compatibility
            gt_vs = current_day_data.get("VS")
            if isinstance(gt_vs, dict):
                if "_SpO2" in gt_vs and "OXYSAT_VSORRES" not in gt_vs:
                    gt_vs["OXYSAT_VSORRES"] = gt_vs["_SpO2"]

            # Location for display
            gt_obj = current_day_data.get("objective", {})
            current_day_data["_display_location"] = gt_obj.get("location", "HOME")
            hr = current_day_data.get("hospital_record", {})
            obs_types = hr.get("observation_types", [])
            current_day_data["_hr_is_visit"] = ("scheduled_visit" in obs_types or "er_visit" in obs_types)
            current_day_data["_hr_obs_types"] = obs_types

    # Filter: only show data up to and including the current viewing day
    visible_days = [d for d in all_days if d.get("day", 0) <= day]

    if view_mode == "hr":
        # Hospital Record: AEs, labs, vitals from hospital_record only
        ae_timeline = []
        _hr_ae_seen = {}  # track per-AE to avoid duplicates per day
        for d in visible_days:
            day_num = d.get("day", 0)
            hr = d.get("hospital_record", {}).get("objective", {})
            for ae in hr.get("active_aes", []):
                ae_term = ae.get("ae")
                detected_day = ae.get("detected_day", day_num)
                resolved_day = ae.get("resolved_day")
                status = ae.get("status", "active")

                ae_timeline.append({
                    "day": day_num,
                    "term": ae_term,
                    "grade": ae.get("grade"),
                    "status": status,
                    "days_active": ae.get("days_active", 0),
                    "channel": ae.get("channel", ""),
                    "detection_delay": ae.get("detection_delay"),
                    "detected_day": detected_day,
                    "resolved_day": resolved_day,
                    "onset_day": ae.get("onset_day"),
                })

        lab_trends = {}
        for d in visible_days:
            day_num = d.get("day", 0)
            hr = d.get("hospital_record", {}).get("objective", {})
            hr_labs = hr.get("labs", {})
            for lab_name, lab_info in hr_labs.items():
                if lab_name not in lab_trends:
                    lab_trends[lab_name] = []
                val = lab_info.get("value") if isinstance(lab_info, dict) else lab_info
                if val is not None:
                    lab_trends[lab_name].append({
                        "day": day_num, "value": val, "unit": "",
                    })
    else:
        # Ground Truth: full data
        ae_timeline = []
        for d in visible_days:
            day_num = d.get("day", 0)
            for ae in d.get("AE", []):
                ae_timeline.append({
                    "day": day_num,
                    "term": ae.get("AETERM"),
                    "grade": ae.get("_grade"),
                    "status": ae.get("_status"),
                    "days_active": ae.get("_days_active"),
                })

        lab_trends = {}
        for d in visible_days:
            day_num = d.get("day", 0)
            results = d.get("LB", {}).get("results", {})
            for lab_name, lab_val in results.items():
                if lab_name not in lab_trends:
                    lab_trends[lab_name] = []
                lab_trends[lab_name].append({
                    "day": day_num,
                    "value": lab_val.get("LBORRES"),
                    "unit": lab_val.get("LBORRESU", ""),
                })

    # Build event log (memory stream, like generative agents)
    event_log = []
    for d in visible_days:
        day_num = d.get("day", 0)
        hr_data = d.get("hospital_record", {})
        obs_types = hr_data.get("observation_types", [])

        if view_mode == "gt":
            day_events = _extract_day_events(d)
            for evt in day_events:
                event_log.append({
                    "day": day_num,
                    "type": evt["type"],
                    "text": evt["text"],
                    "icon": evt["icon"],
                })
        else:
            # HR mode: only show events on observation days
            if obs_types:
                hr_aes = hr_data.get("objective", {}).get("active_aes", [])
                for ae in hr_aes:
                    event_log.append({
                        "day": day_num,
                        "type": "ae_detected",
                        "text": f"Detected: {ae.get('ae', '?')} Grade {ae.get('grade', '?')} ({ae.get('channel', '')})",
                        "icon": "⚠️",
                    })
                if "scheduled_visit" in obs_types:
                    event_log.append({
                        "day": day_num, "type": "visit",
                        "text": "Scheduled clinic visit", "icon": "🏥",
                    })

        # Care records (visible in both modes — these are hospital actions)
        for cr in d.get("care_record", []):
            event_log.append({
                "day": day_num,
                "type": "care",
                "text": cr.get("summary", "Video call"),
                "icon": "📹",
            })

    # Format BMI to 1 decimal
    bmi_raw = profile.get("emr", {}).get("demographics", {}).get("bmi", "?")
    bmi_display = f"{bmi_raw:.1f}" if isinstance(bmi_raw, (int, float)) else str(bmi_raw)

    # Build mood trajectory
    mood_trajectory = []
    for d in visible_days:
        m = d.get("mood_state", {})
        if m:
            mood_trajectory.append({"day": d.get("day", 0), **m})

    context = {
        "run_id": run_id,
        "patient_id": patient_id,
        "day": day,
        "mode": mode,
        "view_mode": view_mode,
        "available_modes": avail,
        "available_modes_json": json.dumps(avail),
        "profile": profile,
        "profile_json": json.dumps(profile),
        "current_day": current_day_data,
        "current_day_json": json.dumps(current_day_data),
        "ae_timeline_json": json.dumps(ae_timeline),
        "lab_trends_json": json.dumps(lab_trends),
        "mood_trajectory_json": json.dumps(mood_trajectory),
        "event_log": event_log,
        "total_days": len(all_days),
        "bmi_display": bmi_display,
        "model_name": _load_run_meta(run_path).get("model", ""),
    }
    return render(request, "patient_state/patient_state.html", context)


# ─── JSON API ─────────────────────────────────────────────────

def api_run_meta(request, run_id: str):
    """Run metadata: patients, total days, drug info."""
    run_path = _get_run_path(run_id)
    if not run_path.exists():
        return JsonResponse({"error": "not found"}, status=404)

    mode = request.GET.get("mode", None)
    patient_ids = _list_patients(run_path)
    total_days = _count_days(run_path, mode)
    rule_set = _load_rule_set(run_path)

    return JsonResponse({
        "run_id": run_id,
        "patient_ids": patient_ids,
        "total_days": total_days,
        "drug_name": rule_set.get("drug_name", ""),
        "indication": rule_set.get("indication", ""),
        "cycle_length": rule_set.get("trial_design", {}).get(
            "cycle_length_days", 21),
    })


def api_day_data(request, run_id: str, day: int):
    """All patients' data for a specific day — for AJAX day navigation."""
    run_path = _get_run_path(run_id)
    if not run_path.exists():
        return JsonResponse({"error": "not found"}, status=404)

    mode = request.GET.get("mode", "natural")
    view_mode = request.GET.get("view", "hr")
    patient_ids = _list_patients(run_path)
    patients = []
    all_events = []

    for pid in patient_ids:
        profile = _load_patient_profile(run_path, pid)
        day_data = _load_day_for_patient(run_path, pid, day, mode)
        patients.append(_patient_summary(
            profile, day_data, view_mode, run_path=run_path, mode=mode))
        if day_data:
            all_events.extend(_extract_day_events(day_data))

    severity_order = {"high": 0, "medium": 1, "info": 2, "low": 3}
    all_events.sort(key=lambda e: severity_order.get(e["severity"], 9))

    return JsonResponse({
        "day": day,
        "patients": patients,
        "events": all_events,
    })


def api_patient_timeline(request, run_id: str, patient_id: str):
    """Full timeline for a patient — for charts and detailed view."""
    run_path = _get_run_path(run_id)
    if not run_path.exists():
        return JsonResponse({"error": "not found"}, status=404)

    mode = request.GET.get("mode", "natural")
    all_days = _load_all_days_for_patient(run_path, patient_id, mode)
    profile = _load_patient_profile(run_path, patient_id)

    return JsonResponse({
        "patient_id": patient_id,
        "profile": profile,
        "days": all_days,
    })


# ─── SSE (Server-Sent Events) for auto-play ──────────────────

def sse_stream(request, run_id: str):
    """
    Concordia-style SSE endpoint for auto-play mode.
    Client sends speed via query param: ?speed=1 (days per second).
    Streams day data as events.
    """
    run_path = _get_run_path(run_id)
    if not run_path.exists():
        return HttpResponse("Run not found", status=404)

    speed = float(request.GET.get("speed", "1"))
    start_day = int(request.GET.get("start", "1"))
    mode = request.GET.get("mode", "natural")
    view_mode = request.GET.get("view", "hr")
    total_days = _count_days(run_path, mode)
    patient_ids = _list_patients(run_path)

    def event_stream():
        for day in range(start_day, total_days + 1):
            patients = []
            all_events = []
            for pid in patient_ids:
                profile = _load_patient_profile(run_path, pid)
                day_data = _load_day_for_patient(run_path, pid, day, mode)
                patients.append(_patient_summary(
                    profile, day_data, view_mode,
                    run_path=run_path, mode=mode))
                if day_data:
                    all_events.extend(_extract_day_events(day_data))

            payload = json.dumps({
                "day": day,
                "patients": patients,
                "events": all_events,
            })
            yield f"event: day\ndata: {payload}\n\n"

            if speed > 0:
                time.sleep(1.0 / speed)

        yield "event: done\ndata: {}\n\n"

    response = StreamingHttpResponse(
        event_stream(), content_type="text/event-stream"
    )
    response["Cache-Control"] = "no-cache"
    response["X-Accel-Buffering"] = "no"
    return response


# ═══════════════════════════════════════════════════════════════
# Care Agent Demo
# ═══════════════════════════════════════════════════════════════

def _load_virtual_patients():
    """Load virtual patient config for Care Agent demo."""
    config_path = Path(settings.BASE_DIR).parent / "data" / "multimodal" / "care_agent_patients.json"
    if config_path.exists():
        return json.loads(config_path.read_text(encoding="utf-8"))
    return {"patients": [], "drug_name": "Unknown", "indication": ""}


def demo_care_agent(request):
    """Care Agent demo page — Tab 1: MedGemma Vision, Tab 2: Care Agent Live."""
    config = _load_virtual_patients()
    return render(request, "demo/care_agent.html", {"vpatients": config.get("patients", [])})


@csrf_exempt
@require_POST
def api_care_agent_run(request):
    """Run Care Agent for a virtual patient — SSE stream via Care AI /v1/consult API."""
    from django.http import StreamingHttpResponse
    import queue, threading, base64, requests as _requests

    body = json.loads(request.body)
    patient_id = body.get("patient_id", "")

    config = _load_virtual_patients()
    vpt = None
    for p in config.get("patients", []):
        if p["id"] == patient_id:
            vpt = p
            break
    if not vpt:
        return JsonResponse({"error": "Virtual patient not found"}, status=404)

    drug_name = config.get("drug_name", "Unknown")
    indication = config.get("indication", "")
    rep = vpt.get("representative_day", {})
    demographics = vpt.get("profile", {})

    q = queue.Queue()

    CARE_AI_URL = os.environ.get("CARE_AI_API_URL", "http://clara-care-ai:8300")

    def _run():
        try:
            import time as _time
            import logging
            log = logging.getLogger("care_agent")

            face_idx = vpt["face_idx"]
            ae_term = rep.get("ae")
            ae_grade = rep.get("grade", 0)
            image_file = rep.get("image", "")
            audio_label = rep.get("audio_label", "none")
            baseline_image = f"normal_{face_idx}.png"
            patient_text = vpt.get("patient_text", "")
            audio_file = vpt.get("audio_file", f"patient_{face_idx}.wav")

            current_labs = vpt.get("current_labs", {})
            baseline_labs = vpt.get("baseline_labs", {})
            current_vitals = vpt.get("current_vitals", {})
            current_meds = vpt.get("current_medications", [])
            med_history = vpt.get("medical_history", [])
            ecog = vpt.get("ecog", 1)
            mood_data = vpt.get("mood", {})

            # --- Load image + audio as base64 ---
            data_dir = Path(settings.BASE_DIR).parent / "data" / "multimodal"
            img_path = data_dir / "v3_images" / image_file
            audio_path = data_dir / "audio" / "patient1" / audio_file

            image_b64 = None
            if img_path.exists():
                image_b64 = base64.b64encode(img_path.read_bytes()).decode()

            audio_b64 = None
            if audio_path.exists():
                audio_b64 = base64.b64encode(audio_path.read_bytes()).decode()

            # --- SSE: session_start ---
            q.put(json.dumps({"type": "session_start", "patient_id": patient_id,
                              "total_days": 1}))

            # --- SSE: day_start ---
            q.put(json.dumps({
                "type": "day_start",
                "image": image_file,
                "baseline_image": baseline_image,
                "audio_label": audio_label,
                "audio_url": f"/api/care-agent/media/audio/{audio_file}" if audio_b64 else None,
                "has_visual_ae": bool(ae_term),
                "gt_ae": ae_term or "none",
                "gt_grade": ae_grade,
            }))

            # --- SSE: patient_greet (patient text — before inference starts) ---
            q.put(json.dumps({
                "type": "patient_greet",
                "text": patient_text,
                "audio_url": f"/api/care-agent/media/audio/{audio_file}" if audio_b64 else None,
                "symptoms": [],
                "mood": "neutral",
            }))

            # --- SSE: inference_start (thinking bubble appears) ---
            q.put(json.dumps({"type": "inference_start", "has_audio": bool(audio_b64)}))

            # ===== Step 1: HeAR (cough detection) =====
            audio_assessment = None
            audio_text = "No cough detected"
            hear_result = {}
            if audio_b64:
                try:
                    t0 = _time.time()
                    resp = _requests.post(
                        f"{CARE_AI_URL}/v1/cough",
                        json={"audio_b64": audio_b64},
                        timeout=60,
                    )
                    resp.raise_for_status()
                    cough_data = resp.json()
                    audio_assessment = cough_data.get("audio_assessment")
                    cough_ms = cough_data.get("latency_ms", 0)
                    if audio_assessment:
                        hear_result = {
                            "cough_detected": audio_assessment.get("cough_detected", False),
                            "majority_type": audio_assessment.get("majority_type"),
                            "num_cough_segments": audio_assessment.get("num_cough_segments", 0),
                            "num_energy_segments": audio_assessment.get("num_energy_segments", 0),
                            "duration_sec": audio_assessment.get("duration_sec", 0),
                            "vote_counts": audio_assessment.get("vote_counts", {}),
                            "latency_ms": cough_ms,
                        }
                        if audio_assessment.get("cough_detected"):
                            mtype = audio_assessment.get("majority_type", "dry")
                            audio_text = f"{mtype.capitalize()} cough detected"
                except Exception as exc:
                    log.warning("HeAR /v1/cough error: %s", exc)

                # --- SSE: hear_result ---
                q.put(json.dumps({"type": "hear_result", **hear_result}))
            else:
                q.put(json.dumps({"type": "hear_unavailable"}))

            # ===== Step 2: MedASR (transcription) =====
            medical_transcript = None
            if audio_b64:
                try:
                    t0 = _time.time()
                    resp = _requests.post(
                        f"{CARE_AI_URL}/v1/transcribe",
                        json={"audio_b64": audio_b64},
                        timeout=60,
                    )
                    resp.raise_for_status()
                    asr_data = resp.json()
                    medical_transcript = asr_data.get("medical_transcript")
                    medasr_ms = asr_data.get("latency_ms", 0)
                except Exception as exc:
                    log.warning("MedASR /v1/transcribe error: %s", exc)

                # --- SSE: medasr_result ---
                medasr_info = {}
                if medical_transcript:
                    words = len(medical_transcript.split())
                    medasr_info = {"transcript": medical_transcript, "word_count": words, "latency_ms": medasr_ms}
                q.put(json.dumps({"type": "medasr_result", **medasr_info}))
            else:
                q.put(json.dumps({"type": "medasr_unavailable"}))

            # ===== Step 3: SigLIP (visual classification) =====
            visual_assessment = {}
            siglip_findings = []
            siglip_ms = 0
            raw_pred = {}
            if image_b64:
                try:
                    t0 = _time.time()
                    resp = _requests.post(
                        f"{CARE_AI_URL}/v1/classify",
                        json={"image_b64": image_b64},
                        timeout=60,
                    )
                    resp.raise_for_status()
                    classify_data = resp.json()
                    visual_assessment = classify_data.get("visual_assessment", {})
                    siglip_ms = classify_data.get("latency_ms", 0)
                    va_findings = visual_assessment.get("findings", [])
                    raw_pred = visual_assessment.get("raw_prediction", {})
                    for f in va_findings:
                        siglip_findings.append({
                            "ae_term": f.get("ae_term") or "normal",
                            "grade": f.get("estimated_grade") or 0,
                            "confidence": f.get("confidence", 0),
                            "description": f.get("description", ""),
                        })
                except Exception as exc:
                    log.warning("SigLIP /v1/classify error: %s", exc)

            # --- SSE: siglip_result ---
            q.put(json.dumps({
                "type": "siglip_result",
                "findings": siglip_findings,
                "raw_prediction": raw_pred,
                "general_observations": visual_assessment.get("general_observations", []),
                "latency_ms": siglip_ms,
                "baseline_image": baseline_image,
                "current_image": image_file,
            }))

            # --- Build visual_obs for nurse_context ---
            visual_obs = []
            if siglip_findings:
                for vf in siglip_findings:
                    if vf["ae_term"] != "normal":
                        visual_obs.append(f"{vf['ae_term'].replace('_',' ')} G{vf['grade']}")
            if not visual_obs:
                visual_obs.append("No visual AE changes detected")

            visual_assessment_ctx = {"findings": siglip_findings, "general_observations": visual_obs}

            # --- SSE: nurse_context ---
            q.put(json.dumps({
                "type": "nurse_context",
                "visual_assessment": visual_assessment_ctx,
                "audio_assessment": audio_text,
                "patient_info": demographics,
                "current_labs": current_labs,
                "baseline_labs": baseline_labs,
                "current_vitals": current_vitals,
                "current_medications": current_meds,
                "medical_history": [h.get("condition", "") + (" (" + h.get("medication", "") + ")" if h.get("medication") and h["medication"] != "none" else "") for h in med_history],
                "ecog": ecog,
                "mood": mood_data,
                "medical_transcript": medical_transcript,
            }))

            # ===== Step 4: NurseEngine (MedGemma + TTS) =====
            nurse_structured = {}
            nurse_text = ""
            nurse_audio_b64 = None
            session_id = ""
            consult_elapsed = 0
            try:
                t0 = _time.time()
                resp = _requests.post(
                    f"{CARE_AI_URL}/v1/nurse",
                    json={
                        "patient_text": patient_text,
                        "visual_assessment": visual_assessment,
                        "audio_assessment": audio_assessment,
                        "medical_transcript": medical_transcript,
                        "drug_name": drug_name,
                        "indication": indication,
                        "skip_tts": False,
                    },
                    timeout=120,
                )
                resp.raise_for_status()
                nurse_data = resp.json()
                session_id = nurse_data.get("session_id", "")
                nurse_text = nurse_data.get("nurse_text", "")
                nurse_structured = nurse_data.get("nurse_structured", {})
                nurse_audio_b64 = nurse_data.get("audio_base64")
                consult_elapsed = (nurse_data.get("latency_ms") or {}).get("total_ms", 0)
            except Exception as exc:
                log.error("Nurse /v1/nurse error: %s", exc)
                q.put(json.dumps({"type": "error", "message": f"Nurse API error: {exc}"}))
                q.put(json.dumps({"type": "finished"}))
                return

            nurse_questions = nurse_structured.get("questions", [])
            nurse_concerns = nurse_structured.get("preliminary_concerns", [])

            # Build nurse audio URL if TTS was returned
            nurse_audio_url = None
            if nurse_audio_b64:
                import base64 as _b64
                nurse_wav_name = f"nurse_{face_idx}.wav"
                nurse_audio_dir = data_dir / "audio" / "nurse_1"
                nurse_audio_dir.mkdir(parents=True, exist_ok=True)
                nurse_wav_path = nurse_audio_dir / nurse_wav_name
                nurse_wav_path.write_bytes(_b64.b64decode(nurse_audio_b64))
                nurse_audio_url = f"/api/care-agent/media/audio/{nurse_wav_name}"

            # --- SSE: nurse_turn ---
            q.put(json.dumps({
                "type": "nurse_turn", "turn": 1,
                "text": nurse_text,
                "questions": nurse_questions,
                "concerns": nurse_concerns,
                "audio_url": nurse_audio_url,
                "approach_style": nurse_structured.get("approach_style", "empathetic"),
                "session_id": session_id,
            }))

            # --- SSE: assessment ---
            detected_aes = []
            for f in siglip_findings:
                if f["ae_term"] != "normal" and f["confidence"] > 0.1:
                    detected_aes.append({
                        "ae_term": f["ae_term"],
                        "grade": f["grade"],
                        "confidence": f["confidence"],
                        "source": "visual",
                    })
            for concern in nurse_concerns:
                concern_lower = concern.lower().replace(" ", "_")
                already = any(a["ae_term"] == concern_lower for a in detected_aes)
                if not already:
                    detected_aes.append({
                        "ae_term": concern_lower,
                        "grade": 0,
                        "confidence": 0,
                        "source": "nurse_concern",
                    })

            concern_level = "low"
            if any(a.get("grade", 0) >= 3 for a in detected_aes):
                concern_level = "high"
            elif any(a.get("grade", 0) >= 2 for a in detected_aes):
                concern_level = "moderate"

            q.put(json.dumps({
                "type": "assessment",
                "detected_aes": detected_aes,
                "concern": concern_level,
                "action": "recommend_early_visit" if concern_level in ("high", "moderate") else "continue_monitoring",
                "latency_ms": consult_elapsed,
            }))

            # --- SSE: day_end ---
            gt_list = [{"ae": ae_term, "grade": ae_grade}] if ae_term else []
            q.put(json.dumps({
                "type": "day_end",
                "gt_aes": gt_list,
                "detected_aes": detected_aes,
            }))

            q.put(json.dumps({"type": "finished"}))
        except Exception as e:
            import traceback
            q.put(json.dumps({"type": "error", "message": str(e),
                              "trace": traceback.format_exc()[-500:]}))
        finally:
            q.put(None)

    def _stream():
        thread = threading.Thread(target=_run, daemon=True)
        thread.start()
        while True:
            item = q.get()
            if item is None:
                break
            yield f"data: {item}\n\n"

    resp = StreamingHttpResponse(_stream(), content_type="text/event-stream")
    resp["Cache-Control"] = "no-cache"
    resp["X-Accel-Buffering"] = "no"
    return resp


@require_GET
def api_care_agent_patients(request):
    """List virtual patients for Care Agent demo."""
    config = _load_virtual_patients()
    patients = []
    for p in config.get("patients", []):
        rep = p.get("representative_day", {})
        patients.append({
            "patient_id": p["id"],
            "age": p.get("profile", {}).get("age"),
            "sex": p.get("profile", {}).get("sex"),
            "race": p.get("profile", {}).get("race"),
            "persona": p.get("persona", ""),
            "ae_type": rep.get("ae", ""),
            "ae_grade": rep.get("grade", 0),
            "rep_day": rep.get("day", 0),
            "baseline_image": f"normal_{p.get('face_idx', 0)}.png",
        })
    return JsonResponse({"patients": patients})


@require_GET
def api_care_agent_media(request, media_type: str, filename: str):
    """Serve multimodal assets (images / audio) for Care Agent demo."""
    from django.http import FileResponse
    base = Path(settings.BASE_DIR).parent / "data" / "multimodal"
    if media_type == "image":
        fpath = base / "v3_images" / filename
        content_type = "image/png"
    elif media_type == "audio":
        # Try patient audio first, then nurse audio, then legacy
        fpath = base / "audio" / "patient1" / filename
        if not fpath.exists():
            fpath = base / "audio" / "nurse_1" / filename
        if not fpath.exists():
            fpath = base / "generated_voices_v4" / filename
        content_type = "audio/wav"
    else:
        return HttpResponse("Invalid media type", status=400)

    if not fpath.exists() or ".." in filename:
        return HttpResponse("Not found", status=404)

    return FileResponse(open(fpath, "rb"), content_type=content_type)


@csrf_exempt
@require_POST
def api_care_agent_chat(request):
    """Proxy follow-up chat to Care AI /v1/chat endpoint."""
    import requests as _requests

    body = json.loads(request.body)
    session_id = body.get("session_id", "")
    message = body.get("message", "")

    if not session_id or not message:
        return JsonResponse({"error": "session_id and message are required"}, status=400)

    CARE_AI_URL = os.environ.get("CARE_AI_API_URL", "http://clara-care-ai:8300")

    try:
        resp = _requests.post(
            f"{CARE_AI_URL}/v1/chat",
            json={
                "session_id": session_id,
                "message": message,
                "skip_tts": body.get("skip_tts", False),
            },
            timeout=60,
        )
        resp.raise_for_status()
        data = resp.json()

        # If TTS audio returned, save to file and provide URL
        audio_url = None
        if data.get("audio_base64"):
            import base64
            data_dir = Path(settings.BASE_DIR).parent / "data" / "multimodal" / "audio" / "nurse_1"
            data_dir.mkdir(parents=True, exist_ok=True)
            audio_filename = f"chat_{session_id[:8]}_{int(time.time())}.wav"
            audio_path = data_dir / audio_filename
            audio_path.write_bytes(base64.b64decode(data["audio_base64"]))
            audio_url = f"/api/care-agent/media/audio/{audio_filename}"

        return JsonResponse({
            "nurse_text": data.get("nurse_text", ""),
            "nurse_structured": data.get("nurse_structured", {}),
            "audio_url": audio_url,
            "latency_ms": data.get("latency_ms", {}),
        })
    except _requests.exceptions.HTTPError as exc:
        status = exc.response.status_code if exc.response is not None else 502
        detail = ""
        try:
            detail = exc.response.json().get("detail", str(exc))
        except Exception:
            detail = str(exc)
        return JsonResponse({"error": detail}, status=status)
    except Exception as exc:
        return JsonResponse({"error": f"Care AI chat error: {exc}"}, status=502)


# ═══════════════════════════════════════════════════════════════
# Interactive Game Mode — Care Agent 대신 사람이 참여하는 시뮬레이션
# ═══════════════════════════════════════════════════════════════

def game_landing(request, run_id: str):
    """게임 모드 환자 선택 화면."""
    run_path = _get_run_path(run_id)
    if not run_path.exists():
        return HttpResponse("Run not found", status=404)

    patient_ids = _list_patients(run_path)
    rule_set = _load_rule_set(run_path)

    persona_labels = {
        "stoic_minimizer": "Stoic",
        "anxious_reporter": "Anxious",
        "shame_avoidant": "Avoidant",
        "confused_elderly": "Confused",
        "health_literate": "Informed",
        "minimizer": "Minimizer",
        "catastrophizer": "Worried",
        "caregiver_dependent": "Dependent",
        "language_barrier": "Language Barrier",
        "compliant_but_forgetful": "Forgetful",
    }

    patients = []
    for pid in patient_ids:
        profile = _load_patient_profile(run_path, pid)
        demo = profile.get("emr", {}).get("demographics", {})
        persona = profile.get("persona", {})
        ptype = persona.get("type", "?")
        patients.append({
            "id": pid,
            "age": demo.get("age", "?"),
            "sex": demo.get("sex", "?"),
            "ecog": demo.get("ecog_ps", "?"),
            "persona_type": ptype,
            "persona_label": persona_labels.get(ptype, ptype.replace("_", " ").title()),
        })

    return render(request, "game/game_landing.html", {
        "run_id": run_id,
        "patients": patients,
        "drug_name": rule_set.get("drug_name", ""),
        "indication": rule_set.get("indication", ""),
        "model_name": _load_run_meta(run_path).get("model", ""),
    })


def game_play(request, run_id: str, patient_id: str):
    """게임 플레이 메인 화면."""
    run_path = _get_run_path(run_id)
    if not run_path.exists():
        return HttpResponse("Run not found", status=404)

    profile = _load_patient_profile(run_path, patient_id)
    rule_set = _load_rule_set(run_path)
    difficulty = request.GET.get("difficulty", "easy")

    demo = profile.get("emr", {}).get("demographics", {})
    persona = profile.get("persona", {})

    return render(request, "game/game_play.html", {
        "run_id": run_id,
        "patient_id": patient_id,
        "profile": json.dumps(profile, ensure_ascii=False),
        "drug_name": rule_set.get("drug_name", ""),
        "indication": rule_set.get("indication", ""),
        "difficulty": difficulty,
        "patient_age": demo.get("age", "?"),
        "patient_sex": demo.get("sex", "?"),
        "patient_race": demo.get("race", ""),
        "patient_ecog": demo.get("ecog_ps", "?"),
        "persona_type": persona.get("type", ""),
        "persona_desc": persona.get("description", "")[:200],
        "model_name": _load_run_meta(run_path).get("model", ""),
    })


# ── Game API (JSON) ──────────────────────────────────

@csrf_exempt
@require_POST
def api_game_start(request):
    """게임 세션 시작. run_id + patient_id로 세션 생성."""
    from src.game_session import create_game_session

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    run_id = body.get("run_id")
    patient_id = body.get("patient_id")
    total_days = body.get("total_days", 84)
    seed = body.get("seed", 42)

    if not run_id or not patient_id:
        return JsonResponse({"error": "run_id and patient_id required"}, status=400)

    try:
        session = create_game_session(
            run_id=run_id,
            patient_id=patient_id,
            total_days=total_days,
            seed=seed,
            data_dir=str(DATA_DIR),
        )
    except FileNotFoundError as e:
        return JsonResponse({"error": str(e)}, status=404)

    return JsonResponse({
        "session_id": session.session_id,
        "patient_id": patient_id,
        "total_days": total_days,
        "status": session.status,
    })


@csrf_exempt
@require_POST
def api_game_advance(request):
    """다음 Day로 진행. GT 생성 + HR 뷰 반환."""
    from src.game_session import get_session

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    session_id = body.get("session_id")
    session = get_session(session_id)
    if not session:
        return JsonResponse({"error": "Session not found"}, status=404)

    result = session.advance_day()
    return JsonResponse(result, json_dumps_params={"ensure_ascii": False})


@csrf_exempt
@require_POST
def api_game_greet(request):
    """환자 초기 인사 생성."""
    from src.game_session import get_session

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    session_id = body.get("session_id")
    session = get_session(session_id)
    if not session:
        return JsonResponse({"error": "Session not found"}, status=404)

    result = session.patient_greet()
    return JsonResponse(result, json_dumps_params={"ensure_ascii": False})


@csrf_exempt
@require_POST
def api_game_chat(request):
    """플레이어 메시지 → 환자 AI 응답."""
    from src.game_session import get_session

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    session_id = body.get("session_id")
    message = body.get("message", "")
    session = get_session(session_id)
    if not session:
        return JsonResponse({"error": "Session not found"}, status=404)
    if not message.strip():
        return JsonResponse({"error": "Empty message"}, status=400)

    result = session.player_chat(message)
    return JsonResponse(result, json_dumps_params={"ensure_ascii": False})


@csrf_exempt
@require_POST
def api_game_end_chat(request):
    """대화 종료 + 관찰/조치 제출."""
    from src.game_session import get_session

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    session_id = body.get("session_id")
    observations = body.get("observations", [])
    actions = body.get("actions", [])
    session = get_session(session_id)
    if not session:
        return JsonResponse({"error": "Session not found"}, status=404)

    result = session.end_chat_and_submit(observations, actions)
    return JsonResponse(result, json_dumps_params={"ensure_ascii": False})


@csrf_exempt
@require_POST
def api_game_skip(request):
    """대화 없이 Day 스킵."""
    from src.game_session import get_session

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    session_id = body.get("session_id")
    session = get_session(session_id)
    if not session:
        return JsonResponse({"error": "Session not found"}, status=404)

    result = session.skip_day()
    return JsonResponse(result, json_dumps_params={"ensure_ascii": False})


@require_GET
def api_game_reveal(request, session_id: str):
    """게임 종료 후 GT 공개 + 성적표."""
    from src.game_session import get_session

    session = get_session(session_id)
    if not session:
        return JsonResponse({"error": "Session not found"}, status=404)

    result = session.reveal_ground_truth()
    return JsonResponse(result, json_dumps_params={"ensure_ascii": False})


@csrf_exempt
@require_POST
def api_game_debrief(request):
    """Day 종료 후 피드백 (HR 기준 비교)."""
    from src.game_session import get_session

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    session_id = body.get("session_id")
    session = get_session(session_id)
    if not session:
        return JsonResponse({"error": "Session not found"}, status=404)

    result = session.day_debrief()
    return JsonResponse(result, json_dumps_params={"ensure_ascii": False})


@csrf_exempt
@require_POST
def api_game_copilot(request):
    """Gemini 코파일럿: 질문/관찰 제안."""
    from src.game_session import get_session

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    session_id = body.get("session_id")
    mode = body.get("mode", "on")
    session = get_session(session_id)
    if not session:
        return JsonResponse({"error": "Session not found"}, status=404)

    result = session.get_copilot_suggestion(mode=mode)
    return JsonResponse(result, json_dumps_params={"ensure_ascii": False})


@require_GET
def api_game_sessions(request):
    """활성 게임 세션 목록."""
    from src.game_session import list_sessions
    return JsonResponse({"sessions": list_sessions()})


# ═══════════════════════════════════════════════════════
# Map Generation
# ═══════════════════════════════════════════════════════

def _ensure_map_for_run(run_path: Path, n_patients: int) -> tuple[dict, dict]:
    """Generate map for a specific run if needed. Returns (tilemap, meta)."""
    map_dir = run_path / "map"
    meta_path = map_dir / "map_meta.json"
    tilemap_path = map_dir / "tilemap.json"

    if meta_path.exists() and tilemap_path.exists():
        with open(meta_path) as f:
            meta = json.load(f)
        if meta.get("n_patients", 0) == n_patients:
            with open(tilemap_path) as f:
                tilemap = json.load(f)
            return tilemap, meta

    # Generate map sized for this run's patient count
    import sys
    tools_dir = Path(settings.BASE_DIR) / "tools"
    if str(tools_dir) not in sys.path:
        sys.path.insert(0, str(tools_dir))
    from generate_map import generate_map
    tilemap, meta = generate_map(n_patients)

    map_dir.mkdir(parents=True, exist_ok=True)
    tilemap_path.write_text(json.dumps(tilemap), encoding="utf-8")
    meta_path.write_text(json.dumps(meta, indent=2), encoding="utf-8")
    return tilemap, meta


@require_GET
def api_map_meta(request, run_id: str):
    """Map metadata (home positions, waypoints) for the tilemap."""
    run_path = _get_run_path(run_id)
    if not run_path.exists():
        return JsonResponse({"error": "not found"}, status=404)

    patient_ids = _list_patients(run_path)
    _, meta = _ensure_map_for_run(run_path, len(patient_ids))
    return JsonResponse(meta)


@require_GET
def api_map_tilemap(request, run_id: str):
    """Serve the Tiled JSON tilemap for a specific run."""
    run_path = _get_run_path(run_id)
    if not run_path.exists():
        return JsonResponse({"error": "not found"}, status=404)

    patient_ids = _list_patients(run_path)
    tilemap, _ = _ensure_map_for_run(run_path, len(patient_ids))
    return JsonResponse(tilemap)


# ═══════════════════════════════════════════════════════
# A/B Comparison Dashboard
# ═══════════════════════════════════════════════════════

def compare_dashboard(request, run_id: str):
    """A/B 비교 대시보드 페이지."""
    run_path = _get_run_path(run_id)
    report_path = run_path / "comparison_report.json"
    sim_dir = run_path / "simulations"
    _run_model = _load_run_meta(run_path).get("model", "")

    # 기본 검증: simulations 디렉토리 존재 여부
    if not sim_dir.exists():
        return render(request, "compare/compare.html", {
            "run_id": run_id, "drug_name": "Unknown", "indication": "",
            "report_json": json.dumps({"error": "Simulation directory not found"}),
            "error": "시뮬레이션 디렉토리가 존재하지 않습니다.",
            "model_name": _run_model,
        })

    # natural / care_ai 파일 존재 확인
    natural_files = list(sim_dir.glob("*_natural.jsonl"))
    care_ai_files = list(sim_dir.glob("*_care_ai.jsonl"))
    if not natural_files or not care_ai_files:
        rule_set = _load_rule_set(run_path)
        missing = []
        if not natural_files:
            missing.append("Natural")
        if not care_ai_files:
            missing.append("Care AI")
        return render(request, "compare/compare.html", {
            "run_id": run_id,
            "drug_name": rule_set.get("drug_name", "Unknown"),
            "indication": rule_set.get("indication", ""),
            "report_json": json.dumps({"error": f"Missing data: {', '.join(missing)}"}),
            "error": f"A/B 비교에 필요한 데이터가 부족합니다: {', '.join(missing)} 모드 데이터 없음. "
                     f"(Natural: {len(natural_files)}명, Care AI: {len(care_ai_files)}명)",
            "model_name": _run_model,
        })

    # comparison_report.json이 없거나 낡았으면 재생성
    try:
        needs_regen = not report_path.exists()
        if not needs_regen:
            sim_files = list(sim_dir.glob("*.jsonl"))
            if sim_files:
                newest_sim = max(f.stat().st_mtime for f in sim_files)
                if report_path.stat().st_mtime < newest_sim:
                    needs_regen = True

        if needs_regen:
            from src.evaluator import run_evaluation
            run_evaluation(run_path)

        with open(report_path) as f:
            report = json.load(f)
    except Exception as e:
        rule_set = _load_rule_set(run_path)
        return render(request, "compare/compare.html", {
            "run_id": run_id,
            "drug_name": rule_set.get("drug_name", "Unknown"),
            "indication": rule_set.get("indication", ""),
            "report_json": json.dumps({"error": str(e)}),
            "error": f"비교 리포트 생성 중 오류: {e}",
            "model_name": _run_model,
        })

    rule_set = _load_rule_set(run_path)

    return render(request, "compare/compare.html", {
        "run_id": run_id,
        "drug_name": rule_set.get("drug_name", "Unknown"),
        "indication": rule_set.get("indication", ""),
        "report_json": json.dumps(report, ensure_ascii=False),
        "model_name": _run_model,
    })


@require_GET
def api_compare_data(request, run_id: str):
    """A/B 비교 데이터 JSON API."""
    run_path = _get_run_path(run_id)
    sim_dir = run_path / "simulations"
    report_path = run_path / "comparison_report.json"

    if not sim_dir.exists():
        return JsonResponse({"error": "Simulation directory not found"}, status=404)

    natural_count = len(list(sim_dir.glob("*_natural.jsonl")))
    care_ai_count = len(list(sim_dir.glob("*_care_ai.jsonl")))
    if natural_count == 0 or care_ai_count == 0:
        return JsonResponse({
            "error": "Both natural and care_ai data required for comparison",
            "natural_count": natural_count,
            "care_ai_count": care_ai_count,
        }, status=400)

    try:
        if not report_path.exists():
            from src.evaluator import run_evaluation
            run_evaluation(run_path)

        with open(report_path) as f:
            report = json.load(f)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse(report, json_dumps_params={"ensure_ascii": False})


@require_GET
def api_compare_regenerate(request, run_id: str):
    """비교 리포트를 강제 재생성한다."""
    from src.evaluator import run_evaluation
    run_path = _get_run_path(run_id)
    sim_dir = run_path / "simulations"

    if not sim_dir.exists():
        return JsonResponse({"error": "Simulation directory not found"}, status=404)

    natural_count = len(list(sim_dir.glob("*_natural.jsonl")))
    care_ai_count = len(list(sim_dir.glob("*_care_ai.jsonl")))
    if natural_count == 0 or care_ai_count == 0:
        return JsonResponse({
            "error": "Both natural and care_ai data required",
            "natural_count": natural_count,
            "care_ai_count": care_ai_count,
        }, status=400)

    try:
        report = run_evaluation(run_path)
        return JsonResponse({"status": "regenerated", "cohort_sizes": report["cohort_sizes"]})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# ─── Live Simulation API ─────────────────────────────────────

import threading
import signal
_live_sims: dict[str, dict] = {}  # run_id -> {thread, status, runner, ...}


@csrf_exempt
@require_POST
def api_sim_start(request):
    """Start a new live simulation run.

    POST body: {
        "drug": "Padcev + Pembrolizumab",
        "indication": "metastatic urothelial carcinoma",
        "patients": 10,
        "days": 126,
        "mode": "both",
        "seed": 42  (optional)
    }
    Returns: {"run_id": "...", "status": "started"}
    """
    try:
        body = json.loads(request.body)
    except Exception:
        return JsonResponse({"error": "Invalid JSON body"}, status=400)

    drug = body.get("drug", "Padcev + Pembrolizumab")
    indication = body.get("indication", "metastatic urothelial carcinoma")
    n_patients = int(body.get("patients", 10))
    n_days = int(body.get("days", 126))
    mode = body.get("mode", "both")
    seed = body.get("seed")
    rule_set_preset = body.get("rule_set_preset")
    skip_rules = rule_set_preset is not None or body.get("skip_rules", True)
    user_api_key = body.get("api_key", "").strip()

    # Create run directory
    from datetime import datetime as _dt
    ts = _dt.now().strftime("%Y%m%d_%H%M%S")
    safe_drug = drug.replace(" ", "_").replace("+", "_")[:30]
    run_name = f"{ts}_{safe_drug}_{n_patients}pt_{n_days}d"
    run_dir = DATA_DIR / "runs" / run_name
    run_dir.mkdir(parents=True, exist_ok=True)

    def _run_simulation():
        """Background thread for running the simulation."""
        import sys as _sys
        _sys.path.insert(0, str(Path(settings.BASE_DIR).parent))

        # 사용자 API 키가 있으면 우선 사용, 없으면 .env 폴백
        if user_api_key:
            os.environ["GOOGLE_API_KEY"] = user_api_key
            from src.agents.llm_client import set_api_key
            set_api_key(user_api_key)
        else:
            # Load .env
            env_path = Path(settings.BASE_DIR).parent / ".env"
            if env_path.exists():
                for line in env_path.read_text().splitlines():
                    line = line.strip()
                    if line and not line.startswith("#") and "=" in line:
                        key, val = line.split("=", 1)
                        os.environ.setdefault(key.strip(), val.strip())

        from src.orchestrator_v2 import SimulationRunnerV2

        try:
            runner = SimulationRunnerV2(
                drug_name=drug,
                indication=indication,
                data_dir=str(run_dir),
                seed=seed,
            )
            _live_sims[run_name]["runner"] = runner

            # Phase 0: Rules
            if skip_rules:
                base_rule_path = None
                if rule_set_preset:
                    _RULE_SETS_DIR = DATA_DIR / "rule_sets"
                    _preset_map = {
                        "rule_set_calibrated_ev302": DATA_DIR / "rule_set_calibrated_ev302.json",
                        "rule_set_darbepoetin_sclc": DATA_DIR / "rule_set_darbepoetin_sclc.json",
                        "rule_set_ep_sclc": DATA_DIR / "rule_set_ep_sclc.json",
                        "rule_set_default": DATA_DIR / "rule_set.json",
                        "rs_1_Darbepoetin_alfa": _RULE_SETS_DIR / "1_Darbepoetin_alfa.json",
                        "rs_2_Etoposide_Cisplatin": _RULE_SETS_DIR / "2_Etoposide_Cisplatin.json",
                        "rs_3_CALGB9732_Paclitaxel_Cisplatin_Etoposide": _RULE_SETS_DIR / "3_CALGB9732_Paclitaxel_Cisplatin_Etoposide.json",
                        "rs_4_Carboplatin_Etoposide": _RULE_SETS_DIR / "4_Carboplatin_Etoposide.json",
                        "rs_6_Paclitaxel_Carboplatin_Bevacizumab": _RULE_SETS_DIR / "6_Paclitaxel_Carboplatin_Bevacizumab.json",
                        "rs_7_Paclitaxel_Carboplatin": _RULE_SETS_DIR / "7_Paclitaxel_Carboplatin.json",
                        "rs_8_Gemcitabine_Cisplatin": _RULE_SETS_DIR / "8_Gemcitabine_Cisplatin.json",
                    }
                    if rule_set_preset in _preset_map:
                        base_rule_path = _preset_map[rule_set_preset]
                    elif rule_set_preset.startswith("gt_"):
                        gt_folder = rule_set_preset[3:]
                        gt_path = _RULESET_DIR / "ground_truth" / gt_folder / "base.json"
                        if gt_path.exists():
                            base_rule_path = gt_path
                if not base_rule_path or not base_rule_path.exists():
                    base_rule_path = DATA_DIR / "rule_set_calibrated_ev302.json"
                    if not base_rule_path.exists():
                        base_rule_path = DATA_DIR / "rule_set.json"
                runner.load_rules(str(base_rule_path))
                import shutil
                shutil.copy2(base_rule_path, run_dir / "rule_set.json")
            else:
                runner.discover_rules()

            runner.write_run_meta(n_patients, n_days, mode, 'generating_patients')

            # Phase 1: Patients
            patients = runner.create_patients_parallel(n_patients, max_workers=10)

            # Phase 2: Daily simulation
            modes = [mode] if mode != "both" else ["natural", "care_ai"]
            for sim_mode in modes:
                runner.write_run_meta(n_patients, n_days, sim_mode, 'running')

                all_results = runner.run_parallel(
                    patients, total_days=n_days, mode=sim_mode, max_workers=10
                )

            # Phase 3: Comparison (if both modes)
            if mode == "both":
                runner.write_run_meta(n_patients, n_days, mode, 'comparing')
                try:
                    from src.evaluator import run_evaluation
                    run_evaluation(run_dir)
                except Exception as e:
                    print(f"Comparison failed: {e}")

            if runner.is_cancelled:
                runner.write_run_meta(n_patients, n_days, mode, 'cancelled')
                _live_sims[run_name]["status"] = "cancelled"
                runner.log("⛔ Simulation cancelled by user")
            else:
                runner.write_run_meta(n_patients, n_days, mode, 'completed')
                _live_sims[run_name]["status"] = "completed"
                runner.log("🏁 Simulation completed successfully")

        except Exception as e:
            try:
                runner.write_run_meta(n_patients, n_days, mode, 'failed',
                                     extra={'error': str(e)})
            except Exception:
                pass
            _live_sims[run_name]["status"] = "failed"
            _live_sims[run_name]["error"] = str(e)
            import traceback
            traceback.print_exc()

    # Start background thread
    t = threading.Thread(target=_run_simulation, daemon=True)
    _live_sims[run_name] = {
        "thread": t,
        "status": "starting",
        "run_dir": str(run_dir),
        "config": {
            "drug": drug, "indication": indication,
            "patients": n_patients, "days": n_days,
            "mode": mode, "seed": seed,
        },
    }
    t.start()

    return JsonResponse({
        "run_id": run_name,
        "status": "started",
        "url": f"/trial/{run_name}/",
    })


@require_GET
def api_sim_status(request, run_id: str):
    """Get status of a live simulation run.

    Returns progress, current day per patient, overall status.
    """
    run_path = _get_run_path(run_id)
    meta_path = run_path / "run_meta.json"

    result = {"run_id": run_id, "exists": run_path.exists()}

    # Check in-memory status
    if run_id in _live_sims:
        result["live"] = True
        result["status"] = _live_sims[run_id].get("status", "unknown")
        result["config"] = _live_sims[run_id].get("config", {})

    # Check on-disk metadata
    if meta_path.exists():
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
            result["meta"] = meta
            result["status"] = meta.get("status", result.get("status", "unknown"))
        except Exception:
            pass

    # Count current simulation files
    sim_dir = run_path / "simulations"
    if sim_dir.exists():
        natural_files = list(sim_dir.glob("*_natural.jsonl"))
        care_ai_files = list(sim_dir.glob("*_care_ai.jsonl"))
        result["files"] = {
            "natural": len(natural_files),
            "care_ai": len(care_ai_files),
        }
        # Count latest day from a random file
        if natural_files:
            try:
                with open(natural_files[0]) as f:
                    lines = f.readlines()
                if lines:
                    last = json.loads(lines[-1])
                    result["latest_day"] = last.get("day", 0)
            except Exception:
                pass
    else:
        result["files"] = {"natural": 0, "care_ai": 0}

    # Patient count
    patients_dir = run_path / "patients"
    if patients_dir.exists():
        result["patients_generated"] = len(list(patients_dir.glob("*.json")))
    else:
        result["patients_generated"] = 0

    # If no live status and no meta, it's a completed replay run
    if "status" not in result:
        if run_path.exists() and sim_dir.exists():
            result["status"] = "completed"
        else:
            result["status"] = "not_found"

    # Include log line count for live panel
    if run_id in _live_sims:
        runner = _live_sims[run_id].get("runner")
        if runner:
            result["log_count"] = len(runner._log_lines)

    return JsonResponse(result)


@require_GET
def api_sim_list(request):
    """List all runs with their status (live + completed)."""
    runs = _get_runs()

    result = []
    for run in runs:
        run_id = run["id"]
        run_path = _get_run_path(run_id)
        meta_path = run_path / "run_meta.json"

        entry = {
            "id": run_id,
            "modes": run["modes"],
            "status": "completed",  # default for old runs
        }

        # Check if running
        if run_id in _live_sims:
            entry["status"] = _live_sims[run_id].get("status", "unknown")
            entry["live"] = True

        # Check meta
        if meta_path.exists():
            try:
                meta = json.loads(meta_path.read_text(encoding="utf-8"))
                entry["drug_name"] = meta.get("drug_name")
                entry["indication"] = meta.get("indication")
                entry["n_patients"] = meta.get("n_patients")
                entry["total_days"] = meta.get("total_days")
                entry["status"] = meta.get("status", entry["status"])
                entry["started_at"] = meta.get("started_at")
                entry["completed_at"] = meta.get("completed_at")
                if meta.get("status") == "completed" and run_id in _live_sims:
                    del _live_sims[run_id]
                    entry.pop("live", None)
            except Exception:
                pass
        else:
            # Extract info from directory name
            parts = run_id.split("_")
            entry["drug_name"] = run_id

        # Patient count
        patients_dir = run_path / "patients"
        if patients_dir.exists():
            entry["n_patients"] = len(list(patients_dir.glob("*.json")))

        result.append(entry)

    return JsonResponse({"runs": result})


@csrf_exempt
@require_POST
def api_sim_stop(request, run_id: str):
    """Stop a running simulation.

    Signals the runner to cancel, then optionally deletes the run data.
    POST body: {"delete": true/false}
    """
    # Optionally delete run data
    try:
        body = json.loads(request.body) if request.body else {}
    except Exception:
        body = {}

    should_delete = body.get("delete", False)

    if run_id not in _live_sims:
        # Server may have restarted — _live_sims lost but run dir still exists.
        # Allow delete even if not tracked in memory.
        if should_delete:
            run_path = _get_run_path(run_id)
            if run_path.exists():
                import shutil
                shutil.rmtree(run_path, ignore_errors=True)
            return JsonResponse({"status": "stopped_and_deleted", "run_id": run_id})
        return JsonResponse({"error": "Run not found or not a live simulation"},
                            status=404)

    sim_info = _live_sims[run_id]
    runner = sim_info.get("runner")

    if runner:
        runner.cancel()
        runner.log("⛔ Stop requested by user")

    sim_info["status"] = "cancelling"

    if should_delete:
        run_path = _get_run_path(run_id)
        if run_path.exists():
            import shutil
            shutil.rmtree(run_path, ignore_errors=True)
        if run_id in _live_sims:
            del _live_sims[run_id]
        return JsonResponse({"status": "stopped_and_deleted", "run_id": run_id})

    return JsonResponse({"status": "stopping", "run_id": run_id})


@require_GET
def api_sim_log(request, run_id: str):
    """Get real-time log lines for a live simulation.

    Query params:
        since: int — return lines from this index (default 0)
    Returns: {"lines": [...], "next": int, "status": str}
    """
    since = int(request.GET.get("since", 0))

    result = {"run_id": run_id, "lines": [], "next": since, "status": "unknown"}

    # Try in-memory log first
    if run_id in _live_sims:
        runner = _live_sims[run_id].get("runner")
        result["status"] = _live_sims[run_id].get("status", "unknown")
        if runner:
            lines = runner.get_log(since)
            result["lines"] = lines
            result["next"] = since + len(lines)
            return JsonResponse(result)

    # Fallback: read from disk log
    run_path = _get_run_path(run_id)
    log_path = run_path / "sim_log.txt"
    if log_path.exists():
        try:
            all_lines = log_path.read_text(encoding="utf-8").splitlines()
            result["lines"] = all_lines[since:]
            result["next"] = len(all_lines)
            result["status"] = "completed"
        except Exception:
            pass

    return JsonResponse(result)


# ═══════════════════════════════════════════════════════
# Doc Agent — SAE Document Generation
# ═══════════════════════════════════════════════════════

def _load_patient_data(run_path, patient_id, mode="natural"):
    """Load patient profile + day records for doc agent.

    Uses hospital record (HR) when available — SAE reporting should be
    based on what the hospital actually observed, not ground truth.
    Falls back to GT for older runs that lack *_hospital.jsonl.
    """
    profile_path = run_path / "patients" / f"{patient_id}.json"
    hr_path = run_path / "simulations" / f"{patient_id}_{mode}_hospital.jsonl"
    gt_path = run_path / "simulations" / f"{patient_id}_{mode}.jsonl"
    sim_path = hr_path if hr_path.exists() else gt_path

    if not profile_path.exists() or not sim_path.exists():
        return None, None

    with open(profile_path, encoding="utf-8") as f:
        profile = json.load(f)

    records = []
    with open(sim_path, encoding="utf-8") as f:
        for line in f:
            if line.strip():
                records.append(json.loads(line))

    return profile, records


@csrf_exempt
@require_POST
def api_doc_generate(request):
    """Generate MedWatch 3500A + E2B XML for a specific patient SAE.

    POST body: {
        "run_id": str,
        "patient_id": str,
        "ae_term": str,
        "ae_day": int (optional),
        "mode": "natural" | "care_ai" (default: "natural"),
        "use_ai": bool (default: false),
    }
    """
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    run_id = body.get("run_id", "")
    patient_id = body.get("patient_id", "")
    ae_term = body.get("ae_term", "")
    ae_day = body.get("ae_day")
    mode = body.get("mode", "natural")
    use_ai = body.get("use_ai", False)

    if not all([run_id, patient_id, ae_term]):
        return JsonResponse({"error": "run_id, patient_id, ae_term required"}, status=400)

    run_path = _get_run_path(run_id)
    if not run_path.exists():
        return JsonResponse({"error": f"Run '{run_id}' not found"}, status=404)

    profile, records = _load_patient_data(run_path, patient_id, mode)
    if profile is None:
        return JsonResponse({"error": f"Patient '{patient_id}' not found"}, status=404)

    meta_path = run_path / "run_meta.json"
    drug_name = "Enfortumab vedotin (Padcev)"
    indication = "Metastatic urothelial carcinoma"
    if meta_path.exists():
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
            drug_name = meta.get("drug_name", drug_name)
            indication = meta.get("indication", indication)
        except Exception:
            pass

    from datetime import date as dt_date
    from src.doc_agent.service import generate_documents

    result = generate_documents(
        patient_profile=profile,
        day_records=records,
        target_ae_term=ae_term,
        run_id=run_id,
        sim_start_date=dt_date(2026, 1, 6),
        drug_name=drug_name,
        indication=indication,
        target_ae_day=ae_day,
        use_ai=use_ai,
    )

    # Record ai_fields in status file when AI is used
    if use_ai and result.get("success"):
        from datetime import datetime
        ae_slug = ae_term.replace(" ", "_").replace("/", "_")
        ai_fields = {
            "section_b.narrative": True,
            "section_c.dechallenge": True,
            "section_c.rechallenge": True,
        }
        status_data = _read_status(run_id, patient_id, ae_slug)
        status_data["ai_fields"] = ai_fields
        # Store MedDRA info if available
        meddra = result.get("meddra", {})
        if meddra:
            status_data["meddra_confidence"] = meddra.get("confidence")
            status_data["meddra_source"] = meddra.get("source")
        status_data["updated_at"] = datetime.utcnow().isoformat()
        if "created_at" not in status_data:
            status_data["created_at"] = datetime.utcnow().isoformat()
        if "status" not in status_data:
            status_data["status"] = "draft"
        _write_status(run_id, patient_id, ae_slug, status_data)

    return JsonResponse(result)


@require_GET
def api_doc_list_saes(request, run_id, patient_id):
    """List all serious AEs for a patient in a run.

    GET /api/doc/saes/<run_id>/<patient_id>/?mode=natural
    """
    mode = request.GET.get("mode", "natural")
    run_path = _get_run_path(run_id)
    if not run_path.exists():
        return JsonResponse({"error": "Run not found"}, status=404)

    profile, records = _load_patient_data(run_path, patient_id, mode)
    if profile is None:
        return JsonResponse({"error": "Patient not found"}, status=404)

    from src.doc_agent.sim_to_crf_adapter import find_serious_aes

    saes = find_serious_aes(records)
    result = []
    for sae in saes:
        ae = sae["ae_record"]
        result.append({
            "ae_term": ae.get("AETERM", ""),
            "grade": ae.get("_grade", 0),
            "onset_day": ae.get("AESTDAT"),
            "severity": ae.get("AESEV", ""),
            "action": ae.get("AEACN", ""),
            "outcome": ae.get("AEOUT", ""),
        })

    return JsonResponse({"patient_id": patient_id, "saes": result})


@require_GET
def api_doc_download(request, run_id, patient_id, filename):
    """Download a generated document (PDF or XML).

    GET /api/doc/download/<run_id>/<patient_id>/<filename>
    """
    from src.doc_agent.service import DOCS_OUTPUT_DIR

    file_path = DOCS_OUTPUT_DIR / run_id / patient_id / filename

    if not file_path.exists() or not file_path.is_file():
        return JsonResponse({"error": "File not found"}, status=404)

    try:
        file_path.resolve().relative_to(DOCS_OUTPUT_DIR.resolve())
    except ValueError:
        return JsonResponse({"error": "Access denied"}, status=403)

    if filename.endswith(".pdf"):
        content_type = "application/pdf"
    elif filename.endswith(".xml"):
        content_type = "application/xml"
    else:
        content_type = "application/octet-stream"

    disposition = "inline" if filename.endswith(".pdf") else "attachment"
    with open(file_path, "rb") as f:
        response = HttpResponse(f.read(), content_type=content_type)
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response


@require_GET
def api_doc_list(request, run_id):
    """List all generated documents for a run.

    GET /api/doc/list/<run_id>/
    """
    from src.doc_agent.service import DOCS_OUTPUT_DIR

    docs_dir = DOCS_OUTPUT_DIR / run_id
    if not docs_dir.exists():
        return JsonResponse({"run_id": run_id, "documents": []})

    documents = []
    for patient_dir in sorted(docs_dir.iterdir()):
        if not patient_dir.is_dir():
            continue
        for doc_file in sorted(patient_dir.iterdir()):
            if doc_file.is_file():
                documents.append({
                    "patient_id": patient_dir.name,
                    "filename": doc_file.name,
                    "type": "pdf" if doc_file.suffix == ".pdf" else "xml",
                    "size": doc_file.stat().st_size,
                    "download_url": f"/api/doc/download/{run_id}/{patient_dir.name}/{doc_file.name}",
                })

    return JsonResponse({"run_id": run_id, "documents": documents})


@csrf_exempt
@require_POST
def api_doc_save(request):
    """Save edited MedWatch data and regenerate PDF + E2B XML.

    POST body: {
        "run_id": str,
        "patient_id": str,
        "ae_slug": str,
        "medwatch_data": { section_a: {...}, section_b: {...}, ... },
    }
    """
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    run_id = body.get("run_id", "")
    patient_id = body.get("patient_id", "")
    ae_slug = body.get("ae_slug", "")
    mw_data = body.get("medwatch_data", {})

    if not all([run_id, patient_id, ae_slug, mw_data]):
        return JsonResponse(
            {"error": "run_id, patient_id, ae_slug, medwatch_data required"},
            status=400,
        )

    from src.doc_agent.service import DOCS_OUTPUT_DIR
    from src.doc_agent.schemas.medwatch import MedWatch3500A
    from src.doc_agent.medwatch_pdf import generate_medwatch_pdf
    from src.doc_agent.e2b_converter import convert_to_e2b_xml
    from src.doc_agent.meddra_coder import code_meddra
    from src.doc_agent.config import Settings
    from src.doc_agent.schemas.crf import CRFData

    try:
        medwatch = MedWatch3500A.model_validate(mw_data)
    except Exception as exc:
        return JsonResponse({"error": f"Invalid MedWatch data: {exc}"}, status=400)

    out_dir = DOCS_OUTPUT_DIR / run_id / patient_id
    out_dir.mkdir(parents=True, exist_ok=True)

    json_path = out_dir / f"medwatch_data_{ae_slug}.json"
    json_path.write_text(
        json.dumps(mw_data, indent=2, default=str, ensure_ascii=False),
        encoding="utf-8",
    )

    pdf_path = out_dir / f"medwatch_3500a_{ae_slug}.pdf"
    xml_path = out_dir / f"e2b_r3_{ae_slug}.xml"

    try:
        generate_medwatch_pdf(medwatch, str(pdf_path))
    except Exception as exc:
        return JsonResponse({"error": f"PDF generation failed: {exc}"}, status=500)

    # Regenerate E2B XML
    run_path = _get_run_path(run_id)
    meta_path = run_path / "run_meta.json"
    drug_name = "Enfortumab vedotin (Padcev)"
    indication = "Metastatic urothelial carcinoma"
    if meta_path.exists():
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
            drug_name = meta.get("drug_name", drug_name)
            indication = meta.get("indication", indication)
        except Exception:
            pass

    settings = Settings.from_simulation(drug_name=drug_name, indication=indication)
    ae_term = medwatch.section_g.ae_term or ae_slug.replace("_", " ")
    meddra = code_meddra(ae_term, use_medgemma=False)

    # Build minimal CRF for E2B (narrative fields already in medwatch)
    from datetime import date as dt_date
    profile, records = _load_patient_data(run_path, patient_id)
    if profile and records:
        from src.doc_agent.sim_to_crf_adapter import build_crf_for_sae
        crf = build_crf_for_sae(
            patient_profile=profile,
            day_records=records,
            target_ae_term=ae_term,
            sim_start_date=dt_date(2026, 1, 6),
        )
    else:
        crf = None

    if crf:
        try:
            e2b_xml = convert_to_e2b_xml(medwatch, crf, meddra, settings)
            xml_path.write_text(e2b_xml, encoding="utf-8")
        except Exception as exc:
            logger_msg = f"E2B regeneration failed: {exc}"

    pdf_url = f"/api/doc/download/{run_id}/{patient_id}/{pdf_path.name}"
    xml_url = f"/api/doc/download/{run_id}/{patient_id}/{xml_path.name}"

    # Update status file with refreshed MedDRA confidence
    from src.doc_agent.service import DOCS_OUTPUT_DIR as _DOCS_DIR
    from datetime import datetime
    status_path = _DOCS_DIR / run_id / patient_id / f"report_status_{ae_slug}.json"
    status_path.parent.mkdir(parents=True, exist_ok=True)
    if status_path.exists():
        try:
            status_data = json.loads(status_path.read_text(encoding="utf-8"))
        except Exception:
            status_data = {}
    else:
        status_data = {
            "status": "draft",
            "ai_fields": {},
            "reviewed_by": None,
            "reviewed_at": None,
            "created_at": datetime.utcnow().isoformat(),
        }
    status_data["meddra_confidence"] = meddra.confidence
    status_data["meddra_source"] = meddra.source
    status_data["updated_at"] = datetime.utcnow().isoformat()
    status_path.write_text(
        json.dumps(status_data, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    return JsonResponse({
        "success": True,
        "pdf_url": pdf_url,
        "xml_url": xml_url,
        "message": "Documents saved and regenerated.",
    })


# ─── SAE Status Management ──────────────────────────────────────────

def _get_status_path(run_id: str, patient_id: str, ae_slug: str) -> Path:
    """Return path to the report status JSON file."""
    from src.doc_agent.service import DOCS_OUTPUT_DIR
    return DOCS_OUTPUT_DIR / run_id / patient_id / f"report_status_{ae_slug}.json"


def _read_status(run_id: str, patient_id: str, ae_slug: str) -> dict:
    """Read status file, returning default draft status if missing."""
    status_path = _get_status_path(run_id, patient_id, ae_slug)
    if status_path.exists():
        try:
            return json.loads(status_path.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"status": "draft"}


def _write_status(run_id: str, patient_id: str, ae_slug: str, data: dict):
    """Write status file."""
    status_path = _get_status_path(run_id, patient_id, ae_slug)
    status_path.parent.mkdir(parents=True, exist_ok=True)
    status_path.write_text(
        json.dumps(data, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


@require_GET
def api_doc_get_status(request, run_id: str, patient_id: str, ae_slug: str):
    """GET /api/doc/status/<run_id>/<patient_id>/<ae_slug>/ — return report status."""
    data = _read_status(run_id, patient_id, ae_slug)
    return JsonResponse(data)


@csrf_exempt
@require_POST
def api_doc_update_status(request):
    """POST /api/doc/status — transition SAE report status.

    Body: {
        "run_id": str,
        "patient_id": str,
        "ae_slug": str,
        "status": "draft" | "under_review" | "accepted",
        "reviewed_by": str (optional),
    }
    """
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    run_id = body.get("run_id", "")
    patient_id = body.get("patient_id", "")
    ae_slug = body.get("ae_slug", "")
    new_status = body.get("status", "")
    reviewed_by = body.get("reviewed_by")

    if not all([run_id, patient_id, ae_slug, new_status]):
        return JsonResponse(
            {"error": "run_id, patient_id, ae_slug, status required"}, status=400
        )

    valid_statuses = {"draft", "under_review", "accepted"}
    if new_status not in valid_statuses:
        return JsonResponse(
            {"error": f"Invalid status. Must be one of: {', '.join(sorted(valid_statuses))}"}, status=400
        )

    from datetime import datetime

    data = _read_status(run_id, patient_id, ae_slug)

    # Validate transitions
    current = data.get("status", "draft")
    allowed_transitions = {
        "draft": {"accepted"},
        "accepted": {"draft"},
    }
    if new_status != current and new_status not in allowed_transitions.get(current, set()):
        return JsonResponse(
            {"error": f"Cannot transition from '{current}' to '{new_status}'"}, status=400
        )

    data["status"] = new_status
    data["updated_at"] = datetime.utcnow().isoformat()

    if new_status == "accepted":
        data["reviewed_by"] = reviewed_by or "Reviewer"
        data["reviewed_at"] = datetime.utcnow().isoformat()
    elif new_status == "draft":
        data["reviewed_by"] = None
        data["reviewed_at"] = None

    _write_status(run_id, patient_id, ae_slug, data)

    return JsonResponse({"success": True, **data})


def sae_report_editor(request, run_id: str, patient_id: str, ae_slug: str):
    """SAE Report Editor — MedWatch 3500A form with inline editing.

    Loads existing generated data if available, otherwise generates fresh.
    """
    run_path = _get_run_path(run_id)
    if not run_path.exists():
        return HttpResponse("Run not found", status=404)

    from src.doc_agent.service import DOCS_OUTPUT_DIR

    # Check for existing saved medwatch data
    json_path = DOCS_OUTPUT_DIR / run_id / patient_id / f"medwatch_data_{ae_slug}.json"
    medwatch_data = None
    meddra_data = None
    ai_used = False

    if json_path.exists():
        try:
            medwatch_data = json.loads(json_path.read_text(encoding="utf-8"))
        except Exception:
            pass

    if medwatch_data is None:
        # Generate fresh
        mode = request.GET.get("mode", "natural")
        profile, records = _load_patient_data(run_path, patient_id, mode)
        if profile is None:
            return HttpResponse("Patient not found", status=404)

        ae_term = ae_slug.replace("_", " ")
        ae_day_str = request.GET.get("ae_day", "").strip()
        ae_day = int(ae_day_str) if ae_day_str.isdigit() else None

        meta_path = run_path / "run_meta.json"
        drug_name = "Enfortumab vedotin (Padcev)"
        indication = "Metastatic urothelial carcinoma"
        if meta_path.exists():
            try:
                meta = json.loads(meta_path.read_text(encoding="utf-8"))
                drug_name = meta.get("drug_name", drug_name)
                indication = meta.get("indication", indication)
            except Exception:
                pass

        use_ai = request.GET.get("use_ai", "0") == "1"

        from datetime import date as dt_date
        from src.doc_agent.service import generate_documents

        result = generate_documents(
            patient_profile=profile,
            day_records=records,
            target_ae_term=ae_term,
            run_id=run_id,
            sim_start_date=dt_date(2026, 1, 6),
            drug_name=drug_name,
            indication=indication,
            target_ae_day=ae_day,
            use_ai=use_ai,
        )

        if result.get("success"):
            medwatch_data = result.get("medwatch_data", {})
            meddra_data = result.get("meddra")
            ai_used = result.get("ai_used", False)
            # Persist the medwatch data
            out_dir = DOCS_OUTPUT_DIR / run_id / patient_id
            out_dir.mkdir(parents=True, exist_ok=True)
            json_path.write_text(
                json.dumps(medwatch_data, indent=2, default=str, ensure_ascii=False),
                encoding="utf-8",
            )
        else:
            return render(request, "doc/sae_report.html", {
                "run_id": run_id,
                "patient_id": patient_id,
                "ae_slug": ae_slug,
                "error": result.get("error", "Unknown error"),
            })

    # Load profile for header info
    profile = _load_patient_profile(run_path, patient_id)
    demo = profile.get("emr", {}).get("demographics", {})
    rule_set = _load_rule_set(run_path)

    pdf_url = f"/api/doc/download/{run_id}/{patient_id}/medwatch_3500a_{ae_slug}.pdf"
    xml_url = f"/api/doc/download/{run_id}/{patient_id}/e2b_r3_{ae_slug}.xml"

    context = {
        "run_id": run_id,
        "patient_id": patient_id,
        "ae_slug": ae_slug,
        "ae_term": ae_slug.replace("_", " ").title(),
        "medwatch_json": json.dumps(medwatch_data, default=str, ensure_ascii=False),
        "meddra_json": json.dumps(meddra_data or {}, ensure_ascii=False),
        "ai_used": ai_used,
        "pdf_url": pdf_url,
        "xml_url": xml_url,
        "drug_name": rule_set.get("drug_name", ""),
        "indication": rule_set.get("indication", ""),
        "patient_age": demo.get("age", "?"),
        "patient_sex": demo.get("sex", "?"),
        "model_name": _load_run_meta(run_path).get("model", ""),
    }
    return render(request, "doc/sae_report.html", context)


def doc_hub(request, run_id: str):
    """Documents Hub — overview of all SAEs across patients, with links to report editor."""
    run_path = _get_run_path(run_id)
    if not run_path.exists():
        return HttpResponse("Run not found", status=404)

    mode = request.GET.get("mode", "natural")
    patient_ids = _list_patients(run_path)

    from src.doc_agent.sim_to_crf_adapter import find_serious_aes

    all_saes = []
    for pid in patient_ids:
        profile, records = _load_patient_data(run_path, pid, mode)
        if not records:
            continue
        saes = find_serious_aes(records)
        for sae in saes:
            ae = sae["ae_record"]
            ae_term = ae.get("AETERM", "")
            ae_slug = ae_term.replace(" ", "_").replace("/", "_")
            # Read report status
            report_status = _read_status(run_id, pid, ae_slug)
            all_saes.append({
                "patient_id": pid,
                "ae_term": ae_term,
                "ae_slug": ae_slug,
                "grade": ae.get("_grade", 0),
                "onset_day": sae["day"],
                "severity": ae.get("AESEV", ""),
                "action": ae.get("AEACN", ""),
                "serious": ae.get("AESER", False),
                "report_status": report_status.get("status", "draft"),
            })

    from src.doc_agent.service import DOCS_OUTPUT_DIR
    docs_dir = DOCS_OUTPUT_DIR / run_id
    existing_docs = set()
    if docs_dir.exists():
        for patient_dir in docs_dir.iterdir():
            if patient_dir.is_dir():
                for f in patient_dir.iterdir():
                    if f.suffix == ".pdf":
                        existing_docs.add(f"{patient_dir.name}/{f.stem}")

    meta = {}
    meta_path = run_path / "meta.json"
    if meta_path.exists():
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception:
            pass

    context = {
        "run_id": run_id,
        "mode": mode,
        "saes": all_saes,
        "sae_count": len(all_saes),
        "patient_count": len(set(s["patient_id"] for s in all_saes)),
        "existing_docs": existing_docs,
        "drug_name": meta.get("drug_name", ""),
        "indication": meta.get("indication", ""),
        "model_name": _load_run_meta(run_path).get("model", ""),
    }
    return render(request, "doc/doc_hub.html", context)


# ─── CRF Tables ─────────────────────────────────────────────────────────

import math
from .crf_aggregator import aggregate_domain, export_domain_to_excel, DOMAIN_COLUMNS, DOMAIN_LABELS

VALID_DOMAINS = {"dm", "mh", "ae", "ec", "cm", "vs", "lb", "ds", "dd", "tu", "rs", "pe", "eg"}


def crf_tables(request, run_id: str):
    """CRF Tables page — renders the main shell, data loaded via AJAX."""
    run_path = _get_run_path(run_id)
    if not run_path.exists():
        return JsonResponse({"error": "Run not found"}, status=404)

    patient_ids = _list_patients(run_path)
    sim_dir = run_path / "simulations"
    available_modes = []
    if sim_dir.exists():
        if list(sim_dir.glob("*_natural.jsonl")):
            available_modes.append("natural")
        if list(sim_dir.glob("*_care_ai.jsonl")):
            available_modes.append("care_ai")

    # Load run meta
    drug_name = ""
    indication = ""
    meta_path = run_path / "run_meta.json"
    if meta_path.exists():
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
            drug_name = meta.get("drug_name", "")
            indication = meta.get("indication", "")
        except Exception:
            pass

    rule_set = _load_rule_set(run_path)
    lab_ranges = rule_set.get("lab_reference_ranges", {})
    if not lab_ranges:
        lab_ranges = _extract_lab_ranges(run_path)
    context = {
        "run_id": run_id,
        "patient_ids": patient_ids,
        "patient_ids_json": json.dumps(patient_ids),
        "available_modes": available_modes,
        "drug_name": drug_name,
        "indication": indication,
        "domain_labels_json": json.dumps(DOMAIN_LABELS),
        "model_name": _load_run_meta(run_path).get("model", ""),
        "lab_ranges_json": json.dumps(lab_ranges),
    }
    return render(request, "doc/crf_tables.html", context)


def _get_sim_start_date(run_path: Path):
    """Read sim_start_date from run_meta.json, default 2026-01-06."""
    from datetime import date, timedelta
    meta_path = run_path / "run_meta.json"
    if meta_path.exists():
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
            d = meta.get("sim_start_date")
            if d:
                return date.fromisoformat(d)
        except Exception:
            pass
    return date(2026, 1, 6)


def _inject_dates(rows, columns, start_date):
    """Replace day-number columns with calendar date columns in CRF rows."""
    from datetime import timedelta

    # Map of day-number keys → date label
    DAY_KEY_LABELS = {
        "day": "Date",
        "AESTDAT": "Start Date", "AEENDAT": "End Date",
        "ECSTDAT": "Start Date", "ECENDAT": "End Date",
        "CMSTDAT": "Start Date", "CMENDAT": "End Date",
        "MHSTDAT": "Start Date", "MHENDAT": "End Date",
        "onset_day": "Onset Date", "detected_day": "Detected Date",
        "PEDAT": "Exam Date", "EGDAT": "ECG Date",
        "TUDAT": "Assessment Date",
        "DSSTDAT": "Disposition Date", "DTHDAT": "Date of Death",
    }

    # Find which day-number columns exist
    day_col_keys = [col["key"] for col in columns if col["key"] in DAY_KEY_LABELS]
    if not day_col_keys:
        return rows, columns

    # Replace day columns with date columns
    new_columns = []
    for col in columns:
        if col["key"] in DAY_KEY_LABELS:
            new_columns.append({
                "key": col["key"] + "_date",
                "label": DAY_KEY_LABELS[col["key"]],
            })
        else:
            new_columns.append(col)

    # Convert day numbers to date strings
    for row in rows:
        for k in day_col_keys:
            v = row.pop(k, None)
            if isinstance(v, (int, float)) and v > 0:
                row[k + "_date"] = str(start_date + timedelta(days=int(v) - 1))
            else:
                row[k + "_date"] = None

    return rows, new_columns


@require_GET
def api_crf_domain_data(request, run_id: str, domain: str):
    """JSON API: return domain-specific CRF rows with pagination."""
    if domain.lower() not in VALID_DOMAINS:
        return JsonResponse({"error": f"Invalid domain: {domain}"}, status=400)

    run_path = _get_run_path(run_id)
    if not run_path.exists():
        return JsonResponse({"error": "Run not found"}, status=404)

    source = request.GET.get("source", "hr")
    mode = request.GET.get("mode", "natural")
    patient_filter = request.GET.get("patient", "")
    page = int(request.GET.get("page", 1))
    per_page = int(request.GET.get("per_page", 100))

    patient_ids = None
    if patient_filter:
        patient_ids = [p.strip() for p in patient_filter.split(",") if p.strip()]

    rows, total, columns = aggregate_domain(
        domain, run_path, patient_ids, mode, source, page, per_page,
    )

    # Inject calendar dates next to day-number columns
    start_date = _get_sim_start_date(run_path)
    rows, columns = _inject_dates(rows, columns, start_date)

    total_pages = math.ceil(total / per_page) if per_page > 0 else 1

    return JsonResponse({
        "domain": domain.upper(),
        "source": source,
        "mode": mode,
        "columns": columns,
        "rows": rows,
        "total_rows": total,
        "page": page,
        "per_page": per_page,
        "total_pages": total_pages,
    })


@require_GET
def api_crf_excel_download(request, run_id: str):
    """Download CRF data as Excel file."""
    run_path = _get_run_path(run_id)
    if not run_path.exists():
        return JsonResponse({"error": "Run not found"}, status=404)

    source = request.GET.get("source", "hr")
    mode = request.GET.get("mode", "natural")
    domains_param = request.GET.get("domains", "")

    if domains_param:
        domains = [d.strip().upper() for d in domains_param.split(",") if d.strip()]
    else:
        domains = [d.upper() for d in VALID_DOMAINS]

    # Multi-domain export: one sheet per domain
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    start_date = _get_sim_start_date(run_path)

    for dom in domains:
        if dom.lower() not in VALID_DOMAINS:
            continue
        rows, total, columns = aggregate_domain(
            dom, run_path, None, mode, source, page=1, per_page=0,
        )
        rows, columns = _inject_dates(rows, columns, start_date)
        ws = wb.create_sheet(title=dom)
        col_keys = [c["key"] for c in columns]

        # Headers
        for ci, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=ci, value=col["label"])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data
        for ri, row in enumerate(rows, 2):
            for ci, key in enumerate(col_keys, 1):
                val = row.get(key)
                if isinstance(val, bool):
                    val = "Yes" if val else "No"
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = thin_border

        # Auto-width
        for ci, col in enumerate(columns, 1):
            max_len = len(col["label"])
            for ri in range(2, min(len(rows) + 2, 102)):  # sample first 100 rows
                val = ws.cell(row=ri, column=ci).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    response = HttpResponse(
        xlsx_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="CRF_{run_id}_{source}.xlsx"'
    return response


_mm_bridges = {}


@csrf_exempt
@require_POST
def api_multimodal_enhance(request):
    """Generate face image + voice audio for a care_record turn on demand.

    POST JSON: {run_id, patient_id, day, turn_index (0-based, patient turns only)}
    Returns:   {face_b64, audio_b64, mm_meta} or {error}
    """
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    run_id = body.get("run_id", "")
    patient_id = body.get("patient_id", "")
    day = int(body.get("day", 0))
    turn_index = int(body.get("turn_index", 0))

    if not run_id or not patient_id or day < 1:
        return JsonResponse({"error": "run_id, patient_id, day required"}, status=400)

    run_path = _get_run_path(run_id)
    if not run_path:
        return JsonResponse({"error": f"Run not found: {run_id}"}, status=404)

    # Load patient profile
    patient_file = run_path / "patients" / f"{patient_id}.json"
    if not patient_file.exists():
        return JsonResponse({"error": f"Patient not found: {patient_id}"}, status=404)

    with open(patient_file) as f:
        patient_json = json.load(f)

    # Load day data from care_ai JSONL
    sim_dir = run_path / "simulations"
    care_file = sim_dir / f"{patient_id}_care_ai.jsonl"
    if not care_file.exists():
        return JsonResponse({"error": "No care_ai data for this patient"}, status=404)

    day_data = None
    with open(care_file) as f:
        for line in f:
            d = json.loads(line)
            if d.get("day") == day:
                day_data = d
                break

    if not day_data:
        return JsonResponse({"error": f"Day {day} not found"}, status=404)

    care_records = day_data.get("care_record", [])
    if not care_records:
        return JsonResponse({"error": "No care_record for this day"}, status=404)

    cr = care_records[0] if isinstance(care_records, list) else care_records
    turns = cr.get("turns", [])

    patient_turns = [t for t in turns if t.get("role") == "patient"]
    if turn_index >= len(patient_turns):
        return JsonResponse({"error": f"turn_index {turn_index} out of range"}, status=400)

    # Extract text from patient turn
    turn = patient_turns[turn_index]
    content = turn.get("content", {})
    text_parts = []
    if isinstance(content, str):
        text_parts.append(content)
    else:
        if g := content.get("greeting"):
            text_parts.append(g)
        if wb := content.get("general_wellbeing"):
            text_parts.append(wb)
        for sym in content.get("reported_symptoms", []):
            if isinstance(sym, dict) and sym.get("verbal_expression"):
                text_parts.append(sym["verbal_expression"])
        for resp in content.get("responses", []):
            if a := resp.get("answer"):
                text_parts.append(a)
    text = " ".join(text_parts) if text_parts else "I'm not feeling great today."

    active_aes = day_data.get("AE", [])
    mood_snapshot = cr.get("mood_snapshot", {})

    # Get or create bridge (cached per patient within run)
    bridge_key = f"{run_id}:{patient_id}"
    try:
        from src.multimodal.game_bridge import MultimodalGameBridge

        if bridge_key not in _mm_bridges:
            _mm_bridges[bridge_key] = MultimodalGameBridge(patient_json, enabled=True)

        bridge = _mm_bridges[bridge_key]
        media = bridge.generate_turn_media(
            text=text,
            active_aes=active_aes,
            day=day,
            mood_snapshot=mood_snapshot,
        )
        return JsonResponse({
            "face_b64": media.get("face_b64"),
            "audio_b64": media.get("audio_b64"),
            "mm_meta": media.get("mm_meta", {}),
            "text": text,
            "day": day,
            "patient_id": patient_id,
        })

    except ImportError as e:
        return JsonResponse({"error": f"Multimodal module not available: {e}"}, status=500)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({"error": f"Generation failed: {e}"}, status=500)

# ═══════════════════════════════════════════════════════════════
# Statistical Analysis (CSR Tables & Charts)
# ═══════════════════════════════════════════════════════════════

def statistical_analysis(request, run_id: str):
    """Statistical Analysis page — CSR-style tables and charts."""
    run_path = _get_run_path(run_id)
    if not run_path.exists():
        return HttpResponse("Run not found", status=404)

    rule_set = _load_rule_set(run_path)
    meta = _load_run_meta(run_path)
    drug_name = rule_set.get("drug_name") or meta.get("drug_name", "Unknown")
    indication = rule_set.get("indication") or meta.get("indication", "")

    sim_dir = run_path / "simulations"
    available_modes = []
    if sim_dir.exists():
        if list(sim_dir.glob("*_natural.jsonl")):
            available_modes.append("natural")
        if list(sim_dir.glob("*_care_ai.jsonl")):
            available_modes.append("care_ai")

    n_patients = len(_list_patients(run_path))

    return render(request, "doc/statistical_analysis.html", {
        "run_id": run_id,
        "drug_name": drug_name,
        "indication": indication,
        "n_patients": n_patients,
        "available_modes": available_modes,
        "model_name": meta.get("model", ""),
    })


@require_GET
def api_stats_data(request, run_id: str):
    """JSON API: compute and return all CSR statistics."""
    run_path = _get_run_path(run_id)
    if not run_path.exists():
        return JsonResponse({"error": "Run not found"}, status=404)

    mode = request.GET.get("mode", "natural")

    cache_path = run_path / "validation" / f"csr_stats_{mode}.json"
    sim_dir = run_path / "simulations"
    needs_compute = True
    if cache_path.exists():
        sim_files = list(sim_dir.glob("*.jsonl")) if sim_dir.exists() else []
        if sim_files:
            newest_sim = max(f.stat().st_mtime for f in sim_files)
            if cache_path.stat().st_mtime >= newest_sim:
                needs_compute = False

    if needs_compute:
        try:
            import sys as _sys
            _proj_root = str(Path(settings.BASE_DIR).parent)
            if _proj_root not in _sys.path:
                _sys.path.insert(0, _proj_root)
            from validation.csr_stats import compute_csr_stats
            stats = compute_csr_stats(str(run_path), mode)
            cache_path.parent.mkdir(parents=True, exist_ok=True)
            with open(cache_path, "w") as f:
                json.dump(stats, f, ensure_ascii=False)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({"error": str(e)}, status=500)
    else:
        with open(cache_path) as f:
            stats = json.load(f)

    return JsonResponse(stats, json_dumps_params={"ensure_ascii": False})


# ─── Stats Chatbot ──────────────────────────────────────────

_STATS_CHAT_URL = os.environ.get("CTE_VLLM_BASE_URL", "").rstrip("/")
_STATS_CHAT_MODEL = os.environ.get("CTE_VLLM_MODEL_ID", "medgemma-4b-antihallu")


def _sanitize_messages(messages):
    """Ensure roles alternate user/assistant after system. Drop consecutive same-role messages."""
    if not messages:
        return messages
    out = []
    for msg in messages:
        role = msg.get("role", "user")
        if role == "system":
            out.append(msg)
            continue
        if out and out[-1].get("role") == role:
            # merge into previous to avoid consecutive same-role
            out[-1] = {**out[-1], "content": out[-1]["content"] + "\n" + msg.get("content", "")}
        else:
            out.append(msg)
    # vLLM requires last message to be user
    if out and out[-1].get("role") != "user" and out[-1].get("role") != "system":
        out.append({"role": "user", "content": "(continue)"})
    return out


def _call_chat_llm(messages, query_meta):
    """Call vLLM for chat completions. Returns JsonResponse."""
    import re as _re

    if not _STATS_CHAT_URL:
        return JsonResponse({"error": "vLLM not configured (CTE_VLLM_BASE_URL)"}, status=500)

    messages = _sanitize_messages(messages)

    payload = {
        "model": _STATS_CHAT_MODEL,
        "messages": messages,
        "max_tokens": 512,
        "temperature": 0.2,
        "repetition_penalty": 1.15,
    }
    body_bytes = json.dumps(payload).encode("utf-8")
    req = Request(
        f"{_STATS_CHAT_URL}/chat/completions",
        data=body_bytes,
        headers={"Content-Type": "application/json", "Authorization": "Bearer EMPTY"},
        method="POST",
    )
    try:
        t0 = time.time()
        with urlopen(req, timeout=120) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        latency_ms = round((time.time() - t0) * 1000)
        answer = data["choices"][0]["message"]["content"]
        answer = _re.sub(r'<unused\d+>.*?<unused\d+>', '', answer, flags=_re.DOTALL).strip()
        query_meta["backend"] = "vllm"
        query_meta["model"] = _STATS_CHAT_MODEL
        return JsonResponse({
            "response": answer,
            "latency_ms": latency_ms,
            "query": query_meta,
        })
    except Exception as exc:
        detail = str(exc)
        if hasattr(exc, "read"):
            try:
                detail = exc.read().decode("utf-8", errors="replace")
            except Exception:
                pass
        logging.warning("Chat LLM call failed: %s — %s", exc, detail)
        return JsonResponse({"error": f"LLM call failed: {detail}"}, status=500)

_STATS_SYSTEM_PROMPT = (
    "You are CLARA's Statistical Analysis Assistant, an expert clinical trial biostatistician.\n"
    "You help researchers interpret CSR (Clinical Study Report) statistical results.\n\n"
    "STRICT RULES — violating any rule is a critical failure:\n"
    "1. Answer based ONLY on the provided data. NEVER fabricate or infer numbers.\n"
    "2. When quoting a number, copy it EXACTLY from the data. Do NOT round, combine, or paraphrase.\n"
    "3. DISTINGUISH between 'all-grade' and 'Grade 3+' (G3+) columns carefully.\n"
    "   - 'all=56.0%' means all-grade incidence is 56%.\n"
    "   - 'G3+=2.0%' means Grade 3+ incidence is 2%.\n"
    "   - These are DIFFERENT numbers. NEVER use the all-grade number when asked about G3+.\n"
    "4. Answer ONLY the question asked. Do NOT dump unrelated sections.\n"
    "5. Be concise: 2-5 sentences unless the user asks for detail.\n"
    "6. If the data does not contain the answer, say so. Do NOT guess.\n"
    "7. Use the same language as the user.\n"
    "8. When user references data with @[...] tags, focus on that specific data point.\n"
)


def _compact_stats(stats: dict, tab: str) -> str:
    """Legacy: tab-based extraction. Use _retrieve_context() instead."""
    return _retrieve_context(stats, "", tab)


# ─── Keyword → section mapping for lightweight RAG ───
_SECTION_KEYWORDS = {
    "demographics": [
        "age", "sex", "gender", "race", "ethnicity", "weight", "height", "bmi",
        "demographic", "population", "patient characteristics", "baseline",
        "나이", "성별", "인종", "체중", "환자 특성",
    ],
    "efficacy": [
        "orr", "dcr", "response", "tumor", "waterfall", "survival", "pfs", "os",
        "progression", "recist", "cr", "pr", "sd", "pd", "best response",
        "time to response", "ttr", "duration of response", "dor",
        "반응률", "종양", "생존", "효능",
    ],
    "safety": [
        "ae", "adverse", "toxicity", "side effect", "safety", "grade",
        "sae", "serious", "fatal",
        "이상반응", "부작용", "독성", "안전",
    ],
    "safety_detail": [
        "fatigue", "nausea", "diarrhea", "rash", "alopecia", "neuropathy",
        "stomatitis", "pruritus", "hyperglycemia", "anemia", "pneumonitis",
        "appetite", "infusion", "vomiting", "constipation", "pain",
    ],
    "treatment": [
        "dose", "rdi", "interruption", "reduction", "modification", "discontinu",
        "cycle", "duration", "administration", "drug",
        "투여", "용량", "중단",
    ],
    "labs": [
        "lab", "glucose", "hemoglobin", "platelet", "creatinine", "alt", "ast",
        "bilirubin", "albumin", "sodium", "anc", "neutrophil", "blood",
        "검사", "혈액",
    ],
    "ecog": [
        "ecog", "performance status", "functional", "ps",
        "수행능력",
    ],
    "conmeds": [
        "concomitant", "medication", "conmed", "supportive", "steroid",
        "병용약",
    ],
    "disposition": [
        "disposition", "enrolled", "completed", "discontinued", "death", "dropout",
        "withdrawal",
        "등록", "완료", "중단", "사망",
    ],
}


def _match_sections(message: str, tab: str) -> list:
    """Return list of matched section names based on message keywords + active tab."""
    msg_lower = message.lower()
    scores = {}
    for section, keywords in _SECTION_KEYWORDS.items():
        score = sum(1 for kw in keywords if kw in msg_lower)
        if score > 0:
            scores[section] = score

    # Always include the active tab
    tab_to_section = {
        "safety": "safety", "demographics": "demographics", "efficacy": "efficacy",
        "labs": "labs", "ecog": "ecog", "conmeds": "conmeds",
        "treatment": "treatment", "disposition": "disposition",
    }
    if tab in tab_to_section:
        sec = tab_to_section[tab]
        scores[sec] = scores.get(sec, 0) + 2  # boost active tab

    # If safety_detail matched, ensure safety is also included
    if "safety_detail" in scores:
        scores["safety"] = scores.get("safety", 0) + scores["safety_detail"]

    # If nothing matched, default to safety (most common)
    if not scores:
        scores["safety"] = 1

    # Sort by score descending, return top sections
    return [s for s, _ in sorted(scores.items(), key=lambda x: -x[1]) if s != "safety_detail"]


def _extract_section(stats: dict, section: str, message: str) -> list:
    """Extract formatted lines for a given section."""
    lines = []
    msg_lower = message.lower()

    if section == "disposition":
        disp = stats.get("disposition", {})
        lines.append(f"[Disposition] Enrolled: {disp.get('enrolled',0)}, "
                     f"Completed: {disp.get('completed',{}).get('n',0)} ({disp.get('completed',{}).get('pct',0)}%), "
                     f"Discontinued: {disp.get('discontinued',{}).get('n',0)}, "
                     f"Deaths: {disp.get('deaths',{}).get('n',0)}")
        reasons = disp.get("reasons", {})
        if reasons:
            parts = [f"{k}: {v}" for k, v in reasons.items() if v]
            if parts:
                lines.append(f"  Reasons: {', '.join(parts)}")

    elif section == "demographics":
        demo = stats.get("demographics", {})
        age = demo.get("age", {})
        if age:
            lines.append(f"[Demographics] Age: mean={age.get('mean','?')} SD={age.get('std','?')}, "
                         f"median={age.get('median','?')}, range={age.get('min','?')}-{age.get('max','?')}")
        for cat_key in ("sex", "race", "ecog"):
            cat = demo.get(cat_key, {})
            if cat:
                parts = [f"{k}={v.get('n',0)}({v.get('pct',0)}%)" for k, v in cat.items()]
                lines.append(f"  {cat_key.title()}: {', '.join(parts)}")
        # BMI if asked
        bmi = demo.get("bmi", {})
        if bmi and any(kw in msg_lower for kw in ("bmi", "weight", "체중")):
            lines.append(f"  BMI: mean={bmi.get('mean','?')} SD={bmi.get('std','?')}, "
                         f"range={bmi.get('min','?')}-{bmi.get('max','?')}")

    elif section == "efficacy":
        eff = stats.get("efficacy", {})
        orr = eff.get("orr", {})
        dcr = eff.get("dcr", {})
        km_os = eff.get("km_os", {})
        km_pfs = eff.get("km_pfs", {})
        lines.append(f"[Efficacy] ORR: {orr.get('pct',0)}% (n={orr.get('n',0)}, "
                     f"CI: {orr.get('ci','?')}), DCR: {dcr.get('pct',0)}%")
        lines.append(f"  Median OS: {km_os.get('median','NR')} days, "
                     f"Median PFS: {km_pfs.get('median','NR')} days")
        br = eff.get("best_response", {})
        if br:
            parts = [f"{k}={v.get('n',0)}({v.get('pct',0)}%)" for k, v in br.items()]
            lines.append(f"  Best response: {', '.join(parts)}")
        ttr = eff.get("time_to_response", {})
        if ttr.get("n"):
            lines.append(f"  TTR: median={ttr.get('median','?')} days (n={ttr['n']})")
        dor = eff.get("dor", {})
        if dor.get("n"):
            lines.append(f"  DoR: median={dor.get('median','NR')} days, events={dor.get('events',0)}")
        # Waterfall individual data if asked
        wf = eff.get("waterfall", [])
        if wf and any(kw in msg_lower for kw in ("waterfall", "tumor change", "pd ", "pr ", "cr ", "종양")):
            lines.append("  Waterfall (per patient):")
            for w in wf:
                lines.append(f"    {w.get('pid','?')}: {w.get('change',0)}% ({w.get('response','?')})")

    elif section == "safety":
        safe = stats.get("safety", {})
        sm = safe.get("summary", {})
        lines.append(f"[Safety Summary]")
        lines.append(f"  Any AE: {sm.get('any_ae',{}).get('pct',0)}% (n={sm.get('any_ae',{}).get('n',0)})")
        lines.append(f"  Grade>=3 AE: {sm.get('grade_gte3',{}).get('pct',0)}% (n={sm.get('grade_gte3',{}).get('n',0)})")
        lines.append(f"  SAE (Serious): {sm.get('sae',{}).get('pct',0)}% (n={sm.get('sae',{}).get('n',0)})")
        lines.append(f"  Fatal AE: {sm.get('fatal',{}).get('pct',0)}% (n={sm.get('fatal',{}).get('n',0)})")
        lines.append(f"  Led to discontinuation: {sm.get('led_to_discont',{}).get('pct',0)}% (n={sm.get('led_to_discont',{}).get('n',0)})")
        lines.append(f"  Led to interruption: {sm.get('led_to_interrupt',{}).get('pct',0)}% (n={sm.get('led_to_interrupt',{}).get('n',0)})")
        by_term = safe.get("by_term", [])
        # Check if user asks about a specific AE
        specific_aes = [ae for ae in by_term
                        if ae["term"].replace("_", " ") in msg_lower
                        or ae["term"].replace("_", "") in msg_lower.replace(" ", "")]
        if specific_aes:
            lines.append("  AE table (all-grade% ≠ G3+%, do NOT confuse):")
            for ae in specific_aes:
                gd = ae.get("grade_dist", {})
                gd_str = ", ".join(f"G{g}={n}" for g, n in sorted(gd.items()) if int(n) > 0)
                lines.append(f"    {ae['term']}: ALL-GRADE={ae['all_grade']['pct']}% (n={ae['all_grade']['n']}), "
                             f"GRADE3+={ae['grade_gte3']['pct']}% (n={ae['grade_gte3']['n']}), "
                             f"onset=Day {ae.get('onset_median','?')} "
                             f"(IQR: {ae.get('onset_iqr','?')}), grades: {gd_str}")
        else:
            # Top 10 AEs summary
            lines.append("  AE table (all-grade% ≠ G3+%, do NOT confuse):")
            lines.append("    TERM | ALL-GRADE% | GRADE3+% | ONSET")
            for ae in by_term[:10]:
                lines.append(f"    {ae['term']} | {ae['all_grade']['pct']}% | "
                             f"{ae['grade_gte3']['pct']}% | Day {ae.get('onset_median','?')}")

    elif section == "treatment":
        tx = stats.get("treatment", {})
        n_total = stats.get("n_patients", stats.get("n_simulated", "?"))
        dr = tx.get('dose_reduction_all', {})
        di = tx.get('dose_interruption_all', {})
        dc = tx.get('discontinuation_all', {})
        lines.append(f"[Treatment] N={n_total}")
        lines.append(f"  Duration: median={tx.get('duration',{}).get('median','?')} days")
        lines.append(f"  Cycles: median={tx.get('cycles',{}).get('median','?')}")
        lines.append(f"  Dose REDUCTION: {dr.get('n',0)} patients = {dr.get('pct',0)}%")
        lines.append(f"  Dose INTERRUPTION: {di.get('n',0)} patients = {di.get('pct',0)}%")
        lines.append(f"  Discontinuation: {dc.get('n',0)} patients = {dc.get('pct',0)}%")
        # Per-drug detail
        per_drug = tx.get("per_drug", {})
        for drug_name, drug_data in per_drug.items():
            admins = drug_data.get("n_admins", {})
            rdi = drug_data.get("rdi_median", "?")
            lines.append(f"  {drug_name}: admins median={admins.get('median','?')} "
                         f"(range {admins.get('min','?')}-{admins.get('max','?')}), RDI={rdi}%")

    elif section == "labs":
        abn = stats.get("labs", {}).get("abnormalities", {})
        if abn:
            lines.append("[Labs] Abnormalities:")
            for test, v in abn.items():
                lines.append(f"  {test}: any={v.get('any_pct',0)}%, G3+={v.get('g3_pct',0)}%")

    elif section == "ecog":
        ec = stats.get("ecog_shift", {})
        sm_ec = ec.get("summary", {})
        if sm_ec:
            lines.append(f"[ECOG Shift] Improved: {sm_ec.get('improved',{}).get('pct',0)}%, "
                         f"Stable: {sm_ec.get('stable',{}).get('pct',0)}%, "
                         f"Worsened: {sm_ec.get('worsened',{}).get('pct',0)}%")
        # Individual shifts
        patients = ec.get("patients", [])
        if patients and any(kw in msg_lower for kw in ("shift", "individual", "patient", "detail")):
            for p in patients[:10]:
                lines.append(f"  {p.get('pid','?')}: {p.get('baseline',0)} → {p.get('worst',0)} → {p.get('last',0)}")

    elif section == "conmeds":
        cm = stats.get("concomitant_meds", {})
        tbl = cm.get("table", [])
        if tbl:
            lines.append(f"[Concomitant Meds] {cm.get('total_unique_meds',0)} unique medications:")
            for m in tbl[:10]:
                lines.append(f"  {m['medication']}: {m['n']} ({m['pct']}%)")

    return lines


def _retrieve_context(stats: dict, message: str, tab: str) -> str:
    """Keyword-based retrieval: select relevant sections from stats based on message content.
    Target: <2500 chars to fit within 4096 token model context.
    """
    sections = _match_sections(message, tab)

    lines = []
    total_chars = 0
    char_budget = 2400

    for section in sections:
        section_lines = _extract_section(stats, section, message)
        section_text = "\n".join(section_lines)
        if total_chars + len(section_text) > char_budget and lines:
            break  # budget exceeded, stop adding sections
        lines.extend(section_lines)
        total_chars += len(section_text) + 1

    return "\n".join(lines)


@csrf_exempt
def api_stats_chat(request, run_id: str):
    """Chat API for statistical analysis — answers questions using vLLM."""
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=405)

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    message = body.get("message", "").strip()
    if not message:
        return JsonResponse({"error": "Empty message"}, status=400)

    mode = body.get("mode", "natural")
    tab = body.get("tab", "")
    history = body.get("history", [])

    # Load cached stats
    run_path = _get_run_path(run_id)
    if not run_path.exists():
        return JsonResponse({"error": "Run not found"}, status=404)

    cache_path = run_path / "validation" / f"csr_stats_{mode}.json"
    if not cache_path.exists():
        return JsonResponse({"error": "Stats not computed yet. Load the stats page first."}, status=400)

    with open(cache_path) as f:
        stats = json.load(f)

    # Load run metadata
    rule_set = _load_rule_set(run_path)
    meta = _load_run_meta(run_path)
    drug_name = rule_set.get("drug_name") or meta.get("drug_name", "Unknown")
    indication = rule_set.get("indication") or meta.get("indication", "")
    n_patients = len(_list_patients(run_path))

    # Keyword-based retrieval — select relevant sections from stats
    matched_sections = _match_sections(message, tab)
    compact = _retrieve_context(stats, message, tab)
    context_block = (
        f"Drug: {drug_name} | Indication: {indication} | "
        f"Mode: {mode} | N={n_patients}\n\n{compact}"
    )

    # Put context in system message
    context_msg = _STATS_SYSTEM_PROMPT + "\n---\nData:\n" + context_block

    # Build messages for vLLM — keep history minimal
    messages = [{"role": "system", "content": context_msg}]
    for msg in history[-4:]:  # last 2 turns
        role = msg.get("role", "user")
        if role == "model":
            role = "assistant"
        messages.append({"role": role, "content": msg.get("content", "")})
    messages.append({"role": "user", "content": message})

    # Build query metadata for debug view
    query_meta = {
        "source": f"csr_stats_{mode}.json",
        "model": _STATS_CHAT_MODEL,
        "matched_sections": matched_sections,
        "tab": tab or "(none)",
        "context_data": compact,
        "history_turns": len(history) // 2,
        "message": message,
    }

    # Call LLM (vLLM or Gemini fallback)
    return _call_chat_llm(messages, query_meta)


@csrf_exempt
def api_stats_chat_demo(request):
    """Demo chat API — auto-selects the latest run with computed stats."""
    runs_dir = DATA_DIR / "runs"
    if not runs_dir.exists():
        return JsonResponse({"error": "No simulation runs found. Please run a simulation first."}, status=404)

    # Find latest run that has validation stats
    for d in sorted(runs_dir.iterdir(), reverse=True):
        if d.is_dir():
            for mode in ("natural", "care_ai"):
                cache = d / "validation" / f"csr_stats_{mode}.json"
                if cache.exists():
                    try:
                        return api_stats_chat(request, d.name)
                    except Exception as exc:
                        logging.warning("Stats chat demo failed for run %s: %s", d.name, exc)
                        return JsonResponse(
                            {"error": f"Stats chat failed for run '{d.name}': {exc}. "
                             "The statistics data may be incomplete or corrupted. "
                             "Try re-running the simulation or computing stats again."},
                            status=500,
                        )

    return JsonResponse(
        {"error": "No runs with computed stats found. "
         "Run a simulation first, then navigate to the Statistical Analysis page "
         "to compute the stats before using the chat."},
        status=404,
    )


# ─── Unified Doc Chat API ────────────────────────────────────

_CRF_SYSTEM_PROMPT = (
    "You are CLARA's CRF Data Assistant, an expert clinical data manager.\n"
    "You help researchers explore CRF (Case Report Form) tabular data.\n\n"
    "Rules:\n"
    "- Answer based ONLY on the provided CRF data below.\n"
    "- NEVER fabricate data not present in the provided rows.\n"
    "- The 'Pre-computed Summary' contains EXACT counts from ALL rows. ALWAYS trust summary numbers over counting the visible table rows (the table may be truncated).\n"
    "- If the data shows 0 matching rows, say 0 — do NOT invent a result.\n"
    "- Be concise (2-5 sentences) unless asked for detail.\n"
    "- Reference specific patient IDs, values, and counts from the data.\n"
    "- Use the same language as the user.\n"
    "- When user references data with @[...] tags, focus on that specific record.\n"
)

_SAE_SYSTEM_PROMPT = (
    "You are CLARA's SAE Report Assistant, an expert in pharmacovigilance and MedWatch reporting.\n"
    "You help researchers analyze Serious Adverse Event reports.\n\n"
    "Rules:\n"
    "- Answer based ONLY on the provided MedWatch/SAE data below.\n"
    "- NEVER fabricate information not in the data.\n"
    "- Be concise (2-5 sentences) unless asked for detail.\n"
    "- Reference specific form sections, dates, and clinical details.\n"
    "- Use the same language as the user.\n"
    "- When user references data with @[...] tags, focus on that specific field.\n"
)


def _build_stats_chat_context(run_path, body, drug_name, indication, n_patients, message):
    """Build context for stats page chat."""
    mode = body.get("mode", "natural")
    tab = body.get("tab", "")
    cache_path = run_path / "validation" / f"csr_stats_{mode}.json"
    if not cache_path.exists():
        return JsonResponse({"error": "Stats not computed yet."}, status=400), None, None
    with open(cache_path) as f:
        stats = json.load(f)
    matched_sections = _match_sections(message, tab)
    compact = _retrieve_context(stats, message, tab)
    context_block = (
        f"Drug: {drug_name} | Indication: {indication} | "
        f"Mode: {mode} | N={n_patients}\n\n{compact}"
    )
    query_meta = {
        "source": f"csr_stats_{mode}.json",
        "model": _STATS_CHAT_MODEL,
        "matched_sections": matched_sections,
        "tab": tab or "(none)",
        "context_data": compact,
        "history_turns": len(body.get("history", [])) // 2,
        "message": message,
    }
    return context_block, _STATS_SYSTEM_PROMPT, query_meta


def _build_crf_chat_context(run_path, body, drug_name, indication, n_patients, message):
    """Build context for CRF tables page chat."""
    domain = body.get("domain", "ae")
    mode = body.get("mode", "natural")
    patient = body.get("patient", "")

    patient_ids = [patient] if patient else None
    try:
        rows, total, columns = aggregate_domain(
            domain, run_path, patient_ids, mode, "hr", 1, 20)
    except Exception:
        rows, total, columns = [], 0, []

    lines = [f"Drug: {drug_name} | Indication: {indication} | Mode: {mode} | N={n_patients}"]
    lines.append(f"Domain: {domain.upper()} | Total rows: {total}")
    lines.append(f"Columns: {', '.join(c.get('label', c.get('key', '')) for c in columns[:8])}")
    lines.append("")

    # Pick columns that matter most for each domain (include Grade for AE)
    _PRIORITY_KEYS = {"_grade", "AESEV", "AESER", "AEREL", "AEACN", "LBSTRESN", "LBSTNRHI", "LBSTNRLO"}
    key_cols = []
    for c in columns[:6]:
        key_cols.append(c)
    for c in columns[6:]:
        if c.get("key", "") in _PRIORITY_KEYS and len(key_cols) < 10:
            key_cols.append(c)

    # Pre-computed summary to prevent hallucination on counts
    if domain.lower() == "ae" and rows:
        from collections import Counter
        all_rows, _, _ = aggregate_domain(domain, run_path, patient_ids, mode, "hr", 1, 500)
        grade_dist = Counter()
        serious_count = 0
        ae_per_pt = Counter()
        for r in all_rows:
            g = r.get("_grade", 0)
            try:
                g = int(g)
            except (ValueError, TypeError):
                g = 0
            grade_dist[g] += 1
            if r.get("AESER") in (True, "True", "Y", "YES"):
                serious_count += 1
            ae_per_pt[r.get("patient_id", "")] += 1
        lines.append("=== Pre-computed Summary (AUTHORITATIVE — always use these counts, ignore the truncated table below if they differ) ===")
        lines.append(f"Total AEs: {len(all_rows)}")
        for g in sorted(grade_dist.keys()):
            # List which patients/AEs for non-G1 grades
            if g >= 2:
                g_rows = [r for r in all_rows if int(r.get("_grade", 0)) == g]
                detail = "; ".join(f'{r.get("patient_id")} {r.get("AETERM","")}' for r in g_rows)
                lines.append(f"  Grade {g} AEs: {grade_dist[g]} ({detail})")
            else:
                lines.append(f"  Grade {g} AEs: {grade_dist[g]}")
        lines.append(f"  Grade 3+ AEs: {sum(n for g, n in grade_dist.items() if g >= 3)}")
        lines.append(f"Serious AEs: {serious_count}")
        lines.append(f"AEs per patient: {', '.join(f'{p}={n}' for p, n in ae_per_pt.most_common())}")
        lines.append(f"NOTE: The table below shows only the first 20 of {len(all_rows)} rows. The summary above covers ALL rows.")
        lines.append("")

    if rows:
        header = " | ".join(c.get("label", c.get("key", "")) for c in key_cols)
        lines.append(header)
        lines.append("-" * len(header))
        for row in rows[:20]:
            vals = []
            for c in key_cols:
                v = row.get(c.get("key", ""), "")
                vals.append(str(v)[:30])
            lines.append(" | ".join(vals))

    context_block = "\n".join(lines)
    if len(context_block) > 2400:
        context_block = context_block[:2400] + "\n...(truncated)"

    query_meta = {
        "source": f"CRF/{domain.upper()}",
        "model": _STATS_CHAT_MODEL,
        "matched_sections": [domain.upper()],
        "tab": domain,
        "context_data": context_block,
        "history_turns": len(body.get("history", [])) // 2,
        "message": message,
    }
    return context_block, _CRF_SYSTEM_PROMPT, query_meta


def _build_sae_chat_context(run_path, body, drug_name, indication, n_patients, message):
    """Build context for SAE report page chat."""
    patient_id = body.get("patient_id", "")
    ae_slug = body.get("ae_slug", "")

    from src.doc_agent.service import DOCS_OUTPUT_DIR
    json_path = DOCS_OUTPUT_DIR / run_path.name / patient_id / f"medwatch_data_{ae_slug}.json"

    if not json_path.exists():
        return JsonResponse({"error": "SAE data not found. Generate the report first."}, status=400), None, None

    try:
        mw = json.loads(json_path.read_text(encoding="utf-8"))
    except Exception as e:
        return JsonResponse({"error": f"Failed to load SAE data: {e}"}, status=500), None, None

    lines = [f"Drug: {drug_name} | Indication: {indication} | Patient: {patient_id}"]
    lines.append(f"AE: {ae_slug.replace('_', ' ')}")
    lines.append("")

    # Section A: Patient Info
    a = mw.get("section_a", mw.get("A", {}))
    if a:
        lines.append("=== Section A: Patient ===")
        for k in ["age", "sex", "weight", "ethnicity"]:
            if k in a:
                lines.append(f"  {k}: {a[k]}")

    # Section B: Adverse Event
    b = mw.get("section_b", mw.get("B", {}))
    if b:
        lines.append("=== Section B: Adverse Event ===")
        for k in ["event_description", "onset_date", "outcome", "narrative"]:
            v = b.get(k, "")
            if v:
                lines.append(f"  {k}: {str(v)[:400]}")
        # Extract seriousness criteria from individual boolean fields
        serious = [k.replace("seriousness_", "") for k, v in b.items()
                   if k.startswith("seriousness_") and v]
        if serious:
            lines.append(f"  serious_criteria: {', '.join(serious)}")
        elif b.get("serious_criteria"):
            lines.append(f"  serious_criteria: {b['serious_criteria']}")

    # Section C: Suspect Product
    c = mw.get("section_c", mw.get("C", {}))
    if c:
        lines.append("=== Section C: Suspect Product ===")
        for k in ["product_name", "drug_name", "dose", "dose_frequency_route", "route",
                   "indication", "start_date", "therapy_start", "stop_date", "therapy_end",
                   "dechallenge", "rechallenge", "concomitant_meds"]:
            v = c.get(k, "")
            if v:
                lines.append(f"  {k}: {str(v)[:300]}")

    # MedDRA coding
    meddra = mw.get("meddra", mw.get("MedDRA", {}))
    if meddra:
        lines.append("=== MedDRA Coding ===")
        for k, v in meddra.items():
            if v:
                lines.append(f"  {k}: {v}")

    context_block = "\n".join(lines)
    if len(context_block) > 2400:
        context_block = context_block[:2400] + "\n...(truncated)"

    query_meta = {
        "source": f"medwatch_data_{ae_slug}.json",
        "model": _STATS_CHAT_MODEL,
        "matched_sections": ["A", "B", "C", "MedDRA"],
        "tab": ae_slug,
        "context_data": context_block,
        "history_turns": len(body.get("history", [])) // 2,
        "message": message,
    }
    return context_block, _SAE_SYSTEM_PROMPT, query_meta


def _build_sae_hub_chat_context(run_path, body, drug_name, indication, n_patients, message):
    """Build context for SAE hub listing page chat — summarises all SAEs."""
    from src.doc_agent.sim_to_crf_adapter import find_serious_aes

    mode = body.get("mode", "natural")
    patient_ids = _list_patients(run_path)

    lines = [f"Drug: {drug_name} | Indication: {indication} | Patients: {n_patients}"]
    lines.append("")
    lines.append("=== All Serious Adverse Events ===")

    sae_rows = []
    for pid in patient_ids:
        profile, records = _load_patient_data(run_path, pid, mode)
        if not records:
            continue
        saes = find_serious_aes(records)
        for sae in saes:
            ae = sae["ae_record"]
            sae_rows.append({
                "patient": pid,
                "term": ae.get("AETERM", ""),
                "grade": ae.get("_grade", 0),
                "day": sae["day"],
                "action": ae.get("AEACN", ""),
                "serious": ae.get("AESER", False),
            })

    if not sae_rows:
        lines.append("No serious adverse events found in this run.")
    else:
        # Pre-computed summary so LLM never needs to count rows
        from collections import Counter
        sae_per_patient = Counter(r["patient"] for r in sae_rows)
        term_counts = Counter(r["term"] for r in sae_rows)
        grade_counts = Counter(r["grade"] for r in sae_rows)
        action_counts = Counter(r["action"] or "NONE" for r in sae_rows)
        onset_days = sorted(r["day"] for r in sae_rows)

        lines.append("=== Pre-computed Summary (use these numbers, do NOT re-count) ===")
        lines.append(f"Total SAEs: {len(sae_rows)}")
        lines.append(f"Affected patients: {len(sae_per_patient)} — {', '.join(f'{p}={n} SAEs' for p, n in sae_per_patient.most_common())}")
        lines.append(f"SAE terms: {', '.join(f'{t}={n}' for t, n in term_counts.most_common())}")
        lines.append(f"Grade distribution: {', '.join(f'G{g}={n}' for g, n in sorted(grade_counts.items()))}")
        lines.append(f"Actions: {', '.join(f'{a}={n}' for a, n in action_counts.most_common())}")
        lines.append(f"Onset range: Day {onset_days[0]} – Day {onset_days[-1]} (first SAE: Day {onset_days[0]})")
        lines.append(f"Fatal SAEs: 0")
        lines.append("")
        lines.append("=== Full SAE Table ===")
        lines.append("Patient | AE Term | Grade | Onset | Action")
        lines.append("--------|---------|-------|-------|-------")
        for r in sae_rows[:40]:
            lines.append(f"{r['patient']} | {r['term']} | G{r['grade']} | Day {r['day']} | {r['action'] or '—'}")

    context_block = "\n".join(lines)
    # Truncate to ~2.4KB
    if len(context_block) > 2400:
        context_block = context_block[:2400] + "\n... (truncated)"

    system_prompt = (
        "You are CLARA's SAE Overview Assistant, an expert in pharmacovigilance.\n"
        "You help researchers analyze the overall SAE profile of a clinical trial run.\n\n"
        "Rules:\n"
        "- Answer based ONLY on the provided SAE listing data below.\n"
        "- NEVER fabricate information not in the data.\n"
        "- The 'Pre-computed Summary' section contains exact counts. ALWAYS use those numbers instead of counting rows yourself.\n"
        "- Be concise (2-5 sentences) unless asked for detail.\n"
        "- Reference specific patients, AE terms, grades, and onset days.\n"
        "- Use the same language as the user.\n"
        "- When user references data with @[...] tags, focus on that specific row.\n"
    )

    query_meta = {
        "page_type": "sae_hub",
        "mode": mode,
        "sae_count": len(sae_rows),
        "message": message,
    }
    return context_block, system_prompt, query_meta


def _build_compare_chat_context(run_path, body, drug_name, indication, n_patients, message):
    """Build context for A/B Comparison page chat — summarises comparison_report.json."""
    report_path = run_path / "comparison_report.json"
    if not report_path.exists():
        return (
            JsonResponse({"error": "comparison_report.json not found"}, status=404),
            None,
            None,
        )

    with open(report_path) as f:
        report = json.load(f)

    lines = [
        f"Drug: {drug_name} | Indication: {indication} | Patients: {n_patients}",
        "",
        "=== A/B Comparison: Natural vs Care AI ===",
        "=== Pre-computed Summary (AUTHORITATIVE — use these numbers, do NOT re-count) ===",
    ]

    # Cohort sizes
    cohort = report.get("cohort_sizes", {})
    lines.append(f"Cohort: Natural={cohort.get('natural', '?')}, Care AI={cohort.get('care_ai', '?')}")
    lines.append("")

    # Detection Delay
    dd = report.get("detection_delay", {})
    deltas = report.get("deltas", {})
    lines.append(
        f"Detection Delay: Natural={dd.get('natural_mean', '?')}d, "
        f"Care AI={dd.get('care_ai_mean', '?')}d "
        f"(\u0394={deltas.get('detection_delay', '?')}d) "
        f"[Undetected: Natural={dd.get('natural_undetected', '?')}, Care AI={dd.get('care_ai_undetected', '?')}]"
    )

    # AE Burden
    ab = report.get("ae_burden", {})
    lines.append(
        f"AE Burden (grade\u00d7days): Natural={ab.get('natural_mean', '?')}, "
        f"Care AI={ab.get('care_ai_mean', '?')} "
        f"[Unique AEs: Natural={ab.get('natural_unique_aes', '?')}, Care AI={ab.get('care_ai_unique_aes', '?')}]"
    )

    # Severe AEs
    sa = report.get("severe_aes", {})
    lines.append(
        f"Grade 3+ AE Days: Natural={sa.get('natural_g3plus_mean', '?')}, "
        f"Care AI={sa.get('care_ai_g3plus_mean', '?')} | "
        f"Grade 4+: Natural={sa.get('natural_g4plus_mean', '?')}, "
        f"Care AI={sa.get('care_ai_g4plus_mean', '?')}"
    )

    # Treatment Duration
    td = report.get("treatment_duration", {})
    lines.append(
        f"Treatment Duration: Natural={td.get('natural_mean', '?')}d, "
        f"Care AI={td.get('care_ai_mean', '?')}d "
        f"(\u0394={deltas.get('treatment_duration', '?')}d)"
    )

    # ECOG
    ecog = report.get("ecog", {})
    lines.append(
        f"ECOG Change: Natural=+{ecog.get('natural_mean_delta', '?')}, "
        f"Care AI=+{ecog.get('care_ai_mean_delta', '?')} "
        f"[End ECOG: Natural={ecog.get('natural_mean_end', '?')}, "
        f"Care AI={ecog.get('care_ai_mean_end', '?')}]"
    )

    # Discontinuation
    disc = report.get("discontinuation", {})
    lines.append(
        f"Discontinued: Natural={disc.get('natural_count', '?')}/{cohort.get('natural', '?')} "
        f"({disc.get('natural_pct', '?')}%), "
        f"Care AI={disc.get('care_ai_count', '?')}/{cohort.get('care_ai', '?')} "
        f"({disc.get('care_ai_pct', '?')}%)"
    )

    # Mortality
    mort = report.get("mortality", {})
    lines.append(
        f"Deaths: Natural={mort.get('natural_deaths', '?')}, "
        f"Care AI={mort.get('care_ai_deaths', '?')}"
    )

    # Care AI Activity
    ca = report.get("care_ai_activity", {})
    if ca:
        lines.append("")
        lines.append("Care AI Activity:")
        lines.append(f"  Mean interventions/patient: {ca.get('mean_interventions', '?')}")
        lines.append(f"  Mean AE detections/patient: {ca.get('mean_detections', '?')}")
        lines.append(f"  Mean turns/call: {ca.get('mean_turns_per_call', '?')}")
        lines.append(f"  Early terminations: {ca.get('total_early_terminations', '?')}")
        lines.append(f"  Force hospital visits: {ca.get('total_force_hospital', '?')}")
        itypes = ca.get("intervention_type_totals", {})
        if itypes:
            lines.append(f"  Intervention types: {', '.join(f'{k}={v}' for k, v in itypes.items())}")

    # Statistical Tests
    stats = report.get("statistics", {})
    if stats:
        lines.append("")
        lines.append("Statistical Tests:")
        for test_name, test_data in stats.items():
            if isinstance(test_data, dict) and "p_value" in test_data:
                p = test_data["p_value"]
                sig = "sig" if isinstance(p, (int, float)) and p < 0.05 else "ns"
                stat_val = test_data.get("statistic", "?")
                n_val = test_data.get("n", "?")
                lines.append(f"  {test_name}: W={stat_val}, p={p} ({sig}, n={n_val})")

    # Pre-computed per-patient analysis (NO raw table — model can't parse it reliably)
    nat_pts = report.get("natural_patients", [])
    cai_pts = report.get("care_ai_patients", [])
    if nat_pts and cai_pts:
        lines.append("")
        lines.append("=== Per-Patient Analysis (pre-computed) ===")

        for label, pts in [("Natural", nat_pts), ("CareAI", cai_pts)]:
            burdens = [p.get("total_ae_burden", 0) for p in pts]
            delays = [p.get("mean_detection_delay", 0) for p in pts]
            g3ds = [p.get("grade3plus_ae_days", 0) for p in pts]
            worst_b = max(pts, key=lambda p: p.get("total_ae_burden", 0))
            best_b = min(pts, key=lambda p: p.get("total_ae_burden", 0))
            deceased = [p["patient_id"] for p in pts if p.get("deceased")]
            disc = [p["patient_id"] for p in pts if p.get("discontinued")]
            g3_pts = [f'{p["patient_id"]}={p.get("grade3plus_ae_days", 0)}d' for p in pts if p.get("grade3plus_ae_days", 0) > 0]
            lines.append(
                f"  {label}: burden range {min(burdens)}-{max(burdens)}, "
                f"worst={worst_b['patient_id']}({worst_b.get('total_ae_burden', 0)}), "
                f"best={best_b['patient_id']}({best_b.get('total_ae_burden', 0)})"
            )
            lines.append(
                f"    delay range {min(delays)}-{max(delays)}d, "
                f"G3+ patients: {', '.join(g3_pts) if g3_pts else 'none'}"
            )
            if deceased or disc:
                lines.append(
                    f"    deceased: {', '.join(deceased) if deceased else 'none'}, "
                    f"discontinued: {', '.join(disc) if disc else 'none'}"
                )

        # Paired comparison: per-patient burden change (sorted by delta)
        nat_by_pid = {p["patient_id"]: p for p in nat_pts}
        cai_by_pid = {p["patient_id"]: p for p in cai_pts}
        common_pids = sorted(set(nat_by_pid) & set(cai_by_pid))
        if common_pids:
            pairs = []
            for pid in common_pids:
                nb = nat_by_pid[pid].get("total_ae_burden", 0)
                cb = cai_by_pid[pid].get("total_ae_burden", 0)
                pairs.append((pid, nb, cb, cb - nb))
            improved = [(pid, nb, cb, d) for pid, nb, cb, d in pairs if d < 0]
            worsened = [(pid, nb, cb, d) for pid, nb, cb, d in pairs if d > 0]
            improved.sort(key=lambda x: x[3])  # most improved first (most negative)
            lines.append(f"  Burden improved with CareAI: {len(improved)}/{len(common_pids)} patients")
            # Show sorted by improvement magnitude
            imp_strs = [f"{pid}({nb}\u2192{cb}, \u0394{d})" for pid, nb, cb, d in improved]
            if imp_strs:
                lines.append(f"    {', '.join(imp_strs)}")
            if improved:
                big = improved[0]
                lines.append(f"  Biggest improvement: {big[0]} (burden {big[1]}\u2192{big[2]}, reduced by {abs(big[3])})")
            if worsened:
                w_strs = [f"{pid}({nb}\u2192{cb}, +{d})" for pid, nb, cb, d in worsened]
                lines.append(f"  Burden worsened: {len(worsened)} — {', '.join(w_strs)}")

    context_block = "\n".join(lines)
    # Truncate to ~2.4KB
    if len(context_block) > 2400:
        context_block = context_block[:2400] + "\n... (truncated)"

    system_prompt = (
        "You are CLARA's A/B Comparison Assistant, an expert in clinical trial analysis.\n"
        "You help researchers analyze Natural vs Care AI simulation results.\n\n"
        "Rules:\n"
        "- Answer based ONLY on the provided comparison data below.\n"
        "- NEVER fabricate information not in the data.\n"
        "- ALL numbers are pre-computed. Quote them directly — do NOT attempt to calculate, count, or find max/min yourself.\n"
        "- Be concise (2-5 sentences) unless asked for detail.\n"
        "- Highlight statistically significant differences (p < 0.05) when relevant.\n"
        "- When discussing Care AI value, focus on detection delay reduction, AE burden, and patient outcomes.\n"
        "- Reference specific metrics, patient IDs, and statistical test results.\n"
        "- Use the same language as the user.\n"
    )

    query_meta = {
        "page_type": "compare",
        "natural_n": cohort.get("natural", 0),
        "care_ai_n": cohort.get("care_ai", 0),
        "message": message,
    }
    return context_block, system_prompt, query_meta


@csrf_exempt
def api_doc_chat(request, run_id: str):
    """Unified chat API — routes to page-specific context builders."""
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=405)

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    message = body.get("message", "").strip()
    if not message:
        return JsonResponse({"error": "Empty message"}, status=400)

    page_type = body.get("page_type", "stats")
    history = body.get("history", [])

    run_path = _get_run_path(run_id)
    if not run_path.exists():
        return JsonResponse({"error": "Run not found"}, status=404)

    # Load common metadata
    rule_set = _load_rule_set(run_path)
    meta = _load_run_meta(run_path)
    drug_name = rule_set.get("drug_name") or meta.get("drug_name", "Unknown")
    indication = rule_set.get("indication") or meta.get("indication", "")
    n_patients = len(_list_patients(run_path))

    # Route to context builder
    if page_type == "stats":
        context_block, system_prompt, query_meta = _build_stats_chat_context(
            run_path, body, drug_name, indication, n_patients, message)
    elif page_type == "crf":
        context_block, system_prompt, query_meta = _build_crf_chat_context(
            run_path, body, drug_name, indication, n_patients, message)
    elif page_type == "sae":
        context_block, system_prompt, query_meta = _build_sae_chat_context(
            run_path, body, drug_name, indication, n_patients, message)
    elif page_type == "sae_hub":
        context_block, system_prompt, query_meta = _build_sae_hub_chat_context(
            run_path, body, drug_name, indication, n_patients, message)
    elif page_type == "compare":
        context_block, system_prompt, query_meta = _build_compare_chat_context(
            run_path, body, drug_name, indication, n_patients, message)
    else:
        return JsonResponse({"error": f"Unknown page_type: {page_type}"}, status=400)

    if isinstance(context_block, JsonResponse):
        return context_block  # Error response from builder

    # Build LLM messages
    full_system = system_prompt + "\n---\nData:\n" + context_block
    messages = [{"role": "system", "content": full_system}]
    for msg in history[-4:]:
        role = msg.get("role", "user")
        if role == "model":
            role = "assistant"
        messages.append({"role": role, "content": msg.get("content", "")})
    messages.append({"role": "user", "content": message})

    # Call LLM (vLLM or Gemini fallback)
    return _call_chat_llm(messages, query_meta)


# ─── Rule Set Generation: GT vs Predicted Comparison ────────
_RULESET_DIR = Path(settings.BASE_DIR).parent / "src" / "ruleset_generation"

_GT_TO_OUTPUT = {
    "1_Darbepoetin_alfa": "darbepoetin_alfa_small_cell_lung_cancer",
    "2_Etoposide_Cisplatin": "etoposide+cisplatin_small_cell_lung_cancer",
    "3_CALGB9732_Paclitaxel_Cisplatin_Etoposide": "paclitaxel+cisplatin+etoposide_small_cell_lung_cancer",
    "4_Carboplatin_Etoposide": "etoposide+carboplatin_small_cell_lung_cancer",
    "6_Paclitaxel_Carboplatin_Bevacizumab": "paclitaxel+carboplatin+bevacizumab_non-small_cell_lung_cancer",
    "7_Paclitaxel_Carboplatin": "paclitaxel+carboplatin_non-small_cell_lung_cancer",
    "8_Gemcitabine_Cisplatin": "gemcitabine+cisplatin_squamous_non-small_cell_lung_cancer",
}


def _load_json_safe(path):
    try:
        with open(path) as f:
            return json.load(f)
    except Exception:
        return None


def _normalize_ae(term):
    import re
    t = term.lower().strip().replace(" ", "_").replace("-", "_")
    t = re.sub(r"_s$", "", t)
    _syn = {
        "diarrhoea": "diarrhea", "dyspnoea": "dyspnea", "anaemia": "anemia",
        "haemoglobin_decreased": "lower_hemoglobin", "paraesthesia": "paresthesia",
        "hyponatraemia": "hyponatremia", "oedema": "edema", "leucopenia": "leukopenia",
    }
    return _syn.get(t, t)


def _clamp01(x):
    return max(0.0, min(1.0, x))


def _score_ruleset(out, gt):
    """Score predicted vs GT across 9 dimensions. Returns list of {dim, score, detail}."""
    import re

    def _parse_dose(s):
        m = re.search(r"([\d.]+)", str(s or ""))
        return float(m.group(1)) if m else None

    results = []

    # 1. Doses
    out_sched = out.get("administration_schedule", [])
    gt_sched = gt.get("administration_schedule", [])
    out_d = {e.get("drug_name", "").lower().strip(): _parse_dose(e.get("dose_per_administration")) for e in out_sched}
    gt_d = {e.get("drug_name", "").lower().strip(): _parse_dose(e.get("dose_per_administration")) for e in gt_sched}
    if gt_d:
        sc = []
        for drug, gd in gt_d.items():
            if drug in out_d and out_d[drug] and gd:
                r = min(out_d[drug], gd) / max(out_d[drug], gd)
                sc.append(1.0 if r >= 0.9 else r)
            else:
                sc.append(0.0)
        results.append({"dim": "Doses", "score": sum(sc) / len(sc), "detail": f"out={out_d} gt={gt_d}"})
    else:
        results.append({"dim": "Doses", "score": 0.0, "detail": "no GT doses"})

    # 2. ORR
    out_orr = out.get("efficacy", {}).get("overall_response_rate")
    gt_orr = gt.get("efficacy", {}).get("overall_response_rate")
    if gt_orr and out_orr is not None:
        results.append({"dim": "ORR", "score": _clamp01(1.0 - abs(out_orr - gt_orr) / gt_orr), "detail": f"out={out_orr:.3f} gt={gt_orr:.3f}"})
    else:
        results.append({"dim": "ORR", "score": 0.0, "detail": f"out={out_orr} gt={gt_orr}"})

    # 3. Age Range
    oa = out.get("demographics", {}).get("age", {}).get("params", {})
    ga = gt.get("demographics", {}).get("age", {}).get("params", {})
    om, ox, gm, gx = oa.get("min"), oa.get("max"), ga.get("min"), ga.get("max")
    if gm is not None and gx is not None and gx != gm and om is not None and ox is not None:
        results.append({"dim": "Age Range", "score": _clamp01(1.0 - (abs(om - gm) + abs(ox - gx)) / (gx - gm)), "detail": f"out={om}-{ox} gt={gm}-{gx}"})
    else:
        results.append({"dim": "Age Range", "score": 0.0, "detail": f"out={om}-{ox} gt={gm}-{gx}"})

    # 4. Sex Ratio
    os_sex = out.get("demographics", {}).get("sex", {}).get("options", {})
    gs_sex = gt.get("demographics", {}).get("sex", {}).get("options", {})
    om_v = os_sex.get("Male", os_sex.get("M", 0))
    gm_v = gs_sex.get("Male", gs_sex.get("M", 0))
    if isinstance(om_v, dict):
        om_v = om_v.get("probability", 0)
    if isinstance(gm_v, dict):
        gm_v = gm_v.get("probability", 0)
    results.append({"dim": "Sex Ratio", "score": _clamp01(1.0 - abs(float(om_v) - float(gm_v))), "detail": f"out_M={om_v} gt_M={gm_v}"})

    # 5. ECOG
    oe = out.get("demographics", {}).get("ecog_ps", {}).get("options", {})
    ge = gt.get("demographics", {}).get("ecog_ps", {}).get("options", {})
    all_k = set(list(oe.keys()) + list(ge.keys()))
    td = sum(abs(float(oe.get(k, 0)) - float(ge.get(k, 0))) for k in all_k) if all_k else 0
    results.append({"dim": "ECOG", "score": _clamp01(1.0 - td / 2.0), "detail": f"out={dict(oe)} gt={dict(ge)}"})

    # 6. AE Count
    on = len(out.get("ae_profile", []))
    gn = len(gt.get("ae_profile", []))
    mx = max(on, gn)
    results.append({"dim": "AE Count", "score": _clamp01(1.0 - abs(on - gn) / mx) if mx else 1.0, "detail": f"out={on} gt={gn}"})

    # 7. AE Freq
    oa_map = {_normalize_ae(ae.get("ae_term", "")): ae.get("incidence_all_grade", 0) for ae in out.get("ae_profile", [])}
    ga_map = {_normalize_ae(ae.get("ae_term", "")): ae.get("incidence_all_grade", 0) for ae in gt.get("ae_profile", [])}
    common = set(oa_map) & set(ga_map)
    if common:
        fsc = []
        for t in common:
            m = max(oa_map[t], ga_map[t])
            fsc.append(1.0 if m == 0 else 1.0 - abs(oa_map[t] - ga_map[t]) / m)
        results.append({"dim": "AE Freq", "score": sum(fsc) / len(fsc), "detail": f"{len(common)} common AEs"})
    else:
        results.append({"dim": "AE Freq", "score": 0.0, "detail": "no common AEs"})

    # 8. Top AE overlap
    oae_s = sorted([(k, v) for k, v in oa_map.items()], key=lambda x: -x[1])
    gae_s = sorted([(k, v) for k, v in ga_map.items()], key=lambda x: -x[1])
    ot10 = set(t for t, _ in oae_s[:10])
    gt10 = set(t for t, _ in gae_s[:10])
    overlap = ot10 & gt10
    results.append({"dim": "Top AE", "score": len(overlap) / 10.0, "detail": f"overlap={sorted(overlap)}"})

    # 9. PFS/OS
    oe_eff = out.get("efficacy", {})
    ge_eff = gt.get("efficacy", {})
    surv_sc = []
    for key in ["progression_free_survival_months", "overall_survival_months"]:
        ov = oe_eff.get(key, {}).get("params", {}).get("mean")
        gv = ge_eff.get(key, {}).get("params", {}).get("mean")
        if gv and gv > 0 and ov is not None:
            surv_sc.append(_clamp01(1.0 - abs(ov - gv) / gv))
        elif gv is None and ov is None:
            surv_sc.append(1.0)
        else:
            surv_sc.append(0.0)
    results.append({"dim": "PFS/OS", "score": sum(surv_sc) / len(surv_sc) if surv_sc else 0.0,
                     "detail": f"PFS: out={oe_eff.get('progression_free_survival_months',{}).get('params',{}).get('mean')} gt={ge_eff.get('progression_free_survival_months',{}).get('params',{}).get('mean')}"})

    return results


def api_ruleset_drugs(request):
    """List available drugs with GT + predicted data."""
    gt_dir = _RULESET_DIR / "ground_truth"
    out_dir = _RULESET_DIR / "output"
    drugs = []
    for gt_folder, out_folder in _GT_TO_OUTPUT.items():
        gt_path = gt_dir / gt_folder / "base.json"
        out_path = out_dir / out_folder / "base.json"
        gt_data = _load_json_safe(gt_path)
        out_data = _load_json_safe(out_path)
        if gt_data:
            drugs.append({
                "id": gt_folder,
                "drug_name": gt_data.get("drug_name", gt_folder),
                "indication": gt_data.get("indication", ""),
                "has_gt": True,
                "has_predicted": out_data is not None,
            })
    return JsonResponse({"drugs": drugs})


def api_ruleset_compare(request, drug_id):
    """Compare GT vs Predicted for a specific drug."""
    gt_dir = _RULESET_DIR / "ground_truth"
    out_dir = _RULESET_DIR / "output"
    out_folder = _GT_TO_OUTPUT.get(drug_id)
    if not out_folder:
        return JsonResponse({"error": f"Unknown drug: {drug_id}"}, status=404)

    gt_data = _load_json_safe(gt_dir / drug_id / "base.json")
    out_data = _load_json_safe(out_dir / out_folder / "base.json")

    if not gt_data:
        return JsonResponse({"error": "GT data not found"}, status=404)
    if not out_data:
        return JsonResponse({"error": "Predicted data not found"}, status=404)

    scores = _score_ruleset(out_data, gt_data)
    avg = sum(s["score"] for s in scores) / len(scores) if scores else 0

    gt_aes = sorted(
        [{"term": ae.get("ae_term", ""), "incidence": ae.get("incidence_all_grade", 0)}
         for ae in gt_data.get("ae_profile", [])],
        key=lambda x: -x["incidence"]
    )[:15]
    pred_aes = sorted(
        [{"term": ae.get("ae_term", ""), "incidence": ae.get("incidence_all_grade", 0)}
         for ae in out_data.get("ae_profile", [])],
        key=lambda x: -x["incidence"]
    )[:15]

    def _extract_fields(data):
        demo = data.get("demographics", {})
        age_raw = demo.get("age", {})
        sex_raw = demo.get("sex", {})
        eff = data.get("efficacy", {})
        sched = data.get("administration_schedule", [])
        doses = []
        for s in sched:
            doses.append({
                "drug": s.get("drug_name", ""),
                "dose": s.get("dose_per_administration", ""),
                "route": s.get("route", ""),
                "schedule_days": s.get("schedule_days", []),
            })
        ecog_raw = demo.get("ecog_ps", {})

        age_params = age_raw.get("params", age_raw) if isinstance(age_raw, dict) else {}
        sex_opts = sex_raw.get("options", sex_raw) if isinstance(sex_raw, dict) else {}
        ecog_opts = ecog_raw.get("options", ecog_raw) if isinstance(ecog_raw, dict) else {}

        pct_male = sex_opts.get("pct_male") or sex_opts.get("Male")
        pct_female = sex_opts.get("pct_female") or sex_opts.get("Female")
        if pct_male and pct_male <= 1:
            pct_male = pct_male * 100
        if pct_female and pct_female <= 1:
            pct_female = pct_female * 100

        pfs = eff.get("progression_free_survival") or eff.get("progression_free_survival_months", {})
        os_data = eff.get("overall_survival") or eff.get("overall_survival_months", {})

        return {
            "n_aes": len(data.get("ae_profile", [])),
            "orr": eff.get("overall_response_rate"),
            "cycle_days": data.get("trial_design", {}).get("cycle_length_days"),
            "doses": doses,
            "age_mean": age_params.get("mean"),
            "age_std": age_params.get("std"),
            "age_min": age_params.get("min"),
            "age_max": age_params.get("max"),
            "pct_male": pct_male,
            "pct_female": pct_female,
            "ecog": ecog_opts if isinstance(ecog_opts, dict) else {},
            "pfs_median": pfs.get("median") if isinstance(pfs, dict) else None,
            "os_median": os_data.get("median") if isinstance(os_data, dict) else None,
        }

    gt_fields = _extract_fields(gt_data)
    gt_fields["top_aes"] = gt_aes
    pred_fields = _extract_fields(out_data)
    pred_fields["top_aes"] = pred_aes

    return JsonResponse({
        "drug_id": drug_id,
        "drug_name": gt_data.get("drug_name", drug_id),
        "indication": gt_data.get("indication", ""),
        "scores": scores,
        "average_score": round(avg, 3),
        "gt_summary": gt_fields,
        "pred_summary": pred_fields,
    }, json_dumps_params={"ensure_ascii": False})


def api_ruleset_compare_all(request):
    """Compare all 7 drugs at once — summary table."""
    gt_dir = _RULESET_DIR / "ground_truth"
    out_dir = _RULESET_DIR / "output"
    rows = []
    for gt_folder, out_folder in _GT_TO_OUTPUT.items():
        gt_data = _load_json_safe(gt_dir / gt_folder / "base.json")
        out_data = _load_json_safe(out_dir / out_folder / "base.json")
        if not gt_data or not out_data:
            continue
        scores = _score_ruleset(out_data, gt_data)
        avg = sum(s["score"] for s in scores) / len(scores) if scores else 0
        rows.append({
            "drug_id": gt_folder,
            "drug_name": gt_data.get("drug_name", gt_folder),
            "indication": gt_data.get("indication", ""),
            "scores": {s["dim"]: round(s["score"], 3) for s in scores},
            "average": round(avg, 3),
        })
    overall_avg = sum(r["average"] for r in rows) / len(rows) if rows else 0
    dims = list(rows[0]["scores"].keys()) if rows else []
    return JsonResponse({
        "drugs": rows,
        "dimensions": dims,
        "overall_average": round(overall_avg, 3),
    })


def demo_ruleset_generation(request):
    """Rule Set Generation demo page."""
    return render(request, "demo/ruleset_generation.html")


_ruleset_gen_jobs: dict = {}


@csrf_exempt
@require_POST
def api_ruleset_generate(request):
    """Start rule set generation for a custom drug."""
    try:
        body = json.loads(request.body)
    except Exception:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    drug_name = body.get("drug_name", "").strip()
    indication = body.get("indication", "").strip()
    user_api_key = body.get("api_key", "").strip()
    if not drug_name:
        return JsonResponse({"error": "drug_name is required"}, status=400)

    job_id = f"{drug_name}_{indication}".replace(" ", "_")[:60]

    if job_id in _ruleset_gen_jobs and _ruleset_gen_jobs[job_id].get("status") == "running":
        return JsonResponse({
            "job_id": job_id,
            "status": "running",
            "message": "Generation already in progress",
        })

    _ruleset_gen_jobs[job_id] = {"status": "running", "progress": "Starting..."}

    def _run():
        import sys as _sys
        _sys.path.insert(0, str(Path(settings.BASE_DIR).parent / "src" / "ruleset_generation"))
        _sys.path.insert(0, str(Path(settings.BASE_DIR).parent))

        # 사용자 API 키가 있으면 우선 사용, 없으면 .env 폴백
        if user_api_key:
            os.environ["GOOGLE_API_KEY"] = user_api_key
            os.environ["RULE_ENGINE_LLM_API_KEY"] = user_api_key
        else:
            env_path = Path(settings.BASE_DIR).parent / ".env"
            if env_path.exists():
                for line in env_path.read_text().splitlines():
                    line = line.strip()
                    if line and not line.startswith("#") and "=" in line:
                        key, val = line.split("=", 1)
                        os.environ.setdefault(key.strip(), val.strip())

        try:
            _ruleset_gen_jobs[job_id]["progress"] = "Collecting evidence from 10 databases..."
            import asyncio
            from rule_engine.config import RuleEngineConfig
            from rule_engine.evidence.collector import collect_evidence
            from rule_engine.agent import synthesize_rules

            gkey = os.environ.get("GOOGLE_API_KEY", "")
            if gkey and not os.environ.get("RULE_ENGINE_LLM_API_KEY"):
                os.environ["RULE_ENGINE_LLM_API_KEY"] = gkey

            drugs = [d.strip() for d in drug_name.split("+")]
            config = RuleEngineConfig()

            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)

            _ruleset_gen_jobs[job_id]["progress"] = "Collecting evidence..."
            evidence = loop.run_until_complete(collect_evidence(drugs, indication, config))

            _ruleset_gen_jobs[job_id]["progress"] = "LLM synthesis (multi-stage)..."
            rule_set, agent_log = loop.run_until_complete(
                synthesize_rules(drugs, indication, evidence, config)
            )
            loop.close()

            from rule_engine.converter import convert_ruleset, split_base_overlay
            internal = json.loads(rule_set.model_dump_json())
            converted, schema_type = convert_ruleset(internal)
            base_dict, overlay_dict = split_base_overlay(converted, schema_type)

            _ruleset_gen_jobs[job_id] = {
                "status": "completed",
                "progress": "Done!",
                "result": base_dict,
                "schema_type": schema_type,
            }
        except Exception as e:
            _ruleset_gen_jobs[job_id] = {
                "status": "error",
                "progress": f"Failed: {e}",
                "error": str(e),
            }

    import threading
    t = threading.Thread(target=_run, daemon=True)
    t.start()

    return JsonResponse({"job_id": job_id, "status": "running", "message": "Generation started"})


def api_ruleset_generate_status(request, job_id):
    """Check rule set generation status."""
    job = _ruleset_gen_jobs.get(job_id)
    if not job:
        return JsonResponse({"error": "Job not found"}, status=404)
    resp = {"job_id": job_id, "status": job["status"], "progress": job.get("progress", "")}
    if job["status"] == "completed" and "result" in job:
        resp["result"] = job["result"]
    if job["status"] == "error":
        resp["error"] = job.get("error", "")
    return JsonResponse(resp, json_dumps_params={"ensure_ascii": False})


# ─── Demo API (auto-select latest run) ───────────────────────────────────


def _get_latest_run_id() -> str | None:
    """Return the run_id of the latest completed run (sorted reverse by name)."""
    runs_dir = DATA_DIR / "runs"
    if not runs_dir.exists():
        return None
    for d in sorted(runs_dir.iterdir(), reverse=True):
        if d.is_dir() and (d / "simulations").exists():
            return d.name
    return None


@csrf_exempt
@require_GET
def api_demo_saes(request):
    """Demo SAE list — auto-selects the latest run, returns all SAEs across
    all patients.

    GET /api/demo/saes/?mode=natural
    """
    mode = request.GET.get("mode", "natural")
    run_id = _get_latest_run_id()
    if not run_id:
        return JsonResponse(
            {"error": "No simulation runs found. Run a simulation first."},
            status=404,
        )

    run_path = _get_run_path(run_id)
    patient_ids = _list_patients(run_path)
    if not patient_ids:
        return JsonResponse(
            {"error": f"No patients found in run '{run_id}'."},
            status=404,
        )

    from src.doc_agent.sim_to_crf_adapter import find_serious_aes

    all_saes = []
    for pid in patient_ids:
        profile, records = _load_patient_data(run_path, pid, mode)
        if not records:
            continue
        try:
            saes = find_serious_aes(records)
        except Exception:
            continue
        for sae in saes:
            ae = sae["ae_record"]
            all_saes.append({
                "patient_id": pid,
                "ae_term": ae.get("AETERM", ""),
                "grade": ae.get("_grade", 0),
                "onset_day": sae.get("day") or ae.get("AESTDAT"),
                "severity": ae.get("AESEV", ""),
                "action": ae.get("AEACN", ""),
                "outcome": ae.get("AEOUT", ""),
                "mode": mode,
            })

    return JsonResponse({"run_id": run_id, "saes": all_saes})


@csrf_exempt
@require_POST
def api_demo_generate(request):
    """Demo report generation — auto-selects the latest run.

    POST body: {
        "patient_id": str,
        "ae_term": str,
        "ae_day": int (optional),
        "mode": "natural" | "care_ai" (default: "natural"),
        "use_ai": bool (default: false),
    }
    """
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    patient_id = body.get("patient_id", "")
    ae_term = body.get("ae_term", "")
    ae_day = body.get("ae_day")
    mode = body.get("mode", "natural")
    use_ai = body.get("use_ai", False)

    if not all([patient_id, ae_term]):
        return JsonResponse(
            {"error": "patient_id and ae_term are required"},
            status=400,
        )

    run_id = _get_latest_run_id()
    if not run_id:
        return JsonResponse(
            {"error": "No simulation runs found. Run a simulation first."},
            status=404,
        )

    run_path = _get_run_path(run_id)
    profile, records = _load_patient_data(run_path, patient_id, mode)
    if profile is None:
        return JsonResponse(
            {"error": f"Patient '{patient_id}' not found in run '{run_id}'."},
            status=404,
        )

    meta_path = run_path / "run_meta.json"
    drug_name = "Enfortumab vedotin (Padcev)"
    indication = "Metastatic urothelial carcinoma"
    if meta_path.exists():
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
            drug_name = meta.get("drug_name", drug_name)
            indication = meta.get("indication", indication)
        except Exception:
            pass

    try:
        from datetime import date as dt_date
        from src.doc_agent.service import generate_documents

        result = generate_documents(
            patient_profile=profile,
            day_records=records,
            target_ae_term=ae_term,
            run_id=run_id,
            sim_start_date=dt_date(2026, 1, 6),
            drug_name=drug_name,
            indication=indication,
            target_ae_day=ae_day,
            use_ai=use_ai,
        )
    except Exception as exc:
        logging.exception("Demo generate failed")
        return JsonResponse(
            {"error": f"Document generation failed: {exc}"},
            status=500,
        )

    # Record ai_fields in status file when AI is used
    if use_ai and result.get("success"):
        from datetime import datetime
        ae_slug = ae_term.replace(" ", "_").replace("/", "_")
        ai_fields = {
            "section_b.narrative": True,
            "section_c.dechallenge": True,
            "section_c.rechallenge": True,
        }
        status_data = _read_status(run_id, patient_id, ae_slug)
        status_data["ai_fields"] = ai_fields
        meddra = result.get("meddra", {})
        if meddra:
            status_data["meddra_confidence"] = meddra.get("confidence")
            status_data["meddra_source"] = meddra.get("source")
        status_data["updated_at"] = datetime.utcnow().isoformat()
        if "created_at" not in status_data:
            status_data["created_at"] = datetime.utcnow().isoformat()
        if "status" not in status_data:
            status_data["status"] = "draft"
        _write_status(run_id, patient_id, ae_slug, status_data)

    # Include run_id in the response so the caller knows which run was used
    if isinstance(result, dict):
        result["run_id"] = run_id

    return JsonResponse(result)


@csrf_exempt
@require_GET
def api_demo_reports(request):
    """List all generated reports for the latest run.

    GET /api/demo/reports/
    """
    run_id = _get_latest_run_id()
    if not run_id:
        return JsonResponse(
            {"error": "No simulation runs found. Run a simulation first."},
            status=404,
        )

    from src.doc_agent.service import DOCS_OUTPUT_DIR

    docs_dir = DOCS_OUTPUT_DIR / run_id
    if not docs_dir.exists():
        return JsonResponse({"run_id": run_id, "reports": []})

    reports = []
    for patient_dir in sorted(docs_dir.iterdir()):
        if not patient_dir.is_dir():
            continue
        patient_id = patient_dir.name

        # Group files by ae_slug to pair PDF/XML together
        file_map = {}  # ae_slug -> {pdf_path, xml_path, ...}
        for doc_file in sorted(patient_dir.iterdir()):
            if not doc_file.is_file():
                continue
            fname = doc_file.name
            # Skip status files
            if fname.startswith("report_status_"):
                continue
            # Skip medwatch data JSON files
            if fname.startswith("medwatch_data_"):
                continue

            # Extract ae_slug from filename patterns:
            #   medwatch_3500a_{ae_slug}.pdf
            #   e2b_r3_{ae_slug}.xml
            ae_slug = None
            if fname.startswith("medwatch_3500a_") and fname.endswith(".pdf"):
                ae_slug = fname[len("medwatch_3500a_"):-len(".pdf")]
            elif fname.startswith("e2b_r3_") and fname.endswith(".xml"):
                ae_slug = fname[len("e2b_r3_"):-len(".xml")]

            if ae_slug:
                if ae_slug not in file_map:
                    file_map[ae_slug] = {
                        "patient_id": patient_id,
                        "ae_term": ae_slug.replace("_", " ").title(),
                        "ae_slug": ae_slug,
                    }
                if fname.endswith(".pdf"):
                    file_map[ae_slug]["pdf_path"] = (
                        f"/api/doc/download/{run_id}/{patient_id}/{fname}"
                    )
                    file_map[ae_slug]["created_at"] = (
                        time.strftime(
                            "%Y-%m-%dT%H:%M:%S",
                            time.gmtime(doc_file.stat().st_mtime),
                        )
                    )
                elif fname.endswith(".xml"):
                    file_map[ae_slug]["xml_path"] = (
                        f"/api/doc/download/{run_id}/{patient_id}/{fname}"
                    )

        reports.extend(file_map.values())

    return JsonResponse({"run_id": run_id, "reports": reports})
