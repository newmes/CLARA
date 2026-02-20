"""
CRF Aggregator — CDISC domain aggregation from simulation JSONL + patient JSON.

Reads JSONL simulation data and patient profile JSON, returns domain-specific
row lists suitable for CRF table display and Excel export.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any


# ─── Lab name abbreviation mapping ────────────────────────────────────
# New data uses full names; map them to standard abbreviations for display.
LAB_ABBREVIATIONS: dict[str, str] = {
    "absolute_neutrophil_count": "ANC",
    "alanine_aminotransferase": "ALT",
    "aspartate_aminotransferase": "AST",
    "total_bilirubin": "Bilirubin",
    "bilirubin": "Bilirubin",
    "glucose_fasting": "Glucose (fasting)",
    "weight_kg": "Weight (lbs)",
    "height_cm": "Height (in)",
}


def _lab_display_name(raw_name: str) -> str:
    """Return human-friendly lab test name."""
    return LAB_ABBREVIATIONS.get(raw_name, raw_name.replace("_", " ").title()
                                  if "_" in raw_name else raw_name)


# ─── Unit conversions (metric → US) ──────────────────────────────────

def _c_to_f(val):
    """Celsius → Fahrenheit."""
    if val is None or not isinstance(val, (int, float)):
        return val
    return round(val * 9 / 5 + 32, 1)


def _kg_to_lbs(val):
    """Kilograms → Pounds."""
    if val is None or not isinstance(val, (int, float)):
        return val
    return round(val * 2.20462, 1)


def _cm_to_in(val):
    """Centimeters → Inches."""
    if val is None or not isinstance(val, (int, float)):
        return val
    return round(val / 2.54, 1)


# ─── Column definitions per domain ─────────────────────────────────────

DM_COLUMNS = [
    {"key": "patient_id", "label": "Subject"},
    {"key": "BRTHDAT", "label": "Date of Birth"},
    {"key": "AGE", "label": "Age"},
    {"key": "AGEU", "label": "Age Unit"},
    {"key": "SEX", "label": "Sex"},
    {"key": "RACE", "label": "Race"},
    {"key": "RACEOTH", "label": "Race Other"},
    {"key": "ETHNIC", "label": "Ethnicity"},
]

MH_COLUMNS = [
    {"key": "patient_id", "label": "Subject"},
    {"key": "MHYN", "label": "Reported"},
    {"key": "MHTERM", "label": "Condition"},
    {"key": "MHDAT", "label": "Collection Date"},
    {"key": "MHSTDAT", "label": "Start Date"},
    {"key": "MHONGO", "label": "Ongoing"},
    {"key": "MHENDAT", "label": "End Date"},
]

AE_COLUMNS = [
    {"key": "patient_id", "label": "Subject"},
    {"key": "AEYN", "label": "Reported"},
    {"key": "AETERM", "label": "Adverse Event"},
    {"key": "AESTDAT", "label": "Start Day"},
    {"key": "AEONGO", "label": "Ongoing"},
    {"key": "AEENDAT", "label": "End Day"},
    {"key": "AESEV", "label": "Severity"},
    {"key": "_grade", "label": "Grade"},
    {"key": "AESER", "label": "Serious"},
    {"key": "AESDTH", "label": "Results in Death"},
    {"key": "AESLIFE", "label": "Life-Threatening"},
    {"key": "AESHOSP", "label": "Hospitalization"},
    {"key": "AESDISAB", "label": "Disability"},
    {"key": "AESCONG", "label": "Congenital Anomaly"},
    {"key": "AESMIE", "label": "Other Medically Important"},
    {"key": "AEREL", "label": "Related"},
    {"key": "AEACN", "label": "Action Taken"},
    {"key": "AEACNOTH", "label": "Other Action"},
    {"key": "AEOUT", "label": "Outcome"},
    {"key": "_status", "label": "Status"},
    {"key": "_days_active", "label": "Days Active"},
    {"key": "_visual", "label": "Visual Details"},
]

AE_HR_COLUMNS = [
    {"key": "patient_id", "label": "Subject"},
    {"key": "AEYN", "label": "Reported"},
    {"key": "AETERM", "label": "Adverse Event"},
    {"key": "AESTDAT", "label": "Start Day"},
    {"key": "AEONGO", "label": "Ongoing"},
    {"key": "AEENDAT", "label": "End Day"},
    {"key": "AESEV", "label": "Severity"},
    {"key": "_grade", "label": "Grade"},
    {"key": "AESER", "label": "Serious"},
    {"key": "AESDTH", "label": "Results in Death"},
    {"key": "AESLIFE", "label": "Life-Threatening"},
    {"key": "AESHOSP", "label": "Hospitalization"},
    {"key": "AESDISAB", "label": "Disability"},
    {"key": "AESCONG", "label": "Congenital Anomaly"},
    {"key": "AESMIE", "label": "Other Medically Important"},
    {"key": "AEREL", "label": "Related"},
    {"key": "AEACN", "label": "Action Taken"},
    {"key": "AEACNOTH", "label": "Other Action"},
    {"key": "AEOUT", "label": "Outcome"},
    {"key": "_status", "label": "Status"},
    {"key": "_days_active", "label": "Days Active"},
    {"key": "_visual", "label": "Visual Details"},
    {"key": "detected_day", "label": "Detected Day"},
    {"key": "detection_delay", "label": "Detection Delay"},
    {"key": "channel", "label": "Detection Channel"},
]

EC_COLUMNS = [
    {"key": "patient_id", "label": "Subject"},
    {"key": "day", "label": "Day"},
    {"key": "_cycle", "label": "Cycle"},
    {"key": "ECREFID", "label": "Drug"},
    {"key": "ECSTDAT", "label": "Start Day"},
    {"key": "ECENDAT", "label": "End Day"},
    {"key": "ECDSTXT", "label": "Dose"},
    {"key": "ECDOSU", "label": "Dose Unit"},
    {"key": "ECDOSFRQ", "label": "Frequency"},
    {"key": "ECROUTE", "label": "Route"},
    {"key": "ECDOSADJ", "label": "Dose Adjusted"},
    {"key": "ECADJ", "label": "Adjustment Reason"},
    {"key": "ECCINTD", "label": "Interruption Duration"},
    {"key": "ECCINTDU", "label": "Duration Unit"},
    {"key": "ECITRPYN", "label": "Interrupted"},
    {"key": "ECTRTCMP", "label": "Completed"},
    {"key": "_dose_mg", "label": "Dose (mg)"},
    {"key": "_cumulative_dose_mg", "label": "Cumulative Dose (mg)"},
    {"key": "_dose_level", "label": "Dose Level"},
]

CM_COLUMNS = [
    {"key": "patient_id", "label": "Subject"},
    {"key": "CMYN", "label": "Reported"},
    {"key": "CMTRT", "label": "Medication"},
    {"key": "CMINDC", "label": "Indication"},
    {"key": "CMDSTXT", "label": "Dose"},
    {"key": "CMDOSU", "label": "Dose Unit"},
    {"key": "DOSUO", "label": "Dose Unit Other"},
    {"key": "CMDOSFRM", "label": "Form"},
    {"key": "DOSFRMO", "label": "Form Other"},
    {"key": "CMDOSFRQ", "label": "Frequency"},
    {"key": "DOSFRQO", "label": "Frequency Other"},
    {"key": "CMROUTE", "label": "Route"},
    {"key": "ROUTEO", "label": "Route Other"},
    {"key": "CMSTDAT", "label": "Start Day"},
    {"key": "CMONGO", "label": "Ongoing"},
    {"key": "CMENDAT", "label": "End Day"},
    {"key": "_baseline", "label": "Baseline"},
]

VS_COLUMNS = [
    {"key": "patient_id", "label": "Subject"},
    {"key": "day", "label": "Day"},
    {"key": "SYSBP", "label": "SBP (mmHg)"},
    {"key": "DIABP", "label": "DBP (mmHg)"},
    {"key": "PULSE", "label": "HR (bpm)"},
    {"key": "TEMP", "label": "Temp (°F)"},
    {"key": "RESP", "label": "RR (/min)"},
    {"key": "HEIGHT", "label": "Height (in)"},
    {"key": "WEIGHT", "label": "Weight (lbs)"},
    {"key": "SpO2", "label": "SpO2 (%)"},
]

VS_HR_COLUMNS = [
    {"key": "patient_id", "label": "Subject"},
    {"key": "day", "label": "Day"},
    {"key": "_cycle", "label": "Cycle"},
    {"key": "VSPERF", "label": "Performed"},
    {"key": "VSDAT", "label": "Assessment Date"},
    {"key": "SYSBP_VSORRES", "label": "SBP"},
    {"key": "SYSBP_VSORRESU", "label": "SBP Unit"},
    {"key": "DIABP_VSORRES", "label": "DBP"},
    {"key": "DIABP_VSORRESU", "label": "DBP Unit"},
    {"key": "PULSE_VSORRES", "label": "HR"},
    {"key": "PULSE_VSORRESU", "label": "HR Unit"},
    {"key": "TEMP_VSORRES", "label": "Temp"},
    {"key": "TEMP_VSORRESU", "label": "Temp Unit"},
    {"key": "RESP_VSORRES", "label": "RR"},
    {"key": "RESP_VSORRESU", "label": "RR Unit"},
    {"key": "OXYSAT_VSORRES", "label": "SpO2"},
    {"key": "_SpO2_unit", "label": "SpO2 Unit"},
    {"key": "HEIGHT_VSORRES", "label": "Height"},
    {"key": "HEIGHT_VSORRESU", "label": "Height Unit"},
    {"key": "WEIGHT_VSORRES", "label": "Weight"},
    {"key": "WEIGHT_VSORRESU", "label": "Weight Unit"},
    {"key": "BP_VSPOS", "label": "BP Position"},
    {"key": "BP_VSLOC", "label": "BP Location"},
    {"key": "PULSE_VSLOC", "label": "Pulse Location"},
    {"key": "TEMP_VSLOC", "label": "Temp Location"},
    {"key": "_stale", "label": "Stale"},
]

LB_COLUMNS = [
    {"key": "patient_id", "label": "Subject"},
    {"key": "day", "label": "Day"},
    {"key": "test_name", "label": "Test"},
    {"key": "LBORRES", "label": "Result"},
    {"key": "LBORRESU", "label": "Unit"},
    {"key": "LBORNRLO", "label": "Normal Low"},
    {"key": "LBORNRHI", "label": "Normal High"},
    {"key": "LBNRIND", "label": "Normal/Abnormal"},
    {"key": "LBCLSIG", "label": "Clin. Significant"},
    {"key": "LBCAT", "label": "Category"},
    {"key": "_trend", "label": "Trend"},
]

LB_HR_COLUMNS = [
    {"key": "patient_id", "label": "Subject"},
    {"key": "day", "label": "Day"},
    {"key": "_cycle", "label": "Cycle"},
    {"key": "LBPERF", "label": "Performed"},
    {"key": "LBDAT", "label": "Lab Date"},
    {"key": "test_name", "label": "Test"},
    {"key": "LBORRES", "label": "Result"},
    {"key": "LBORRESU", "label": "Unit"},
    {"key": "LBORNRLO", "label": "Normal Low"},
    {"key": "LBORNRHI", "label": "Normal High"},
    {"key": "LBNRIND", "label": "Normal/Abnormal"},
    {"key": "LBCLSIG", "label": "Clin. Significant"},
    {"key": "LBCAT", "label": "Category"},
    {"key": "_trend", "label": "Trend"},
    {"key": "_stale", "label": "Stale"},
]

DS_COLUMNS = [
    {"key": "patient_id", "label": "Subject"},
    {"key": "day", "label": "Day"},
    {"key": "_cycle", "label": "Cycle"},
    {"key": "DSDECOD", "label": "Disposition"},
    {"key": "DSTERM", "label": "Description"},
    {"key": "DSSTDAT", "label": "Disposition Date"},
]

DD_COLUMNS = [
    {"key": "patient_id", "label": "Subject"},
    {"key": "day", "label": "Day"},
    {"key": "_cycle", "label": "Cycle"},
    {"key": "DTHDAT", "label": "Date of Death"},
    {"key": "DDDECOD", "label": "Cause of Death"},
    {"key": "DDTERM", "label": "Description"},
    {"key": "PRCDTH_DDORRES", "label": "Primary Cause"},
    {"key": "AUTOPIND_DDORRES", "label": "Autopsy Performed"},
]

TU_COLUMNS = [
    {"key": "patient_id", "label": "Subject"},
    {"key": "TUYN", "label": "Assessed"},
    {"key": "day", "label": "Day"},
    {"key": "_cycle", "label": "Cycle"},
    {"key": "TULNKID", "label": "Lesion Link ID"},
    {"key": "TULOC", "label": "Location"},
    {"key": "TULAT", "label": "Laterality"},
    {"key": "TUDIR", "label": "Direction"},
    {"key": "TULOCDTL", "label": "Location Detail"},
    {"key": "TUMETHOD", "label": "Method"},
    {"key": "TUDAT", "label": "Assessment Date"},
    {"key": "TUEVAL", "label": "Evaluator"},
    {"key": "TUEVALID", "label": "Evaluator ID"},
    {"key": "TRORRES", "label": "Target Lesion Result"},
    {"key": "TRORRESU", "label": "Result Unit"},
    {"key": "_baseline_mm", "label": "Baseline (mm)"},
    {"key": "_change_pct", "label": "Change (%)"},
    {"key": "TRSTAT", "label": "Status"},
    {"key": "TRREASND", "label": "Reason Not Done"},
    {"key": "TURESULT", "label": "Result"},
]

RS_COLUMNS = [
    {"key": "patient_id", "label": "Subject"},
    {"key": "day", "label": "Day"},
    {"key": "_cycle", "label": "Cycle"},
    {"key": "RSPERF", "label": "Performed"},
    {"key": "RSCAT", "label": "Assessment Category"},
    {"key": "RSEVAL", "label": "Evaluator"},
    {"key": "RSEVALID", "label": "Evaluator ID"},
    {"key": "TRGRESP_RSORRES", "label": "Target Response"},
    {"key": "NTRGRESP_RSORRES", "label": "Non-Target Response"},
    {"key": "OVRLRESP_RSORRES", "label": "Overall Response"},
    {"key": "BESTRESP_RSORRES", "label": "Best Overall Response"},
    {"key": "_tumor_change_pct", "label": "Tumor Change (%)"},
    {"key": "_nadir_pct", "label": "Nadir (%)"},
    {"key": "_description", "label": "Description"},
    {"key": "RSRESULT", "label": "Response"},
    {"key": "RSTESTCD", "label": "Test Code"},
    {"key": "RSREASND", "label": "Reason Not Done"},
]

PE_COLUMNS = [
    {"key": "patient_id", "label": "Subject"},
    {"key": "day", "label": "Day"},
    {"key": "_cycle", "label": "Cycle"},
    {"key": "PEPERF", "label": "Performed"},
    {"key": "PEDAT", "label": "Exam Date"},
]

EG_COLUMNS = [
    {"key": "patient_id", "label": "Subject"},
    {"key": "day", "label": "Day"},
    {"key": "_cycle", "label": "Cycle"},
    {"key": "EGPERF", "label": "Performed"},
    {"key": "EGREFID", "label": "Reference ID"},
    {"key": "EGMETHOD", "label": "Method"},
    {"key": "EGPOS", "label": "Position"},
    {"key": "EGDAT", "label": "ECG Date"},
]

DOMAIN_COLUMNS = {
    "DM": DM_COLUMNS,
    "MH": MH_COLUMNS,
    "AE": AE_COLUMNS,
    "AE_HR": AE_HR_COLUMNS,
    "EC": EC_COLUMNS,
    "CM": CM_COLUMNS,
    "VS": VS_COLUMNS,
    "VS_HR": VS_HR_COLUMNS,
    "LB": LB_COLUMNS,
    "LB_HR": LB_HR_COLUMNS,
    "DS": DS_COLUMNS,
    "DD": DD_COLUMNS,
    "TU": TU_COLUMNS,
    "RS": RS_COLUMNS,
    "PE": PE_COLUMNS,
    "EG": EG_COLUMNS,
}

# Domain display labels (code → full name)
DOMAIN_LABELS = {
    "AE": "AE — Adverse Events",
    "EC": "EC — Exposures",
    "CM": "CM — Concomitant Meds",
    "LB": "LB — Labs",
    "VS": "VS — Vital Signs",
    "DM": "DM — Demographics",
    "MH": "MH — Medical History",
    "DS": "DS — Disposition",
    "DD": "DD — Death Details",
    "TU": "TU — Tumor",
    "RS": "RS — RECIST Response",
    "PE": "PE — Physical Exam",
    "EG": "EG — ECG",
}


# ─── Internal helpers ───────────────────────────────────────────────────

def _iter_jsonl(run_path: Path, patient_id: str, mode: str = "natural"):
    """Yield parsed JSON records from a simulation JSONL file."""
    fpath = run_path / "simulations" / f"{patient_id}_{mode}.jsonl"
    if not fpath.exists():
        return
    with open(fpath, encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if line:
                yield json.loads(line)


def _iter_hospital_jsonl(run_path: Path, patient_id: str, mode: str = "natural"):
    """Yield parsed records from a hospital JSONL file.

    Prefers the dedicated *_hospital.jsonl (v2.3+).
    Falls back to extracting hospital_record from the GT file (older runs).
    """
    hr_path = run_path / "simulations" / f"{patient_id}_{mode}_hospital.jsonl"
    if hr_path.exists():
        with open(hr_path, encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if line:
                    yield json.loads(line)
    else:
        # Legacy: extract hospital_record from GT file
        for record in _iter_jsonl(run_path, patient_id, mode):
            hr = record.get("hospital_record")
            if hr:
                merged = {
                    "day": record.get("day"),
                    "cycle": record.get("cycle"),
                    "cycle_day": record.get("cycle_day"),
                }
                merged.update(hr)
                yield merged
            else:
                # Oldest format: no hospital_record at all, use GT directly
                yield record


def _load_patient_json(run_path: Path, patient_id: str) -> dict | None:
    """Load patient profile JSON."""
    fpath = run_path / "patients" / f"{patient_id}.json"
    if not fpath.exists():
        return None
    with open(fpath, encoding="utf-8") as fh:
        return json.load(fh)


def _list_patient_ids(run_path: Path) -> list[str]:
    """List available patient IDs from patients/ directory."""
    patients_dir = run_path / "patients"
    if not patients_dir.exists():
        return []
    return sorted(
        f.stem for f in patients_dir.glob("PT-*.json")
    )


def _resolve_patients(run_path: Path, patient_ids: list[str] | None) -> list[str]:
    """Resolve patient list: use provided or discover all."""
    if patient_ids:
        return patient_ids
    return _list_patient_ids(run_path)


def _fmt_cycle(cycle, cycle_day) -> str | None:
    """Format cycle/cycle_day as 'C2D8' string."""
    if cycle is None or cycle_day is None:
        return None
    return f"C{cycle}D{cycle_day}"


def _paginate(rows: list[dict], page: int = 1, per_page: int = 50) -> tuple[list[dict], int]:
    """Slice rows for pagination. Returns (page_rows, total_count)."""
    total = len(rows)
    if per_page <= 0:
        return rows, total
    start = (page - 1) * per_page
    end = start + per_page
    return rows[start:end], total


# ─── Domain aggregation functions ───────────────────────────────────────

def aggregate_dm(
    run_path: Path,
    patient_ids: list[str] | None = None,
    page: int = 1,
    per_page: int = 50,
) -> tuple[list[dict], int, list[dict]]:
    """DM domain: 1 row per patient from patient profile JSON."""
    patients = _resolve_patients(run_path, patient_ids)
    rows = []
    for pid in patients:
        profile = _load_patient_json(run_path, pid)
        if not profile:
            continue
        dm = profile.get("DM", {})
        rows.append({
            "patient_id": pid,
            "BRTHDAT": dm.get("BRTHDAT"),
            "AGE": dm.get("AGE"),
            "AGEU": dm.get("AGEU", "YEARS"),
            "SEX": dm.get("SEX"),
            "RACE": dm.get("RACE"),
            "RACEOTH": dm.get("RACEOTH"),
            "ETHNIC": dm.get("ETHNIC"),
        })
    page_rows, total = _paginate(rows, page, per_page)
    return page_rows, total, DM_COLUMNS


def aggregate_mh(
    run_path: Path,
    patient_ids: list[str] | None = None,
    page: int = 1,
    per_page: int = 50,
) -> tuple[list[dict], int, list[dict]]:
    """MH domain: 1 row per medical history item from patient profile."""
    patients = _resolve_patients(run_path, patient_ids)
    rows = []
    for pid in patients:
        profile = _load_patient_json(run_path, pid)
        if not profile:
            continue
        for mh in profile.get("MH", []):
            rows.append({
                "patient_id": pid,
                "MHYN": mh.get("MHYN"),
                "MHTERM": mh.get("MHTERM"),
                "MHDAT": mh.get("MHDAT"),
                "MHSTDAT": mh.get("MHSTDAT"),
                "MHONGO": mh.get("MHONGO"),
                "MHENDAT": mh.get("MHENDAT"),
            })
    page_rows, total = _paginate(rows, page, per_page)
    return page_rows, total, MH_COLUMNS


def aggregate_ae(
    run_path: Path,
    patient_ids: list[str] | None = None,
    mode: str = "natural",
    source: str = "hr",
    page: int = 1,
    per_page: int = 50,
) -> tuple[list[dict], int, list[dict]]:
    """AE domain: deduplicated by (patient_id, AETERM, AESTDAT), latest state.

    source="gt": from JSONL AE array (ground truth)
    source="hr": from hospital_record.objective.active_aes (hospital record)
    """
    patients = _resolve_patients(run_path, patient_ids)

    if source == "hr":
        return _aggregate_ae_hr(run_path, patients, mode, page, per_page)

    rows_map: dict[tuple, dict] = {}  # (pid, AETERM, AESTDAT) -> latest row
    for pid in patients:
        for record in _iter_jsonl(run_path, pid, mode):
            for ae in record.get("AE", []):
                key = (pid, ae.get("AETERM"), ae.get("AESTDAT"))
                rows_map[key] = {
                    "patient_id": pid,
                    "AEYN": ae.get("AEYN"),
                    "AETERM": ae.get("AETERM"),
                    "AESTDAT": ae.get("AESTDAT"),
                    "AEONGO": ae.get("AEONGO"),
                    "AEENDAT": ae.get("AEENDAT"),
                    "AESEV": ae.get("AESEV"),
                    "_grade": ae.get("_grade") or ae.get("AETOXGR"),
                    "AESER": ae.get("AESER"),
                    "AESDTH": ae.get("AESDTH"),
                    "AESLIFE": ae.get("AESLIFE"),
                    "AESHOSP": ae.get("AESHOSP"),
                    "AESDISAB": ae.get("AESDISAB"),
                    "AESCONG": ae.get("AESCONG"),
                    "AESMIE": ae.get("AESMIE"),
                    "AEREL": ae.get("AEREL"),
                    "AEACN": ae.get("AEACN"),
                    "AEACNOTH": ae.get("AEACNOTH"),
                    "AEOUT": ae.get("AEOUT"),
                    "_status": ae.get("_status"),
                    "_days_active": ae.get("_days_active"),
                    "_visual": ae.get("_visual"),
                }
    rows = sorted(rows_map.values(), key=lambda r: (r["patient_id"], r["AESTDAT"] or 0))
    page_rows, total = _paginate(rows, page, per_page)
    return page_rows, total, AE_COLUMNS


def _aggregate_ae_hr(
    run_path: Path,
    patients: list[str],
    mode: str,
    page: int,
    per_page: int,
) -> tuple[list[dict], int, list[dict]]:
    """AE from hospital record — full CDISC fields + detection metadata.

    Reads top-level AE[] from *_hospital.jsonl (full CDISC),
    enriched with detection info from objective.active_aes.
    Falls back to GT hospital_record for older runs.
    """
    rows_map: dict[tuple, dict] = {}
    for pid in patients:
        for record in _iter_hospital_jsonl(run_path, pid, mode):
            day = record.get("day")
            # Build detection lookup from objective.active_aes
            obj = record.get("objective", {})
            if not obj:
                obj = record.get("hospital_record", {}).get("objective", {})
            detection_map: dict[str, dict] = {}
            for det in obj.get("active_aes", []):
                det_name = det.get("ae", "")
                detection_map[det_name] = det

            # Read full CDISC AE from top-level
            for ae in record.get("AE", []):
                ae_term = ae.get("AETERM", "")
                onset = ae.get("AESTDAT")
                key = (pid, ae_term, onset)
                det = detection_map.get(ae_term, {})
                rows_map[key] = {
                    "patient_id": pid,
                    "AEYN": ae.get("AEYN"),
                    "AETERM": ae_term,
                    "AESTDAT": onset,
                    "AEONGO": ae.get("AEONGO"),
                    "AEENDAT": ae.get("AEENDAT"),
                    "AESEV": ae.get("AESEV"),
                    "_grade": ae.get("_grade") or ae.get("AETOXGR"),
                    "AESER": ae.get("AESER"),
                    "AESDTH": ae.get("AESDTH"),
                    "AESLIFE": ae.get("AESLIFE"),
                    "AESHOSP": ae.get("AESHOSP"),
                    "AESDISAB": ae.get("AESDISAB"),
                    "AESCONG": ae.get("AESCONG"),
                    "AESMIE": ae.get("AESMIE"),
                    "AEREL": ae.get("AEREL"),
                    "AEACN": ae.get("AEACN"),
                    "AEACNOTH": ae.get("AEACNOTH"),
                    "AEOUT": ae.get("AEOUT"),
                    "_status": ae.get("_status"),
                    "_days_active": ae.get("_days_active"),
                    "_visual": ae.get("_visual"),
                    "detected_day": det.get("detected_day") or day,
                    "detection_delay": det.get("detection_delay"),
                    "channel": det.get("channel", ""),
                }

            # Legacy fallback: if no top-level AE but objective.active_aes exists
            if not record.get("AE") and detection_map:
                for det in obj.get("active_aes", []):
                    ae_name = det.get("ae", "")
                    onset = det.get("onset_day")
                    key = (pid, ae_name, onset)
                    if key not in rows_map:
                        rows_map[key] = {
                            "patient_id": pid,
                            "AEYN": None,
                            "AETERM": ae_name,
                            "AESTDAT": onset,
                            "AEONGO": None,
                            "AEENDAT": None,
                            "AESEV": None,
                            "_grade": det.get("grade"),
                            "AESER": None,
                            "AESDTH": None,
                            "AESLIFE": None,
                            "AESHOSP": None,
                            "AESDISAB": None,
                            "AESCONG": None,
                            "AESMIE": None,
                            "AEREL": None,
                            "AEACN": None,
                            "AEACNOTH": None,
                            "AEOUT": None,
                            "_status": det.get("status"),
                            "_days_active": None,
                            "_visual": None,
                            "detected_day": det.get("detected_day"),
                            "detection_delay": det.get("detection_delay"),
                            "channel": det.get("channel", ""),
                        }

    rows = sorted(rows_map.values(), key=lambda r: (r["patient_id"], r["AESTDAT"] or 0))
    page_rows, total = _paginate(rows, page, per_page)
    return page_rows, total, AE_HR_COLUMNS


def aggregate_ec(
    run_path: Path,
    patient_ids: list[str] | None = None,
    mode: str = "natural",
    page: int = 1,
    per_page: int = 50,
) -> tuple[list[dict], int, list[dict]]:
    """EC domain: 1 row per drug exposure record per day."""
    patients = _resolve_patients(run_path, patient_ids)
    rows = []
    for pid in patients:
        for record in _iter_jsonl(run_path, pid, mode):
            day = record.get("day")
            cycle_display = _fmt_cycle(record.get("cycle"), record.get("cycle_day"))
            for ec in record.get("EC", []):
                rows.append({
                    "patient_id": pid,
                    "day": day,
                    "_cycle": cycle_display,
                    "ECREFID": ec.get("ECREFID"),
                    "ECSTDAT": ec.get("ECSTDAT"),
                    "ECENDAT": ec.get("ECENDAT"),
                    "ECDSTXT": ec.get("ECDSTXT"),
                    "ECDOSU": ec.get("ECDOSU"),
                    "ECDOSFRQ": ec.get("ECDOSFRQ"),
                    "ECROUTE": ec.get("ECROUTE"),
                    "ECDOSADJ": ec.get("ECDOSADJ"),
                    "ECADJ": ec.get("ECADJ"),
                    "ECCINTD": ec.get("ECCINTD"),
                    "ECCINTDU": ec.get("ECCINTDU"),
                    "ECITRPYN": ec.get("ECITRPYN"),
                    "ECTRTCMP": ec.get("ECTRTCMP"),
                    "_dose_mg": ec.get("_dose_mg"),
                    "_cumulative_dose_mg": ec.get("_cumulative_dose_mg"),
                    "_dose_level": ec.get("_dose_level"),
                })
    page_rows, total = _paginate(rows, page, per_page)
    return page_rows, total, EC_COLUMNS


def aggregate_cm(
    run_path: Path,
    patient_ids: list[str] | None = None,
    mode: str = "natural",
    page: int = 1,
    per_page: int = 50,
) -> tuple[list[dict], int, list[dict]]:
    """CM domain: deduplicated by (patient_id, drug name), latest state."""
    patients = _resolve_patients(run_path, patient_ids)
    rows_map: dict[tuple, dict] = {}
    for pid in patients:
        for record in _iter_jsonl(run_path, pid, mode):
            for cm in record.get("CM", []):
                drug = cm.get("CMTRT", "")
                key = (pid, drug)
                rows_map[key] = {
                    "patient_id": pid,
                    "CMYN": cm.get("CMYN"),
                    "CMTRT": drug,
                    "CMINDC": cm.get("CMINDC"),
                    "CMDSTXT": cm.get("CMDSTXT"),
                    "CMDOSU": cm.get("CMDOSU"),
                    "DOSUO": cm.get("DOSUO"),
                    "CMDOSFRM": cm.get("CMDOSFRM"),
                    "DOSFRMO": cm.get("DOSFRMO"),
                    "CMDOSFRQ": cm.get("CMDOSFRQ"),
                    "DOSFRQO": cm.get("DOSFRQO"),
                    "CMROUTE": cm.get("CMROUTE"),
                    "ROUTEO": cm.get("ROUTEO"),
                    "CMSTDAT": cm.get("CMSTDAT"),
                    "CMONGO": cm.get("CMONGO"),
                    "CMENDAT": cm.get("CMENDAT"),
                    "_baseline": cm.get("_baseline"),
                }
    rows = sorted(rows_map.values(), key=lambda r: (r["patient_id"], r["CMTRT"] or ""))
    page_rows, total = _paginate(rows, page, per_page)
    return page_rows, total, CM_COLUMNS


def aggregate_vs(
    run_path: Path,
    patient_ids: list[str] | None = None,
    mode: str = "natural",
    source: str = "hr",
    page: int = 1,
    per_page: int = 50,
) -> tuple[list[dict], int, list[dict]]:
    """VS domain: 1 row per day.

    source="gt": from JSONL VS (ground truth, every day)
    source="hr": from hospital_record.objective.vitals (observed only)
    """
    patients = _resolve_patients(run_path, patient_ids)

    if source == "hr":
        return _aggregate_vs_hr(run_path, patients, mode, page, per_page)

    rows = []
    for pid in patients:
        for record in _iter_jsonl(run_path, pid, mode):
            vs = record.get("VS")
            if not vs or not vs.get("VSPERF"):
                continue
            rows.append({
                "patient_id": pid,
                "day": record.get("day"),
                "SYSBP": vs.get("SYSBP_VSORRES"),
                "DIABP": vs.get("DIABP_VSORRES"),
                "PULSE": vs.get("PULSE_VSORRES"),
                "TEMP": _c_to_f(vs.get("TEMP_VSORRES")),
                "RESP": vs.get("RESP_VSORRES"),
                "HEIGHT": _cm_to_in(vs.get("HEIGHT_VSORRES")),
                "WEIGHT": _kg_to_lbs(vs.get("WEIGHT_VSORRES")),
                "SpO2": vs.get("_SpO2") or vs.get("OXYSAT_VSORRES"),
            })
    page_rows, total = _paginate(rows, page, per_page)
    return page_rows, total, VS_COLUMNS


def _aggregate_vs_hr(
    run_path: Path,
    patients: list[str],
    mode: str,
    page: int,
    per_page: int,
) -> tuple[list[dict], int, list[dict]]:
    """VS from hospital record — reads top-level VS from *_hospital.jsonl.

    Includes full CDISC fields (units, positions, locations) plus
    backward-compatible simplified keys (SBP, DBP, HR, etc.) for chart rendering.
    Falls back to objective.vitals for older runs.
    """
    rows = []
    for pid in patients:
        for record in _iter_hospital_jsonl(run_path, pid, mode):
            day = record.get("day")
            cycle_display = _fmt_cycle(record.get("cycle"), record.get("cycle_day"))

            vs = record.get("VS")
            if vs:
                # New format: top-level VS with full CDISC fields
                sysbp = vs.get("SYSBP_VSORRES")
                diabp = vs.get("DIABP_VSORRES")
                pulse = vs.get("PULSE_VSORRES")
                temp = vs.get("TEMP_VSORRES")
                resp = vs.get("RESP_VSORRES")
                height = vs.get("HEIGHT_VSORRES")
                weight = vs.get("WEIGHT_VSORRES")
                spo2 = vs.get("_SpO2") or vs.get("OXYSAT_VSORRES")
                rows.append({
                    "patient_id": pid,
                    "day": day,
                    "_cycle": cycle_display,
                    "VSPERF": vs.get("VSPERF"),
                    "VSDAT": vs.get("VSDAT"),
                    "SYSBP_VSORRES": sysbp,
                    "SYSBP_VSORRESU": vs.get("SYSBP_VSORRESU"),
                    "DIABP_VSORRES": diabp,
                    "DIABP_VSORRESU": vs.get("DIABP_VSORRESU"),
                    "PULSE_VSORRES": pulse,
                    "PULSE_VSORRESU": vs.get("PULSE_VSORRESU"),
                    "TEMP_VSORRES": _c_to_f(temp),
                    "TEMP_VSORRESU": "°F",
                    "RESP_VSORRES": resp,
                    "RESP_VSORRESU": vs.get("RESP_VSORRESU"),
                    "OXYSAT_VSORRES": spo2,
                    "_SpO2_unit": vs.get("_SpO2_unit"),
                    "HEIGHT_VSORRES": _cm_to_in(height),
                    "HEIGHT_VSORRESU": "in",
                    "WEIGHT_VSORRES": _kg_to_lbs(weight),
                    "WEIGHT_VSORRESU": "lbs",
                    "BP_VSPOS": vs.get("BP_VSPOS"),
                    "BP_VSLOC": vs.get("BP_VSLOC"),
                    "PULSE_VSLOC": vs.get("PULSE_VSLOC"),
                    "TEMP_VSLOC": vs.get("TEMP_VSLOC"),
                    "_stale": vs.get("_stale"),
                    # Backward-compatible keys for VS chart
                    "SBP": sysbp,
                    "DBP": diabp,
                    "HR": pulse,
                    "BT": _c_to_f(temp),
                    "RR": resp,
                    "height_cm": _cm_to_in(height),
                    "weight_kg": _kg_to_lbs(weight),
                    "SpO2": spo2,
                    "stale_days": 0 if vs.get("VSPERF") else None,
                })
            else:
                # Legacy: objective.vitals
                obj = record.get("objective", {})
                if not obj:
                    obj = record.get("hospital_record", {}).get("objective", {})
                vitals = obj.get("vitals")
                if not vitals:
                    continue
                stale = obj.get("vitals_stale_days", 0)
                sysbp = vitals.get("SBP")
                diabp = vitals.get("DBP")
                pulse = vitals.get("HR")
                temp = vitals.get("BT")
                resp = vitals.get("RR")
                height = vitals.get("height_cm")
                weight = vitals.get("weight_kg")
                spo2 = vitals.get("SpO2")
                rows.append({
                    "patient_id": pid,
                    "day": day,
                    "_cycle": cycle_display,
                    "VSPERF": None,
                    "VSDAT": None,
                    "SYSBP_VSORRES": sysbp,
                    "SYSBP_VSORRESU": None,
                    "DIABP_VSORRES": diabp,
                    "DIABP_VSORRESU": None,
                    "PULSE_VSORRES": pulse,
                    "PULSE_VSORRESU": None,
                    "TEMP_VSORRES": _c_to_f(temp),
                    "TEMP_VSORRESU": "°F",
                    "RESP_VSORRES": resp,
                    "RESP_VSORRESU": None,
                    "OXYSAT_VSORRES": spo2,
                    "_SpO2_unit": None,
                    "HEIGHT_VSORRES": _cm_to_in(height),
                    "HEIGHT_VSORRESU": "in",
                    "WEIGHT_VSORRES": _kg_to_lbs(weight),
                    "WEIGHT_VSORRESU": "lbs",
                    "BP_VSPOS": None,
                    "BP_VSLOC": None,
                    "PULSE_VSLOC": None,
                    "TEMP_VSLOC": None,
                    "_stale": stale > 0 if stale else None,
                    # Backward-compatible keys
                    "SBP": sysbp,
                    "DBP": diabp,
                    "HR": pulse,
                    "BT": _c_to_f(temp),
                    "RR": resp,
                    "height_cm": _cm_to_in(height),
                    "weight_kg": _kg_to_lbs(weight),
                    "SpO2": spo2,
                    "stale_days": stale,
                })
    page_rows, total = _paginate(rows, page, per_page)
    return page_rows, total, VS_HR_COLUMNS


def aggregate_lb(
    run_path: Path,
    patient_ids: list[str] | None = None,
    mode: str = "natural",
    source: str = "hr",
    page: int = 1,
    per_page: int = 50,
) -> tuple[list[dict], int, list[dict]]:
    """LB domain: 1 row per lab test per day.

    source="gt": from JSONL LB.results (ground truth)
    source="hr": from hospital_record.objective.labs (observed)
    """
    patients = _resolve_patients(run_path, patient_ids)

    if source == "hr":
        return _aggregate_lb_hr(run_path, patients, mode, page, per_page)

    rows = []
    for pid in patients:
        for record in _iter_jsonl(run_path, pid, mode):
            lb = record.get("LB")
            if not lb or not lb.get("LBPERF"):
                continue
            day = record.get("day")
            results = lb.get("results", {})
            for test_name, vals in results.items():
                rows.append({
                    "patient_id": pid,
                    "day": day,
                    "test_name": _lab_display_name(test_name),
                    "LBORRES": vals.get("LBORRES"),
                    "LBORRESU": vals.get("LBORRESU"),
                    "LBORNRLO": vals.get("LBORNRLO"),
                    "LBORNRHI": vals.get("LBORNRHI"),
                    "LBNRIND": vals.get("LBNRIND"),
                    "LBCLSIG": vals.get("LBCLSIG"),
                    "LBCAT": vals.get("LBCAT"),
                    "_trend": vals.get("_trend"),
                })
    page_rows, total = _paginate(rows, page, per_page)
    return page_rows, total, LB_COLUMNS


def _aggregate_lb_hr(
    run_path: Path,
    patients: list[str],
    mode: str,
    page: int,
    per_page: int,
) -> tuple[list[dict], int, list[dict]]:
    """LB from hospital record — reads top-level LB from *_hospital.jsonl.

    Includes full CDISC fields (ranges, indicators, categories) plus
    backward-compatible simplified keys (value, unit, trend, stale_days)
    for chart rendering.  Falls back to objective.labs for older runs.
    """
    rows = []
    for pid in patients:
        for record in _iter_hospital_jsonl(run_path, pid, mode):
            day = record.get("day")
            cycle_display = _fmt_cycle(record.get("cycle"), record.get("cycle_day"))

            lb = record.get("LB")
            if lb and lb.get("results"):
                # New format: top-level LB with full CDISC fields
                results = lb["results"]
                lb_perf = lb.get("LBPERF")
                lb_dat = lb.get("LBDAT")
                lb_stale = lb.get("_stale", False)
                obj = record.get("objective", {})
                stale_days = obj.get("labs_stale_days", 0)
                for test_name, vals in results.items():
                    result_val = vals.get("LBORRES")
                    result_unit = vals.get("LBORRESU")
                    rows.append({
                        "patient_id": pid,
                        "day": day,
                        "_cycle": cycle_display,
                        "LBPERF": lb_perf,
                        "LBDAT": lb_dat,
                        "test_name": _lab_display_name(test_name),
                        "LBORRES": result_val,
                        "LBORRESU": result_unit,
                        "LBORNRLO": vals.get("LBORNRLO"),
                        "LBORNRHI": vals.get("LBORNRHI"),
                        "LBNRIND": vals.get("LBNRIND"),
                        "LBCLSIG": vals.get("LBCLSIG"),
                        "LBCAT": vals.get("LBCAT"),
                        "_trend": vals.get("_trend"),
                        "_stale": vals.get("_stale", lb_stale),
                        # Backward-compatible keys for LB chart
                        "value": result_val,
                        "unit": result_unit,
                        "trend": vals.get("_trend"),
                        "stale_days": stale_days,
                    })
            else:
                # Legacy: objective.labs
                obj = record.get("objective", {})
                if not obj:
                    obj = record.get("hospital_record", {}).get("objective", {})
                labs = obj.get("labs")
                if not labs:
                    continue
                stale = obj.get("labs_stale_days", 0)
                for test_name, vals in labs.items():
                    result_val = vals.get("value")
                    result_unit = vals.get("unit")
                    rows.append({
                        "patient_id": pid,
                        "day": day,
                        "_cycle": cycle_display,
                        "LBPERF": None,
                        "LBDAT": None,
                        "test_name": _lab_display_name(test_name),
                        "LBORRES": result_val,
                        "LBORRESU": result_unit,
                        "LBORNRLO": None,
                        "LBORNRHI": None,
                        "LBNRIND": None,
                        "LBCLSIG": None,
                        "LBCAT": None,
                        "_trend": vals.get("trend"),
                        "_stale": None,
                        # Backward-compatible keys
                        "value": result_val,
                        "unit": result_unit,
                        "trend": vals.get("trend"),
                        "stale_days": stale,
                    })
    page_rows, total = _paginate(rows, page, per_page)
    return page_rows, total, LB_HR_COLUMNS


def aggregate_ds(
    run_path: Path,
    patient_ids: list[str] | None = None,
    mode: str = "natural",
    page: int = 1,
    per_page: int = 50,
) -> tuple[list[dict], int, list[dict]]:
    """DS domain: 0-1 row per patient (last non-null DS record)."""
    patients = _resolve_patients(run_path, patient_ids)
    rows = []
    for pid in patients:
        last_ds = None
        last_day = None
        last_cycle = None
        for record in _iter_jsonl(run_path, pid, mode):
            ds = record.get("DS")
            if ds:
                last_ds = ds
                last_day = record.get("day")
                last_cycle = _fmt_cycle(record.get("cycle"), record.get("cycle_day"))
        if last_ds:
            rows.append({
                "patient_id": pid,
                "day": last_day,
                "_cycle": last_cycle,
                "DSDECOD": last_ds.get("DSDECOD", ""),
                "DSTERM": last_ds.get("DSTERM", ""),
                "DSSTDAT": last_ds.get("DSSTDAT"),
            })
    page_rows, total = _paginate(rows, page, per_page)
    return page_rows, total, DS_COLUMNS


def aggregate_dd(
    run_path: Path,
    patient_ids: list[str] | None = None,
    mode: str = "natural",
    page: int = 1,
    per_page: int = 50,
) -> tuple[list[dict], int, list[dict]]:
    """DD domain: 0-1 row per patient (death detail)."""
    patients = _resolve_patients(run_path, patient_ids)
    rows = []
    for pid in patients:
        last_dd = None
        last_day = None
        last_cycle = None
        for record in _iter_jsonl(run_path, pid, mode):
            dd = record.get("DD")
            if dd:
                last_dd = dd
                last_day = record.get("day")
                last_cycle = _fmt_cycle(record.get("cycle"), record.get("cycle_day"))
        if last_dd:
            rows.append({
                "patient_id": pid,
                "day": last_day,
                "_cycle": last_cycle,
                "DTHDAT": last_dd.get("DTHDAT"),
                "DDDECOD": last_dd.get("DDDECOD", ""),
                "DDTERM": last_dd.get("DDTERM", ""),
                "PRCDTH_DDORRES": last_dd.get("PRCDTH_DDORRES"),
                "AUTOPIND_DDORRES": last_dd.get("AUTOPIND_DDORRES"),
            })
    page_rows, total = _paginate(rows, page, per_page)
    return page_rows, total, DD_COLUMNS


def aggregate_tu(
    run_path: Path,
    patient_ids: list[str] | None = None,
    mode: str = "natural",
    page: int = 1,
    per_page: int = 50,
) -> tuple[list[dict], int, list[dict]]:
    """TU domain: tumor assessment rows (1 per scan day)."""
    patients = _resolve_patients(run_path, patient_ids)
    rows = []
    for pid in patients:
        for record in _iter_jsonl(run_path, pid, mode):
            tu = record.get("TU")
            if not tu:
                continue
            day = record.get("day")
            cycle_display = _fmt_cycle(record.get("cycle"), record.get("cycle_day"))
            items = tu if isinstance(tu, list) else [tu] if isinstance(tu, dict) else []
            for item in items:
                rows.append({
                    "patient_id": pid,
                    "TUYN": item.get("TUYN"),
                    "day": day,
                    "_cycle": cycle_display,
                    "TULNKID": item.get("TULNKID"),
                    "TULOC": item.get("TULOC", ""),
                    "TULAT": item.get("TULAT"),
                    "TUDIR": item.get("TUDIR"),
                    "TULOCDTL": item.get("TULOCDTL"),
                    "TUMETHOD": item.get("TUMETHOD", ""),
                    "TUDAT": item.get("TUDAT"),
                    "TUEVAL": item.get("TUEVAL"),
                    "TUEVALID": item.get("TUEVALID"),
                    "TRORRES": item.get("TRORRES"),
                    "TRORRESU": item.get("TRORRESU"),
                    "_baseline_mm": item.get("_baseline_mm"),
                    "_change_pct": item.get("_change_pct"),
                    "TRSTAT": item.get("TRSTAT"),
                    "TRREASND": item.get("TRREASND"),
                    "TURESULT": item.get("TURESULT", ""),
                })
    page_rows, total = _paginate(rows, page, per_page)
    return page_rows, total, TU_COLUMNS


def aggregate_rs(
    run_path: Path,
    patient_ids: list[str] | None = None,
    mode: str = "natural",
    page: int = 1,
    per_page: int = 50,
) -> tuple[list[dict], int, list[dict]]:
    """RS domain: RECIST response assessment (1 per scan)."""
    patients = _resolve_patients(run_path, patient_ids)
    rows = []
    for pid in patients:
        for record in _iter_jsonl(run_path, pid, mode):
            rs = record.get("RS")
            if not rs:
                continue
            day = record.get("day")
            cycle_display = _fmt_cycle(record.get("cycle"), record.get("cycle_day"))
            items = rs if isinstance(rs, list) else [rs] if isinstance(rs, dict) else []
            for item in items:
                rows.append({
                    "patient_id": pid,
                    "day": day,
                    "_cycle": cycle_display,
                    "RSPERF": item.get("RSPERF"),
                    "RSCAT": item.get("RSCAT"),
                    "RSEVAL": item.get("RSEVAL", ""),
                    "RSEVALID": item.get("RSEVALID"),
                    "TRGRESP_RSORRES": item.get("TRGRESP_RSORRES"),
                    "NTRGRESP_RSORRES": item.get("NTRGRESP_RSORRES"),
                    "OVRLRESP_RSORRES": item.get("OVRLRESP_RSORRES"),
                    "BESTRESP_RSORRES": item.get("BESTRESP_RSORRES"),
                    "_tumor_change_pct": item.get("_tumor_change_pct"),
                    "_nadir_pct": item.get("_nadir_pct"),
                    "_description": item.get("_description", ""),
                    "RSRESULT": item.get("RSRESULT", item.get("RSORRES", "")),
                    "RSTESTCD": item.get("RSTESTCD", ""),
                    "RSREASND": item.get("RSREASND"),
                })
    page_rows, total = _paginate(rows, page, per_page)
    return page_rows, total, RS_COLUMNS


def aggregate_pe(
    run_path: Path,
    patient_ids: list[str] | None = None,
    mode: str = "natural",
    page: int = 1,
    per_page: int = 50,
) -> tuple[list[dict], int, list[dict]]:
    """PE domain: physical exam records (1 per exam day)."""
    patients = _resolve_patients(run_path, patient_ids)
    rows = []
    for pid in patients:
        for record in _iter_jsonl(run_path, pid, mode):
            pe = record.get("PE")
            if not pe or not pe.get("PEPERF"):
                continue
            rows.append({
                "patient_id": pid,
                "day": record.get("day"),
                "_cycle": _fmt_cycle(record.get("cycle"), record.get("cycle_day")),
                "PEPERF": pe.get("PEPERF"),
                "PEDAT": pe.get("PEDAT"),
            })
    page_rows, total = _paginate(rows, page, per_page)
    return page_rows, total, PE_COLUMNS


def aggregate_eg(
    run_path: Path,
    patient_ids: list[str] | None = None,
    mode: str = "natural",
    page: int = 1,
    per_page: int = 50,
) -> tuple[list[dict], int, list[dict]]:
    """EG domain: ECG test records."""
    patients = _resolve_patients(run_path, patient_ids)
    rows = []
    for pid in patients:
        for record in _iter_jsonl(run_path, pid, mode):
            eg = record.get("EG")
            if not eg or not eg.get("EGPERF"):
                continue
            rows.append({
                "patient_id": pid,
                "day": record.get("day"),
                "_cycle": _fmt_cycle(record.get("cycle"), record.get("cycle_day")),
                "EGPERF": eg.get("EGPERF"),
                "EGREFID": eg.get("EGREFID"),
                "EGMETHOD": eg.get("EGMETHOD", ""),
                "EGPOS": eg.get("EGPOS", ""),
                "EGDAT": eg.get("EGDAT"),
            })
    page_rows, total = _paginate(rows, page, per_page)
    return page_rows, total, EG_COLUMNS


# ─── Dispatcher ─────────────────────────────────────────────────────────

AGGREGATE_FNS: dict[str, Any] = {
    "DM": aggregate_dm,
    "MH": aggregate_mh,
    "AE": aggregate_ae,
    "EC": aggregate_ec,
    "CM": aggregate_cm,
    "VS": aggregate_vs,
    "LB": aggregate_lb,
    "DS": aggregate_ds,
    "DD": aggregate_dd,
    "TU": aggregate_tu,
    "RS": aggregate_rs,
    "PE": aggregate_pe,
    "EG": aggregate_eg,
}


def aggregate_domain(
    domain: str,
    run_path: Path,
    patient_ids: list[str] | None = None,
    mode: str = "natural",
    source: str = "hr",
    page: int = 1,
    per_page: int = 50,
) -> tuple[list[dict], int, list[dict]]:
    """Dispatch to the appropriate domain aggregation function.

    Returns (rows, total_count, columns).
    """
    fn = AGGREGATE_FNS.get(domain.upper())
    if not fn:
        return [], 0, []

    # Build kwargs based on what the function accepts
    import inspect
    sig = inspect.signature(fn)
    kwargs: dict[str, Any] = {"run_path": run_path, "page": page, "per_page": per_page}
    if "patient_ids" in sig.parameters:
        kwargs["patient_ids"] = patient_ids
    if "mode" in sig.parameters:
        kwargs["mode"] = mode
    if "source" in sig.parameters:
        kwargs["source"] = source
    return fn(**kwargs)


# ─── Excel export helper ────────────────────────────────────────────────

def export_domain_to_excel(
    domain: str,
    run_path: Path,
    patient_ids: list[str] | None = None,
    mode: str = "natural",
    source: str = "hr",
) -> bytes:
    """Export a domain to Excel bytes (openpyxl).

    Returns all rows (no pagination) as .xlsx bytes.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    rows, total, columns = aggregate_domain(
        domain, run_path, patient_ids, mode, source,
        page=1, per_page=0,  # 0 = no limit
    )

    wb = Workbook()
    ws = wb.active
    ws.title = domain.upper()

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    col_keys = [c["key"] for c in columns]
    for ci, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col["label"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Write data
    for ri, row in enumerate(rows, 2):
        for ci, key in enumerate(col_keys, 1):
            val = row.get(key)
            if isinstance(val, bool):
                val = "Yes" if val else "No"
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin_border

    # Auto-width
    for ci, col in enumerate(columns, 1):
        max_len = len(col["label"])
        for ri in range(2, len(rows) + 2):
            val = ws.cell(row=ri, column=ci).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
